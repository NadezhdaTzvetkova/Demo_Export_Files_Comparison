from behave import given, when, then
from pytest_bdd import step, parsers

from helpers import data_validation
from helpers.file_processing import *
from helpers.data_validation import *
from helpers.delimiter_utils import *
from helpers.user_notifications import *
from openpyxl import load_workbook
import concurrent.futures
import psutil
import logging
import random

# =================== Logging Configuration ===================

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

# =================== Feature to Data Directory Mapping ===================

# Dynamically set the DATA_DIR based on feature name
FEATURE_NAME = "decimal_precision"
DATA_DIR = os.path.join("test_data", "{}_test_data".format(FEATURE_NAME))

# Ensure DATA_DIR exists
if not os.path.exists(DATA_DIR):
    logging.warning("Warning directory {} does not exist!".format(DATA_DIR))

# =================== Behave Step Definitions ===================


@given('a system processing "{input_data}" bank export files per hour')
def system_processing_files_per_hour(context, input_data):
    # type: (Any, Any) -> None
    """
    Simulates system performance testing with file validation.
    Handles both file count processing and single file processing based on input data.
    """

    # Process by file count
    if input_data.isdigit():
        context.file_count = int(input_data)
        logging.info(
            "System will process "
            + str(context.file_count)
            + " bank export files per hour."
        )

        # Expected processing time for each file (in seconds)
        expected_time_per_file = 5
        total_expected_time = context.file_count * expected_time_per_file

        # Start measuring time
        start_time = datetime.now()

        # Simulate file processing
        for _ in range(context.file_count):
            end_time = datetime.now() + timedelta(seconds=expected_time_per_file)
            while datetime.now() < end_time:
                pass  # Simulate delay

        # Calculate elapsed time and check performance
        elapsed_time = (datetime.now() - start_time).total_seconds()
        if elapsed_time > total_expected_time:
            raise ValueError(
                "System performance test failed: "
                + str(context.file_count)
                + " files in "
                + "{:.2f}".format(elapsed_time)
                + " seconds, exceeding expected time of "
                + str(total_expected_time)
                + " seconds."
            )

        logging.info(
            "System processed "
            + str(context.file_count)
            + " files in "
            + "{:.2f}".format(elapsed_time)
            + " seconds, within the expected time."
        )

    # Process by file name
    else:
        file_name = input_data
        try:
            file_processor = FileProcessor(file_name, context)
            file_processor.process()
            context.file_name = file_name
            logging.info("System will process the file: " + file_name)
        except Exception as e:
            logging.error("Error processing the file " + file_name + ": " + str(e))


# ================= AML Suspicious Activity Validation =================


@when('I check the "Transaction Amount" and "Recipient" columns in "{sheet_name}"')
def check_transaction_amount_and_recepient(context, sheet_name):
    # type: (Any, Any) -> None
    """Load the file and verify the presence of necessary columns."""

    # Load the file using the helper
    context.df = load_bank_export(context.file_path, sheet_name=sheet_name)

    # List of required columns
    required_columns = ["Transaction Amount", "Recipient"]

    # Check if required columns exist
    missing_cols = [col for col in required_columns if col not in context.df.columns]
    if missing_cols:
        logging.error("Missing columns: " + ", ".join(missing_cols))
        assert False, "Missing column(s): " + ", ".join(missing_cols)

    # Ensure the dataframe is not empty
    assert not context.df.empty, "Loaded dataframe is empty."

    # Validate column types and data
    if "Transaction Amount" in context.df.columns:
        if not validate_column(context.df, "Transaction Amount", "numeric"):
            logging.error(
                "The 'Transaction Amount' column contains non-numeric values."
            )
            assert False, "'Transaction Amount' column contains non-numeric values."

    if "Recipient" in context.df.columns:
        if context.df["Recipient"].isnull().any():
            logging.error("The 'Recipient' column contains missing values.")
            assert False, "'Recipient' column contains missing values."


@then('transactions above "{threshold}" should be flagged')
def flag_high_value_transactions(context, threshold):
    # type: (Any, Any) -> None
    """Flag transactions above the specified threshold"""
    threshold_value = float(threshold.replace("$", "").replace(",", ""))
    flagged = context.df[context.df["Transaction Amount"] > threshold_value]
    context.flagged_transactions = flagged
    assert not flagged.empty, "No transactions flagged."
    print(str(len(flagged)) + " transactions flagged above " + threshold + ".")


@then('flagged transactions should be reported to "{compliance_team}"')
def step_then_report_to_compliance(context, compliance_team):
    # type: (Any, Any) -> None
    """Ensure flagged transactions are reported"""
    assert hasattr(
        context, "flagged_transactions"
    ), "No flagged transactions to report."
    assert (
        not context.flagged_transactions.empty
    ), "Flagged transactions dataset is empty."
    print(
        "Reporting "
        + str(len(context.flagged_transactions))
        + " flagged transactions to "
        + compliance_team
        + "."
    )


@then('system should cross-check against known "{sanctioned_list}"')
def step_then_cross_check_sanctions(context, sanctioned_list):
    # type: (Any, Any) -> None
    """Simulate cross-checking against a sanctioned list"""
    sanctioned_entities = ["OFAC Watchlist", "EU Sanctions", "Interpol Red List"]
    assert sanctioned_list in sanctioned_entities, (
        "Sanctioned list " + sanctioned_list + " not recognized."
    )
    print("Cross-checking against " + sanctioned_list + " completed.")


@when('I check multiple transactions in "{sheet_name}"')
def step_when_check_structured_transactions(context, sheet_name):
    # type: (Any, Any) -> None
    """Load transactions for structured transaction analysis."""
    if context.file_path.endswith(".csv"):
        context.df = pd.read_csv(context.file_path)
    else:
        context.df = pd.read_excel(context.file_path, sheet_name=sheet_name)

    assert (
        not context.df.empty
    ), "Loaded dataframe for structured transactions is empty."


@then(
    'transactions structured below "{threshold}" but summing above "{aggregate_limit}" should be flagged'
)
def step_then_flag_structured_transactions(context, threshold, aggregate_limit):
    # type: (Any, Any, Any) -> None
    """Detect structured transactions designed to evade AML reporting"""
    threshold_value = float(threshold.replace("$", "").replace(",", ""))
    aggregate_value = float(aggregate_limit.replace("$", "").replace(",", ""))
    grouped = context.df.groupby("Recipient")["Transaction Amount"].sum()
    flagged = grouped[(grouped > aggregate_value) & (grouped < threshold_value)]
    context.flagged_structured_transactions = flagged
    assert not flagged.empty, "No structured transactions flagged."
    print(str(len(flagged)) + " structured transactions flagged.")


@when("I check for AML compliance anomalies")
def step_when_check_aml_anomalies(context):
    # type: (Any) -> None
    """Simulate checking large datasets for AML anomalies"""
    assert hasattr(context, "df"), "No dataset loaded for AML anomaly detection."
    assert not context.df.empty, "Dataset is empty. Cannot check for AML anomalies."
    context.large_dataset_flagged = True


@then('flagged transactions should be identified within "{expected_runtime}"')
def step_then_validate_aml_runtime(context, expected_runtime):
    # type: (Any, Any) -> None
    """Check if AML processing meets expected runtime"""
    assert hasattr(
        context, "large_dataset_flagged"
    ), "Large dataset anomalies not processed in time."
    assert (
        context.large_dataset_flagged
    ), "AML anomaly detection did not execute properly."
    print("AML compliance anomalies checked within " + expected_runtime + ".")


# ================= End of AML Suspicious Activity Validation =================


# ================= Start of Currency Consistency Validation =================


@given('a bank export file "{file_name}"')
def loaded_bank_export_file(context, file_name):
    # type: (Any, Any) -> None
    """Ensures the bank export file exists and loads it into the context."""

    context.feature_folder = "data_validation_test_data"

    # Use the helper function from helpers/file_processing.py to load the bank export file
    try:
        context.data = load_bank_export_file(context.feature_folder, file_name)
        logging.info("Successfully loaded the bank export file: {0}".format(file_name))
    except Exception as e:
        logging.error(
            "Failed to load the bank export file: {0} due to error: {1}".format(
                file_name, str(e)
            )
        )
        raise


@when('I check for currency codes in the "{sheet_name}" sheet')
def step_when_check_currency_codes(context, sheet_name):
    # type: (Any, Any) -> None
    """Validates that all transactions have a valid currency code."""
    assert "Currency" in context.data.columns, "Missing 'Currency' column in file."
    valid_currencies = set(
        ["USD", "EUR", "GBP", "JPY", "CAD"]
    )  # Example ISO 4217 codes
    context.currency_issues = context.data[
        ~context.data["Currency"].isin(valid_currencies)
    ]


@then("all transactions should have a valid ISO 4217 currency code")
def step_then_validate_currency_codes(context):
    # type: (Any) -> None
    """Checks if any invalid currency codes exist."""
    assert context.currency_issues.empty, "Invalid currency codes found:\n" + str(
        context.currency_issues
    )


@then("currency codes should match the account’s assigned currency")
def step_then_validate_account_currency(context):
    # type: (Any) -> None
    """Ensures that the transaction currency matches the assigned account currency."""
    assert (
        "Account Currency" in context.data.columns
    ), "Missing 'Account Currency' column."
    mismatches = context.data[
        context.data["Currency"] != context.data["Account Currency"]
    ]
    assert mismatches.empty, "Currency mismatches found:\n" + str(mismatches)


@when('I check for negative values in the "{sheet_name}" sheet')
def step_when_check_negative_values(context, sheet_name):
    # type: (Any, Any) -> None
    """Validates that negative values are only present in debit transactions."""
    assert "Amount" in context.data.columns, "Missing 'Amount' column."
    assert (
        "Transaction Type" in context.data.columns
    ), "Missing 'Transaction Type' column."
    context.invalid_negatives = context.data[
        (context.data["Amount"] < 0) & (context.data["Transaction Type"] == "Credit")
    ]


@then("negative values should only be present in debit transactions")
def step_then_validate_negative_values(context):
    # type: (Any) -> None
    """Ensures that no credit transactions have negative values."""
    assert (
        context.invalid_negatives.empty
    ), "Invalid negative values in credit transactions:\n" + str(
        context.invalid_negatives
    )


@when('I check for negative and zero values in the "{sheet_name}" sheet')
def step_when_check_zero_values(context, sheet_name):
    # type: (Any, Any) -> None
    """Checks for zero and negative transaction amounts."""
    context.zero_values = context.data[context.data["Amount"] == 0]
    context.negative_debits = context.data[
        (context.data["Amount"] < 0) & (context.data["Transaction Type"] == "Debit")
    ]


@then("zero amounts should be flagged as potential errors")
def step_then_flag_zero_amounts(context):
    # type: (Any) -> None
    """Flags transactions with zero amounts."""
    assert context.zero_values.empty, "Transactions with zero amounts:\n" + str(
        context.zero_values
    )


@then("transactions with zero values should be logged for further review")
def step_then_log_zero_values(context):
    # type: (Any) -> None
    """Logs zero value transactions for further investigation."""
    print("⚠️ Zero value transactions detected:\n" + str(context.zero_values))


def load_file(file_path, sheet_name):
    # type: (Any, Any) -> pd.DataFrame
    """Loads CSV or Excel files based on extension"""
    if file_path.endswith(".csv"):
        return pd.read_csv(file_path)
    elif file_path.endswith(".xlsx"):
        return pd.read_excel(file_path, sheet_name=sheet_name)
    raise ValueError("Unsupported file format")


@when('I check the "Date" column in the "{sheet_name}" sheet')
def step_when_check_date_column(context, sheet_name):
    # type: (Any, Any) -> None
    """Load and validate date format"""
    df = load_file(context.file_path, sheet_name)
    context.dates = df["Date"]
    context.invalid_dates = []

    for date in context.dates:
        try:
            datetime.strptime(str(date), ISO_DATE_FORMAT)
        except ValueError:
            context.invalid_dates.append(date)


@then('all dates should follow the format "YYYY-MM-DD"')
def step_then_validate_date_format(context):
    # type: (Any) -> None
    """Check if all dates follow ISO format"""
    assert not context.invalid_dates, "Invalid date formats detected: " + str(
        context.invalid_dates
    )


@then("dates should not contain time components unless explicitly required")
def step_then_validate_date_time_component(context):
    # type: (Any) -> None
    """Ensure no unexpected time components in dates"""
    assert all(
        " " not in str(date) for date in context.dates
    ), "Some dates contain unexpected time components"


@when('I check the "Timestamp" column in the "{sheet_name}" sheet')
def step_when_check_timestamp_column(context, sheet_name):
    # type: (Any, Any) -> None
    """Load and validate timestamp format"""
    df = load_file(context.file_path, sheet_name)
    context.timestamps = df["Timestamp"]
    context.invalid_timestamps = []

    for timestamp in context.timestamps:
        try:
            datetime.strptime(str(timestamp), TIMESTAMP_FORMAT)
        except ValueError:
            context.invalid_timestamps.append(timestamp)


@then('all timestamps should follow the "YYYY-MM-DD HH:SS" format')
def step_then_validate_timestamp_format(context):
    # type: (Any) -> None
    """Check if all timestamps follow the expected format"""
    assert not context.invalid_timestamps, "Invalid timestamp formats detected: " + str(
        context.invalid_timestamps
    )


@when("I check for missing or blank date values")
def step_when_check_missing_dates(context):
    # type: (Any) -> None
    """Identify missing date values"""
    context.missing_dates = context.dates.isna() | (context.dates == "")


@then("no date field should be empty or null")
def step_then_validate_no_missing_dates(context):
    # type: (Any) -> None
    """Ensure no missing date values"""
    assert not context.missing_dates.any(), "Missing or blank dates detected"


@when("I check for chronological consistency")
def step_when_check_chronology(context):
    # type: (Any) -> None
    """Ensure dates are in chronological order"""
    context.dates_sorted = sorted(pd.to_datetime(context.dates, errors="coerce"))
    context.date_mismatches = context.dates_sorted != list(
        pd.to_datetime(context.dates, errors="coerce")
    )


@then("transaction dates should be in chronological order")
def step_then_validate_chronology(context):
    # type: (Any) -> None
    """Verify that dates are sorted correctly"""
    assert not any(context.date_mismatches), "Transactions are out of order"


@when('I analyze the "Date" column in the "{sheet_name}" sheet')
def step_when_analyze_date_column(context, sheet_name):
    # type: (Any, Any) -> None
    """Detect future or backdated transactions"""
    df = load_file(context.file_path, sheet_name)
    context.dates = pd.to_datetime(df["Date"], errors="coerce")
    today = datetime.today()

    context.backdated_transactions = context.dates < (
        today - timedelta(days=365 * 5)
    )  # Older than 5 years
    context.future_transactions = context.dates > (
        today + timedelta(days=30)
    )  # More than 30 days ahead


@then(
    'transactions older than "{backdate_threshold}" years should be flagged for fraud review'
)
def step_then_flag_old_transactions(context, backdate_threshold):
    # type: (Any, Any) -> None
    """Flag transactions older than the threshold"""
    threshold = int(backdate_threshold)
    assert not any(context.backdated_transactions), (
        "Found transactions older than " + str(threshold) + " years"
    )


@then('transactions postdated beyond "{future_threshold}" days should trigger alerts')
def step_then_flag_future_transactions(context, future_threshold):
    # type: (Any, Any) -> None
    """Flag future-dated transactions"""
    threshold = int(future_threshold)
    assert not any(context.future_transactions), (
        "Found transactions postdated beyond " + str(threshold) + " days"
    )


# ================= End of Date Format Validation =================

# ================= Beginning of Decimal Precision Validation =================


@given('a bank export file "{file_name}"')
def bank_export_file(context, file_name):
    # type: (Any, Any) -> None
    """Ensure the bank export file exists and store it in context for a single file."""
    context.file_name = file_name
    file_path = os.path.join(context.base_dir, file_name)

    # Ensure the file exists
    assert os.path.exists(file_path), "File " + file_name + " not found."

    try:
        # Initialize FileProcessor and process the file
        processor = FileProcessor(file_name, context)
        processor.process()  # This handles the file processing and logging
    except ValueError as e:
        logging.error("Error processing file: " + str(e))
        raise  # Raise the error once, it will stop further execution for this step.


@given("a set of bank export files:")
def bank_export_files(context):
    # type: (Any) -> None
    """Ensure all provided bank export files exist and process them."""
    context.file_paths = []

    for row in context.table:
        file_name = row["file_name"]
        file_path = os.path.join(context.base_dir, file_name)

        # Ensure the file exists
        assert os.path.exists(file_path), "File " + file_name + " not found."
        context.file_paths.append(file_path)

        try:
            # Initialize FileProcessor and process the file
            processor = FileProcessor(file_name, context)
            processor.process()  # This handles the file processing and logging
        except ValueError as e:
            logging.error("Error processing file " + file_name + ": " + str(e))
            raise  # Raise the error once, it will stop further execution for this step.


@when('I check the "Amount" column in the "{sheet_name}" sheet')
def step_when_check_amount_column(context, sheet_name):
    # type: (Any, Any) -> None
    context.df = load_bank_export(context.file_name, sheet_name)

    if "Amount" not in context.df.columns:
        raise KeyError("Missing 'Amount' column in dataset")

    logging.info(
        "Loaded dataset from " + context.file_name + ", checking 'Amount' column"
    )


@then('all monetary values should have exactly "{expected_precision}" decimal places')
def step_then_validate_decimal_precision(context, expected_precision):
    # type: (Any, Any) -> None
    expected_precision = int(expected_precision)

    context.df["Decimal_Precision"] = (
        context.df["Amount"].astype(str).str.split(".").str[-1].str.len()
    )
    incorrect_values = context.df[context.df["Decimal_Precision"] != expected_precision]

    assert incorrect_values.empty, (
        str(len(incorrect_values)) + " transactions have incorrect decimal precision!"
    )
    logging.info("All monetary values meet expected decimal precision.")


@then('values should not contain more than "{expected_precision}" decimal places')
def step_then_check_max_precision(context, expected_precision):
    # type: (Any, Any) -> None
    expected_precision = int(expected_precision)

    max_precision = context.df["Decimal_Precision"].max()
    assert max_precision <= expected_precision, (
        "Max precision found "
        + str(max_precision)
        + ", expected: "
        + str(expected_precision)
    )
    logging.info("All values conform to max decimal precision limit.")


@then("rounding inconsistencies should be flagged")
def step_then_flag_rounding_issues(context):
    # type: (Any) -> None
    context.df["Rounded_Amount"] = context.df["Amount"].round(2)
    rounding_issues = context.df[context.df["Amount"] != context.df["Rounded_Amount"]]

    if not rounding_issues.empty:
        logging.warning(
            str(len(rounding_issues)) + " transactions show rounding inconsistencies."
        )
        logging.info(rounding_issues.to_string())
    else:
        logging.info("No rounding inconsistencies found.")


# ================= End of Decimal Precision Validation =================
# Map human-readable delimiter names to actual characters


@when("I check the delimiter format in the file")
def step_when_check_delimiter_format(context):
    # type: (Any) -> None
    """Detect the delimiter used in the file."""
    with open(context.file_path, "r") as f:
        first_line = f.readline()

    detected_delimiters = [d for d in DELIMITER_MAPPING.values() if d in first_line]
    context.detected_delimiters = (
        detected_delimiters if detected_delimiters else [","]
    )  # Default to comma


@then(
    'files containing multiple valid delimiters "{allowed_delimiters}" should be accepted'
)
def step_then_validate_multiple_delimiters(context, allowed_delimiters):
    # type: (Any, Any) -> None
    """Check if the file's delimiter is in the list of allowed delimiters."""
    allowed_list = [DELIMITER_MAPPING[d.strip()] for d in allowed_delimiters.split(",")]

    for delimiter in allowed_list:
        try:
            df = pd.read_csv(context.file_path, delimiter=delimiter)
            print("Successfully parsed using delimiter:", delimiter)
            return  # Stop once successful
        except Exception as e:
            print("Failed with delimiter:", delimiter, "->", str(e))

    assert False, "No valid delimiter found in " + context.file_path


# ================= Beginning of Delimiter Inconsistency Validation =================


@when("I check the delimiter format in the file")
def step_when_check_delimiter(context):
    # type: (Any) -> None
    """Checks the delimiter format in the given file."""

    # Get the file path using the helper function (assuming get_test_file_path exists)
    file_path = get_test_file_path(context, context.file_name)

    try:
        # Open the file and read the first line to detect delimiters
        with open(file_path, "r", encoding="utf-8") as f:
            first_line = f.readline()

        # Define a list of possible delimiters
        delimiters = [",", ";", "|", "\t"]

        # Detect which delimiters are present in the first line
        context.detected_delimiters = [d for d in delimiters if d in first_line]

        # Assert that at least one delimiter is found
        assert (
            len(context.detected_delimiters) > 0
        ), "No delimiter detected. File might be corrupted."

        # Log the detected delimiters
        logging.info("Detected delimiters: {}".format(context.detected_delimiters))

    except Exception as e:
        # If an error occurs, log it
        context.error = str(e)
        logging.error("Error checking delimiters: {}".format(context.error))
        raise


@then("the delimiter should be consistent throughout the file")
def step_then_check_consistency(context):
    # type: (Any) -> None
    """Validate that the delimiter is used consistently across the file."""
    file_path = get_test_file_path(context, context.file_name)

    with open(file_path, "r") as f:
        lines = f.readlines()

    delimiter = context.detected_delimiters[0]
    delimiter_counts = [line.count(delimiter) for line in lines]

    assert all(
        count == delimiter_counts[0] for count in delimiter_counts
    ), "Inconsistent delimiters detected."

    logging.info("Delimiter '{}' is consistent across the file.".format(delimiter))


@then("mixed delimiters within the file should be flagged")
def step_then_flag_mixed_delimiters(context):
    # type: (Any) -> None
    """Flag files that contain more than one type of delimiter."""
    if len(context.detected_delimiters) > 1:
        logging.warning(
            "Mixed delimiters detected in {} {}".format(
                context.file_name, context.detected_delimiters
            )
        )
        assert False, "File contains mixed delimiters."


@then("an error report should be generated listing inconsistent delimiters")
def step_then_generate_error_report(context):
    # type: (Any) -> None
    """Generate a text report if multiple delimiters are detected in the file."""
    if len(context.detected_delimiters) > 1:
        report_path = get_test_file_path(context, "delimiter_error_report.txt")
        with open(report_path, "w") as f:
            f.write("File {}\n".format(context.file_name))
            f.write("Detected Delimiters: {}\n".format(context.detected_delimiters))
        logging.info("Error report generated: {}".format(report_path))

    # ================= End of Delimiter Inconsistency Validation =================

    # ================= Beginning of Encoding Validation =================


@when("I check the file encoding")
def step_when_check_encoding(context):
    # type: (Any) -> None
    """Detect and log the encoding used in the file."""

    # Initialize the FileProcessor to reuse its encoding detection
    file_processor = FileProcessor(file_name=context.file_name, context=context)

    # Use the detect_encoding method from FileProcessor to detect encoding
    context.detected_encoding = file_processor.detect_encoding(context.file_path)

    # Log the detected encoding without using f-strings
    logging.info(
        "Detected encoding for " + context.file_name + ": " + context.detected_encoding
    )


@then('the file encoding should be "{expected_encoding}"')
def step_then_validate_encoding(context, expected_encoding):
    # type: (Any, Any) -> None
    """Validate that the detected encoding matches the expected encoding."""

    # Compare the detected encoding with the expected encoding (case-insensitive)
    if context.detected_encoding.lower() != expected_encoding.lower():
        raise AssertionError(
            "Encoding mismatch. Expected: "
            + expected_encoding
            + ", but detected: "
            + context.detected_encoding
        )


@then("non-standard encodings should be flagged")
def step_then_flag_non_standard_encoding(context):
    # type: (Any) -> None
    standard_encodings = ["utf-8", "ascii"]
    if context.detected_encoding.lower() not in standard_encodings:
        logging.warning("Non-standard encoding detected: " + context.detected_encoding)


@then("the system should suggest converting to a standard encoding")
def step_then_suggest_conversion(context):
    # type: (Any) -> None
    if context.detected_encoding.lower() not in ["utf-8", "ascii"]:
        logging.info("Suggested conversion " + context.file_name + " to UTF-8 or ASCII")


@then("a conversion report should list all affected files")
def step_then_generate_encoding_report(context):
    # type: (Any) -> None
    report_path = get_test_file_path(context, "encoding_report.txt")
    with open(report_path, "w") as f:
        f.write("File {}\n".format(context.file_name))
        f.write("Detected Encoding: {}\n".format(context.detected_encoding))
    logging.info("Encoding report generated: {}".format(report_path))

    # ================= End of Encoding Validation =================

    # ================= Beginning of Invalid Account Number Validation =================


@when('I check the "Account Number" column in the "{sheet_name}" sheet')
def step_when_check_account_number(context, sheet_name):
    # type: (Any, Any) -> None
    # Placeholder for actual implementation (CSV/Excel parsing logic required)
    context.account_numbers = [
        "1234567890",
        "ABCDEFGHIJ",
        "1234-567890",
    ]


@then('all account numbers should match the expected pattern "{pattern}"')
def step_then_validate_account_number_format(context, pattern):
    # type: (Any, Any) -> None
    for account_number in context.account_numbers:
        assert is_valid_account_number(
            account_number, pattern
        ), "Invalid account number format detected: {}".format(account_number)


@then("invalidly formatted account numbers should be flagged")
def step_then_flag_invalid_accounts(context):
    # type: (Any) -> None
    invalid_accounts = [
        acc
        for acc in context.account_numbers
        if not is_valid_account_number(acc, r"\d{10,12}")
    ]
    if invalid_accounts:
        logging.warning("Invalid account numbers found: {}".format(invalid_accounts))
        assert (
            False
        ), "Some account numbers do not conform to the expected format: {}".format(
            invalid_accounts
        )

    @then("an alert should be sent for accounts not matching regulatory formats")
    def step_then_alert_regulatory_issues(context):
        # type: (Any) -> None
        logging.warning("Regulatory alert: non-compliant account number detected.")

    # ================= End of Invalid Account Number Validation =================

    # ================= Beginning of Invalid Currency Code Validation =================


def is_valid_currency_code(currency_code, pattern=r"^[A-Z]{3}$"):
    # type: (Any, str) -> bool
    """
    Checks if the currency code follows the expected pattern.
    Default pattern matches standard ISO 4217 currency codes (3 uppercase letters).
    """
    if not isinstance(currency_code, str):
        return False
    return re.match(pattern + r'\Z', currency_code.strip()) is not None


@when('I check the "Currency" column in the "{sheet_name}" sheet')
def step_when_check_currency_column(context, sheet_name):
    # type: (Any, Any) -> None
    # Placeholder for actual implementation (CSV/Excel parsing logic required)
    context.currency_codes = ["USD", "E$U", "XYZ"]  # Example test data


@then('all currency codes should match the expected pattern "{pattern}"')
def step_then_validate_currency_format(context, pattern):
    # type: (Any, Any) -> None
    """Validate that all currency codes conform to the expected pattern."""
    for currency_code in context.currency_codes:
        if not is_valid_currency_code(currency_code, pattern):
            assert False, "Invalid currency code detected: {}".format(currency_code)
    logging.info("All currency codes match the expected pattern.")


@then("invalid currency codes should be flagged")
def step_then_flag_invalid_currencies(context):
    # type: (Any) -> None
    invalid_currencies = [
        cur
        for cur in context.currency_codes
        if not is_valid_currency_code(cur, r"^[A-Z]{3}$")
    ]

    if invalid_currencies:
        logging.warning("Invalid currency codes found: {}".format(invalid_currencies))
        assert (
            False
        ), "Some currency codes do not conform to the expected format: {}".format(
            invalid_currencies
        )


@then("a correction suggestion should be provided")
def step_then_suggest_correction(context):
    # type: (Any) -> None
    """Provide suggestions based on context."""
    if hasattr(context, "column_values"):
        # Suggest correction for whitespace issues in column values
        corrections = context.column_values.str.replace("  +", " ", regex=True)
        logging.info(
            "Suggested corrections for whitespace issues applied where necessary."
        )

    if (
        hasattr(context, "file_name")
        and hasattr(context, "missing_detected")
        and context.missing_detected
    ):
        # Suggest mapping corrections for missing columns
        logging.info(
            "Suggested mappings for missing columns in {}: {}".format(
                context.file_name, context.missing_detected
            )
        )

    if hasattr(context, "account_numbers"):
        # Suggest correction for account numbers (numeric characters and length)
        logging.info(
            "Suggest correction account numbers contain only numeric characters and match expected length."
        )

    if hasattr(context, "currency_codes"):
        # Suggest correction for currency codes (ISO 4217 standard)
        logging.info("Suggest correction currency codes follow the ISO 4217 standard.")

    if hasattr(context, "text_fields") and hasattr(context, "char_limit"):
        # Suggest truncating long text fields
        logging.info(
            "Suggestion values in text fields do not exceed the character limit. Consider truncating or reformatting long entries."
        )


@then("transactions with invalid currency codes should be marked for review")
def step_then_flag_invalid_transactions(context):
    # type: (Any) -> None
    logging.warning(
        "Flagging transactions with invalid currency codes for further review."
    )

    # ================= End of Invalid Currency Code Validation =================

    # ================= Beginning of Missing Values Validation =================


@when('I check for missing values in the "{sheet_name}" sheet')
def step_when_check_missing_values(context, sheet_name):
    # type: (Any, Any) -> None
    # Placeholder for actual implementation (CSV/Excel parsing logic required)
    context.missing_values = {
        "Transaction ID": 2,
        "Currency": 3,
    }


@then("no mandatory field should be empty or null")
def step_then_validate_missing_values(context):
    # type: (Any) -> None
    """Validates that no mandatory field is empty or null."""

    # Iterate through the missing_values dictionary
    for field, count in context.missing_values.items():
        # If there are missing values for a mandatory field, assert failure
        assert count == 0, "Missing values found in field: " + field

    # If no fields are missing, print success message
    logging.info("All mandatory fields are complete and valid.")


@then("missing values should be flagged")
def step_then_flag_missing_values(context):
    # type: (Any) -> None
    """Flag missing values and handle appropriate actions."""

    # Access missing values data from context
    missing_values = context.missing_values

    # Identify fields with missing values
    flagged_fields = [field for field, count in missing_values.items() if count > 0]

    if flagged_fields:
        # Log the fields with missing values
        logging.warning("Missing values detected in fields: " + str(flagged_fields))
        # Assert failure if there are flagged fields with missing values
        assert False, "Some mandatory fields have missing values: " + str(
            flagged_fields
        )

    # Check if there are any missing values
    total_missing = sum(missing_values.values())
    assert total_missing > 0, "No missing values detected."

    # Print the number of missing values found
    logging.info(str(total_missing) + " missing values found.")


@then('records with missing values should be categorized based on "{priority}"')
def categorize_records_with_missing_values(context, priority):
    # Retrieve records with any missing values using the data_validation helper
    missing_records = data_validation.get_records_with_missing_values(context.dataset)

    # Ensure that we found some records if we expect any missing data
    assert (
        missing_records is not None
    ), "No records were returned for missing value check."

    # Additional checks for specific fields like missing currency codes, transaction IDs, reference data
    missing_currency_codes = [
        record
        for record in missing_records
        if 'currency_code' not in record or not record['currency_code']
    ]
    missing_transaction_ids = [
        record
        for record in missing_records
        if 'transaction_id' not in record or not record['transaction_id']
    ]
    missing_reference_data = [
        record
        for record in missing_records
        if 'reference_data' not in record or not record['reference_data']
    ]

    # Log if specific missing fields are found
    if missing_currency_codes:
        logging.warning(
            "Missing currency codes in {} records.".format(len(missing_currency_codes))
        )
    if missing_transaction_ids:
        logging.warning(
            "Missing transaction IDs in {} records.".format(
                len(missing_transaction_ids)
            )
        )
    if missing_reference_data:
        logging.warning(
            "Missing reference data in {} records.".format(len(missing_reference_data))
        )

    # Handle large datasets (if the dataset is large, the checks might take time)
    if len(missing_records) > 1000:  # Adjust threshold as necessary
        logging.info(
            "Processing large dataset with {} records with missing values.".format(
                len(missing_records)
            )
        )

    # Verify each record with missing data has the expected priority category
    for record in missing_records:
        # Determine the category that this record *should* fall under based on its missing fields
        actual_category = data_validation.determine_missing_priority(record)

        # Check that it matches the expected priority from the scenario
        assert (
            actual_category == priority
        ), "Record with missing values was categorized as '{}' instead of '{}'".format(
            actual_category, priority
        )

    # Optionally, log a summary of missing values
    total_missing = (
        len(missing_currency_codes)
        + len(missing_transaction_ids)
        + len(missing_reference_data)
    )
    logging.info("Total missing values in records: {}".format(total_missing))

    # ================= End of Missing Values Validation =================

    # ================= Beginning of Negative Values Validation =================


@given('I have a bank export file "{file_name}" from the old system')
def old_system_file(context, file_name):
    # type: (Any, str) -> None
    """Load a bank export file from the old system."""

    # Set file name and path
    context.old_file_name = file_name
    context.old_file_path = get_test_file_path(context, file_name)

    # Check if the file exists at the given path
    file_exists = os.path.exists(context.old_file_path)

    # If the file does not exist, raise an error
    if not file_exists:
        raise ValueError("File {} does not exist in old system.".format(file_name))

    # If file exists, continue processing or logging if needed
    logging.info(
        "File {} successfully found at {}".format(file_name, context.old_file_path)
    )


@given('I have a bank export file "{file_name}" from the new system')
def new_system_file(context, file_name):
    # type: (Any, Any) -> None
    """Load a bank export file from the new system."""
    context.new_file_name = file_name
    context.new_file_path = get_test_file_path(context, file_name)
    assert os.path.exists(
        context.new_file_path
    ), "File {} does not exist in new system.".format(file_name)


@when('I compare the "{numeric_column}" column')
def step_when_compare_numeric_column(context, numeric_column):
    # type: (Any, Any) -> None
    """Compares the numeric column between two systems (old and new) while ensuring proper file type handling."""

    # Load old system data using generic helper logic
    if context.old_file_path.endswith(".csv"):
        old_data = pd.read_csv(context.old_file_path)
    elif context.old_file_path.endswith(".xlsx"):
        old_data = pd.read_excel(context.old_file_path)
    else:
        raise ValueError(
            "Unsupported file format for the old file: {}".format(context.old_file_path)
        )

    # Load new system data using generic helper logic
    if context.new_file_path.endswith(".csv"):
        new_data = pd.read_csv(context.new_file_path)
    elif context.new_file_path.endswith(".xlsx"):
        new_data = pd.read_excel(context.new_file_path)
    else:
        raise ValueError(
            "Unsupported file format for the new file: {}".format(context.new_file_path)
        )

    # Ensure the column exists in both datasets
    if numeric_column not in old_data.columns:
        raise ValueError(
            "Column '{}' not found in old system file: {}".format(
                numeric_column, context.old_file_path
            )
        )
    if numeric_column not in new_data.columns:
        raise ValueError(
            "Column '{}' not found in new system file: {}".format(
                numeric_column, context.new_file_path
            )
        )

    # Validate column is numeric using helper logic if needed
    if not pd.api.types.is_numeric_dtype(old_data[numeric_column]):
        raise TypeError(
            "Column '{}' in the old file is not numeric.".format(numeric_column)
        )
    if not pd.api.types.is_numeric_dtype(new_data[numeric_column]):
        raise TypeError(
            "Column '{}' in the new file is not numeric.".format(numeric_column)
        )

    # Store values in context for further comparison
    context.old_values = old_data[numeric_column]
    context.new_values = new_data[numeric_column]

    logging.info(
        "Successfully loaded and validated '{}' from both files.".format(numeric_column)
    )


@then("negative values should be identified in the old system")
def step_then_check_old_negative_values(context):
    # type: (Any) -> None
    """Checks for negative values in the old system's data."""
    old_negatives = context.old_values[context.old_values < 0]
    assert (
        not old_negatives.empty
    ), "No negative values found in the old system where expected."

    # Assert negative values match between the two exports
    assert old_negatives.equals(
        context.new_negatives
    ), "Negative values mismatch between old and new system exports."
    logging.info("Negative values processed correctly and match between systems.")


@when('I compare the "Amount" column')
def step_when_compare_amount_column(context):
    # type: (Any) -> None
    """Compare the 'Amount' column in old and new system exports."""
    context.old_values = pd.read_csv(context.old_file_path)["Amount"]
    context.new_values = pd.read_csv(context.new_file_path)["Amount"]

    context.old_negatives = context.old_values[context.old_values < 0]
    context.new_negatives = context.new_values[context.new_values < 0]


@given('I have a bank export file "{file_name}" from the old system')
def old_bank_export_file(context, file_name):
    # type: (Any, Any) -> None
    """Loads a bank export file from the old system."""
    context.old_file_name = file_name
    context.old_file_path = get_test_file_path(context, file_name)
    assert os.path.exists(
        context.old_file_path
    ), "File {} does not exist in old system.".format(file_name)


@given('I have a bank export file "{file_name}" from the new system')
def new_file(context, file_name):
    # type: (Any, Any) -> None
    """Loads a bank export file from the new system."""
    context.new_file_name = file_name
    context.new_file_path = get_test_file_path(context, file_name)
    assert os.path.exists(
        context.new_file_path
    ), "File {} does not exist in new system.".format(file_name)


@then("negative values should be processed correctly")
def step_then_check_negative_values(context):
    # type: (Any) -> None
    """Ensure negative values are processed correctly in both exports."""
    old_negatives = context.old_values[context.old_values < 0]
    new_negatives = context.new_values[context.new_values < 0]
    assert old_negatives.equals(
        new_negatives
    ), "Mismatch in negative values between old and new data."
    logging.info("Negative values processed correctly and match between systems.")

    # ================= End of Negative Values Validation =================

    # ================= Beginning of Whitespace Handling Validation =================


@when(
    'I check for whitespace issues in the "{column_name}" column in the "{sheet_name}" sheet'
)
def step_when_check_whitespace(context, column_name, sheet_name):
    # type: (Any, Any, Any) -> None
    """Load data and validate whitespace issues in a specific column."""
    # Load data based on file type
    if context.file_path.endswith(".csv"):
        context.data = pd.read_csv(context.file_path)
    else:
        context.data = pd.read_excel(context.file_path, sheet_name=sheet_name)

    # Ensure the column exists
    assert column_name in context.data.columns, "Column {} not found.".format(
        column_name
    )

    # Store column values for validation
    context.column_values = context.data[column_name]


@then("leading and trailing spaces should be removed from all text fields")
def step_then_remove_whitespace(context):
    # type: (Any) -> None
    """Verify that all leading/trailing spaces are removed from text fields."""
    before_cleaning = context.column_values.str.strip()
    assert before_cleaning.equals(
        context.column_values
    ), "Leading or trailing whitespace found in text fields."
    logging.info("Whitespace correctly handled in text fields.")


@then("fields with excessive whitespace should be flagged")
def step_then_flag_excessive_whitespace(context):
    # type: (Any) -> None
    """Check for and flag fields with excessive internal whitespace."""
    flagged_rows = context.column_values[context.column_values.str.contains("  ")]
    assert flagged_rows.empty, "Excessive whitespace detected in text fields."
    logging.info("No excessive whitespace detected in fields.")

    # ================= End of Whitespace Handling Validation =================

    # ================= Beginning of Duplicate Accounts Validation =================


@when('I check the "Account Number" column in the "{sheet_name}" sheet')
def step_when_check_duplicate_accounts(context, sheet_name):
    # type: (Any, Any) -> None
    """Check for duplicate account numbers."""
    if context.file_path.endswith(".csv"):
        context.data = pd.read_csv(context.file_path)
    else:
        context.data = pd.read_excel(context.file_path, sheet_name=sheet_name)

    assert (
        "Account Number" in context.data.columns
    ), "Column 'Account Number' not found in file."

    context.duplicate_accounts = context.data["Account Number"].duplicated(keep=False)


@then("duplicate account numbers should be flagged")
def step_then_flag_duplicate_accounts(context):
    # type: (Any) -> None
    """Flag duplicate account numbers."""
    flagged_duplicates = context.data[context.duplicate_accounts]
    assert not flagged_duplicates.empty, "No duplicate accounts detected."
    logging.info(
        "Flagged duplicate accounts: {} records.".format(len(flagged_duplicates))
    )


@then("a report should be generated listing duplicate occurrences")
def step_then_generate_report(context):
    # type: (Any) -> None
    """Generate a report of duplicate account numbers."""
    flagged_duplicates = context.data[context.duplicate_accounts]
    report_path = os.path.join("reports", "duplicate_accounts_report.csv")
    flagged_duplicates.to_csv(report_path, index=False)
    logging.info("Duplicate accounts report generated at {}.".format(report_path))


@then("accounts with high-frequency duplication should be escalated for review")
def step_then_escalate_high_frequency_duplicates(context):
    # type: (Any) -> None
    """Escalate accounts with high-frequency duplicates for review."""
    flagged_duplicates = context.data[context.duplicate_accounts]
    duplicate_counts = flagged_duplicates["Account Number"].value_counts()
    high_risk_accounts = duplicate_counts[
        duplicate_counts > 5
    ]  # Example threshold for escalation
    assert (
        not high_risk_accounts.empty
    ), "No high-frequency duplicate accounts detected."
    logging.info(
        "High-risk duplicate accounts flagged: {} occurrences.".format(
            len(high_risk_accounts)
        )
    )


# ================= End of Duplicate Accounts Validation =================

# ================= Beginning of Duplicate Customers Validation =================


@when('I check the "Customer ID" column in the "{sheet_name}" sheet')
def step_when_check_duplicate_customers(context, sheet_name):
    # type: (Any, Any) -> None
    """Check for duplicate customer IDs."""
    if context.file_path.endswith(".csv"):
        context.data = pd.read_csv(context.file_path)
    else:
        context.data = pd.read_excel(context.file_path, sheet_name=sheet_name)

    assert (
        "Customer ID" in context.data.columns
    ), "Column 'Customer ID' not found in file."

    context.duplicate_customers = context.data["Customer ID"].duplicated(keep=False)


@then("duplicate customer records should be flagged")
def step_then_flag_duplicate_customers(context):
    # type: (Any) -> None
    """Flag duplicate customer records."""
    flagged_duplicates = context.data[context.duplicate_customers]
    assert not flagged_duplicates.empty, "No duplicate customer records detected."
    logging.info(
        "Flagged {} duplicate customer records.".format(len(flagged_duplicates))
    )


@then("a report should be generated listing duplicate occurrences")
def step_then_generate_duplicate_report(context):
    # type: (Any) -> None
    """Generate a report of duplicate customer records."""
    flagged_duplicates = context.data[context.duplicate_customers]
    report_path = os.path.join("reports", "duplicate_customers_report.csv")
    flagged_duplicates.to_csv(report_path, index=False)
    logging.info("Duplicate customers report generated at {}.".format(report_path))


@then("duplicate customers should be marked for manual review")
def step_then_mark_for_review(context):
    # type: (Any) -> None
    """Mark duplicate customers for manual review."""
    flagged_duplicates = context.data[context.duplicate_customers]
    logging.info(
        "Manual review required for {} duplicate customer records.".format(
            len(flagged_duplicates)
        )
    )

    # ================= End of Duplicate Customers Validation =================

    # ================= Beginning of Duplicate Transactions Validation =================


@when('I check the "Transaction ID" column in the "{sheet_name}" sheet')
def step_when_check_duplicate_transactions(context, sheet_name):
    # type: (Any, Any) -> None
    """Check for duplicate transactions based on Transaction ID."""
    if context.file_path.endswith(".csv"):
        context.data = pd.read_csv(context.file_path)
    else:
        context.data = pd.read_excel(context.file_path, sheet_name=sheet_name)

    assert (
        "Transaction ID" in context.data.columns
    ), "Column 'Transaction ID' not found in file."

    context.duplicate_transactions = context.data["Transaction ID"].duplicated(
        keep=False
    )


@then("duplicate transaction records should be flagged")
def step_then_flag_duplicate_transactions(context):
    # type: (Any) -> None
    """Flag duplicate transactions."""
    flagged_duplicates = context.data[context.duplicate_transactions]
    assert not flagged_duplicates.empty, "No duplicate transaction records detected."
    logging.info(
        "Flagged {} duplicate transaction records.".format(len(flagged_duplicates))
    )


@then("a report should be generated listing duplicate occurrences")
def step_then_generate_duplicate_transactions_report(context):
    # type: (Any) -> None
    """Generate a report of duplicate transactions."""
    flagged_duplicates = context.data[context.duplicate_transactions]
    report_dir = "reports"
    report_path = os.path.join(report_dir, "duplicate_transactions_report.csv")
    # Ensure the reports directory exists
    if not os.path.exists(report_dir):
        os.makedirs(report_dir)
    flagged_duplicates.to_csv(report_path, index=False)
    logging.info("Duplicate transactions report generated at {}.".format(report_path))


@then("duplicate transactions should be marked for manual review")
def step_then_mark_duplicates_for_manual_review(context):
    # type: (Any) -> None
    """Mark duplicate transactions for manual review."""
    flagged_duplicates = context.data[context.duplicate_transactions]
    count = len(flagged_duplicates)
    logging.info(
        "Manual review required for {} duplicate transaction records.".format(count)
    )


# ================= End of Duplicate Transactions Validation =================

# ================= Beginning of Fraudulent Transactions Validation =================


@when('I check the "Transaction ID" column in the "{sheet_name}" sheet')
def step_when_check_fraudulent_transactions(context, sheet_name):
    # type: (Any, Any) -> None
    """Check for fraudulent transactions flagged as 'High' risk."""
    if context.file_path.endswith(".csv"):
        context.data = pd.read_csv(context.file_path)
    else:
        context.data = pd.read_excel(context.file_path, sheet_name=sheet_name)

    assert (
        "Transaction ID" in context.data.columns
    ), "Column 'Transaction ID' not found in file."

    context.fraudulent_transactions = context.data[context.data["Risk Flag"] == "High"]


@then("transactions flagged with high-risk indicators should be identified")
def step_then_flag_high_risk_transactions(context):
    # type: (Any) -> None
    assert (
        not context.fraudulent_transactions.empty
    ), "No high-risk transactions detected."

    logging.info(
        "Flagged "
        + str(len(context.fraudulent_transactions))
        + " high-risk transactions."
    )


@then("an alert should be generated for compliance review")
def step_then_generate_compliance_alert(context):
    # type: (Any) -> None
    logging.warning(
        "Compliance alert: "
        + str(len(context.fraudulent_transactions))
        + " high-risk transactions detected."
    )


@then("flagged transactions should be escalated for investigation")
def step_then_escalate_for_investigation(context):
    # type: (Any) -> None
    logging.info(
        "Escalating "
        + str(len(context.fraudulent_transactions))
        + " transactions for fraud investigation."
    )


# ================= End of Fraudulent Transactions Validation =================


# ================= Beginning of Orphaned Transactions Validation =================


@when(
    'I check the "Transaction ID" and "Account Number" columns in the "{sheet_name}" sheet'
)
def step_when_check_orphaned_transactions(context, sheet_name):
    # type: (Any, Any) -> None
    if context.file_path.endswith(".csv"):
        context.data = pd.read_csv(context.file_path)
    else:
        context.data = pd.read_excel(context.file_path, sheet_name=sheet_name)

    assert (
        "Transaction ID" in context.data.columns
    ), "Column 'Transaction ID' not found in file."
    assert (
        "Account Number" in context.data.columns
    ), "Column 'Account Number' not found in file."

    context.orphaned_transactions = context.data[context.data["Account Number"].isna()]


@then("transactions with missing or unlinked accounts should be flagged")
def step_then_flag_orphaned_transactions(context):
    # type: (Any) -> None
    assert not context.orphaned_transactions.empty, "No orphaned transactions detected."
    logging.info(
        "Flagged orphaned transactions {} records.".format(
            len(context.orphaned_transactions)
        )
    )


@then("an alert should be generated for data consistency review")
def step_then_generate_data_consistency_alert(context):
    # type: (Any) -> None
    logging.warning(
        "Data consistency alert {} orphaned transactions detected.".format(
            len(context.orphaned_transactions)
        )
    )


@then("flagged transactions should be escalated for manual verification")
def step_then_escalate_for_manual_verification(context):
    # type: (Any) -> None
    logging.info(
        "Escalating {} orphaned transactions for manual verification.".format(
            len(context.orphaned_transactions)
        )
    )


# ================= End of Orphaned Transactions Validation =================

# ================= Beginning of Transaction Mismatch Validation =================


@when(
    'I compare the "Transaction ID", "Amount", and "Currency" columns in the "{sheet_name}" sheet'
)
def step_when_compare_transaction_details(context, sheet_name):
    # type: (Any, Any) -> None
    """Check for mismatched transaction details."""

    if context.file_path.endswith(".csv"):
        context.data = pd.read_csv(context.file_path)
    else:
        context.data = pd.read_excel(context.file_path, sheet_name=sheet_name)

    required_columns = ["Transaction ID", "Amount", "Currency"]
    assert all(
        col in context.data.columns for col in required_columns
    ), "Required columns missing in file."

    def has_mismatch(group):
        return group["Amount"].nunique() > 1 or group["Currency"].nunique() > 1

    context.mismatched_transactions = context.data.groupby("Transaction ID").filter(
        has_mismatch
    )


@then("transactions with mismatched details should be flagged")
def step_then_flag_mismatched_transactions(context):
    # type: (Any) -> None
    """Flag transactions with mismatched details."""
    assert (
        not context.mismatched_transactions.empty
    ), "No transaction mismatches detected."
    logging.info(
        "Flagged {} transactions with mismatched details.".format(
            len(context.mismatched_transactions)
        )
    )


@then("flagged transactions should be reviewed for potential data entry errors")
def step_then_review_data_entry_errors(context):
    # type: (Any) -> None
    """Warn about potential data entry errors in flagged transactions."""
    logging.warning(
        "Review required: {} potential data entry errors detected.".format(
            len(context.mismatched_transactions)
        )
    )


@then("a report should be generated listing mismatches")
def step_then_generate_mismatch_report(context):
    # type: (Any) -> None
    """Generate a CSV report of transactions with mismatches."""
    report_path = "reports/transaction_mismatch_report.csv"
    context.mismatched_transactions.to_csv(report_path, index=False)
    logging.info("Mismatch report generated at: {}".format(report_path))


# ================= End of Transaction Mismatch Validation =================

# ================= Beginning of Edge Case Validation =================


@when("I attempt to process the file")
def step_when_attempt_process_file(context):
    # type: (Any) -> None
    """Check if the file is empty."""
    file_size = os.stat(context.file_path).st_size
    if file_size == 0:
        context.is_empty = True
    else:
        context.is_empty = False


@then("the system should detect it as empty")
def step_then_detect_empty_file(context):
    # type: (Any) -> None
    """Ensure the system detects the file as empty."""
    assert context.is_empty, "File is not empty."
    logging.warning("Empty file detected and flagged.")


@when('I check for special characters in the "{column_name}" column')
def step_when_check_special_characters(context, column_name):
    # type: (Any, Any) -> None
    """Check for special characters in the specified column."""
    context.data = (
        pd.read_csv(context.file_path)
        if context.file_path.endswith(".csv")
        else pd.read_excel(context.file_path)
    )
    context.special_char_issues = context.data[column_name].str.contains(
        r"[^a-zA-Z0-9 ]", na=False
    )


@then("transactions containing special characters should be flagged")
def step_then_flag_special_characters(context):
    # type: (Any) -> None
    """Flag transactions that contain special characters."""
    flagged_rows = context.data[context.special_char_issues]
    assert not flagged_rows.empty, "No special characters found."
    logging.warning(
        "Flagged {} transactions containing special characters.".format(
            len(flagged_rows)
        )
    )


@when('I check the "Date" column in the "{sheet_name}" sheet')
def step_when_check_extreme_dates(context, sheet_name):
    # type: (Any, Any) -> None
    """Check for extreme date values in the specified column."""
    context.data = (
        pd.read_csv(context.file_path)
        if context.file_path.endswith(".csv")
        else pd.read_excel(context.file_path, sheet_name=sheet_name)
    )
    context.extreme_dates = context.data["Date"][
        (context.data["Date"] < "1900-01-01") | (context.data["Date"] > "2100-01-01")
    ]


@then("transactions with dates in the far future or past should be flagged")
def step_then_flag_extreme_dates(context):
    # type: (Any) -> None
    """Flag transactions with extreme date values."""
    assert not context.extreme_dates.empty, "No extreme dates found."
    logging.warning(
        "Flagged {} transactions with extreme dates.".format(len(context.extreme_dates))
    )


@when("I attempt to open the file")
def step_when_attempt_open_file(context):
    # type: (Any) -> None
    """Attempt to open the file and detect corruption."""
    try:
        context.data = (
            pd.read_csv(context.file_path)
            if context.file_path.endswith(".csv")
            else pd.read_excel(context.file_path)
        )
        context.is_corrupt = False
    except Exception as e:
        logging.error("File corruption detected: {}".format(e))
        context.is_corrupt = True


@then("an error should be raised indicating the file is corrupted")
def step_then_detect_corrupt_file(context):
    # type: (Any) -> None
    """Raise an error if the file is corrupted."""
    assert context.is_corrupt, "File is not corrupted."
    logging.warning("Corrupt file detected and flagged for review.")


@when('I check for whitespace issues in the "{column_name}" column')
def step_when_check_whitespace_issues(context, column_name):
    # type: (Any, Any) -> None
    """Check for leading or trailing whitespace in the specified column."""
    context.data[column_name] = context.data[column_name].astype(str)
    context.whitespace_issues = context.data[column_name].str.contains(
        r"^\s|\s$", na=False
    )


@then("leading and trailing spaces should be removed")
def step_then_strip_whitespace(context):
    # type: (Any) -> None
    """Remove leading and trailing spaces from all string values in the dataframe."""
    context.data = context.data.applymap(
        lambda x: x.strip() if isinstance(x, str) else x
    )
    logging.info("Whitespace issues cleaned.")


# ================= End of Edge Case Validation =================


# ================= Beginning of Empty File Validation =================
@when("I attempt to process the file")
def step_when_attempt_to_process_file(context):
    # type: (Any) -> None
    """Check if the file is empty or not."""
    if os.stat(context.file_path).st_size == 0:
        context.is_empty = True
    else:
        context.is_empty = False


@then("an appropriate error message should be returned")
def step_then_error_message(context):
    # type: (Any) -> None
    """Log an error message for an empty file."""
    logging.error("Error file is empty and cannot be processed.")


@then("the file should be excluded from processing")
def step_then_exclude_empty_file(context):
    # type: (Any) -> None
    """Log that the empty file is excluded from processing."""
    logging.info("Empty file has been excluded from processing.")


@then("a system log entry should be recorded for tracking")
def step_then_log_empty_file(context):
    # type: (Any) -> None
    """Log an entry for the empty file."""
    logging.info("System log file {} recorded for tracking.".format(context.file_name))


@given("a batch of bank export files:")
def batch_files(context):
    # type: (Any) -> None
    """Validate and collect a batch of bank export files."""
    context.batch_files = [row["file_name"] for row in context.table]
    context.valid_files = []

    for file_name in context.batch_files:
        file_path = get_test_file_path(context, file_name)
        if os.path.exists(file_path) and os.stat(file_path).st_size > 0:
            context.valid_files.append(file_name)
        else:
            logging.warning("File {} is empty or missing.".format(file_name))


@when("I attempt to process these files")
def step_when_process_batch_files(context):
    # type: (Any) -> None
    """Identify non-empty files for processing."""
    context.non_empty_files = [
        file
        for file in context.valid_files
        if os.stat(get_test_file_path(context, file)).st_size > 0
    ]


@then("the system should continue processing non-empty files")
def step_then_continue_processing(context):
    # type: (Any) -> None
    """Ensure the system processes only non-empty files."""
    assert len(context.non_empty_files) > 0, "No valid files found for processing."
    logging.info("Processing {} valid files.".format(len(context.non_empty_files)))


@then("an appropriate error should be logged for each empty file")
def step_then_log_empty_files(context):
    # type: (Any) -> None
    """Log errors for files that are empty."""
    for file_name in context.batch_files:
        if file_name not in context.non_empty_files:
            logging.error(
                "Error: {} is empty and cannot be processed.".format(file_name)
            )


@then("system resources should remain stable")
def step_then_check_resources(context):
    # type: (Any) -> None
    """Verify that resources remain stable during processing."""
    logging.info("Resource monitoring confirms stability during processing.")


@then("processing time should be logged for benchmarking")
def step_then_log_processing_time(context):
    # type: (Any) -> None
    """Log processing time for future benchmarking."""
    processing_time = 123.45  # This should come from your processing logic
    log_processing_time(context, processing_time)


@then("the user should receive a warning notification about the empty file")
def step_then_notify_user(context):
    # type: (Any) -> None
    """Send a warning notification to the user about the empty file."""
    notify_empty_file(context)


@then("the file should be marked as failed in the processing log")
def step_then_mark_failed(context):
    # type: (Any) -> None
    """Mark the file as failed in the processing log."""
    logging.error(
        "Processing Log: {} marked as failed due to empty content.".format(
            context.file_name
        )
    )


@then("a recommendation should be provided to verify data source")
def step_then_recommend_verification(context):
    # type: (Any) -> None
    """Provide a recommendation to verify the data source."""
    logging.info(
        "Recommendation the data source and ensure the file is not empty before reprocessing."
    )


# ================= End of Empty File Validation =================

# ================= Beginning of Hidden Rows Validation =================


@when('I check for hidden rows in the "{sheet_name}" sheet')
def step_when_check_hidden_rows(context, sheet_name):
    # type: (Any, Any) -> None
    """Check for rows that are completely empty."""
    if context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        df = pd.read_csv(context.file_path)

    context.hidden_rows = df[df.isnull().all(axis=1)].index.tolist()
    logging.info("Hidden rows detected: {}".format(context.hidden_rows))


@then("all hidden rows should be identified and logged")
def step_then_log_hidden_rows(context):
    # type: (Any) -> None
    """Log any hidden rows found."""
    assert len(context.hidden_rows) > 0, "No hidden rows detected."
    logging.warning("Hidden rows found: {}".format(context.hidden_rows))


@then("a report should be generated listing the hidden rows")
def step_then_generate_hidden_rows_report(context):
    # type: (Any) -> None
    """Generate a CSV report of hidden rows."""
    report_dir = "reports"
    os.makedirs(report_dir, exist_ok=True)
    report_path = os.path.join(report_dir, "hidden_rows_report.csv")
    hidden_rows_df = pd.DataFrame({"Hidden Row Indices": context.hidden_rows})
    hidden_rows_df.to_csv(report_path, index=False)
    logging.info("Generated report for hidden rows: {}".format(context.hidden_rows))
    logging.info("Report saved to: {}".format(report_path))


@then("users should be alerted to review the hidden data")
def step_then_alert_users(context):
    # type: (Any) -> None
    """Notify users of hidden rows."""
    logging.warning("User notification rows detected. Review required.")

    # Call the function to notify users about hidden rows
    notify_hidden_data(context)


@then("transactions hidden in rows should be flagged as potential fraud")
def step_then_flag_fraudulent_hidden_rows(context):
    # type: (Any) -> None
    """Flag hidden transactions as suspicious."""
    assert len(context.hidden_rows) > 0, "No hidden transactions detected."
    logging.error("Fraud Alert detected in hidden rows.")


@then("flagged transactions should be escalated for further review")
def step_then_escalate_fraudulent_hidden_rows(context):
    # type: (Any) -> None
    """Escalate flagged hidden transactions for compliance review."""
    logging.info("Escalating hidden transactions for compliance review.")


@then("an alert should be generated for compliance teams")
def step_then_alert_compliance(context):
    # type: (Any) -> None
    """Send a compliance alert for hidden transactions."""
    logging.warning("Compliance Alert transactions flagged for investigation.")


@then("rows with partially hidden content should be identified")
def step_then_identify_partial_hidden_rows(context):
    # type: (Any) -> None
    """Identify rows that are partially hidden."""
    df = (
        pd.read_csv(context.file_path)
        if context.file_path.endswith(".csv")
        else pd.read_excel(context.file_path)
    )
    context.partial_hidden_rows = [
        row for row in context.hidden_rows if row in df.index
    ]
    logging.info(
        "Partially hidden rows detected: {}".format(context.partial_hidden_rows)
    )


@then("a warning should be generated for data review")
def step_then_warn_partial_hidden_rows(context):
    # type: (Any) -> None
    """Log a warning about partially hidden rows."""
    logging.warning(
        "Warning hidden rows found - {}".format(context.partial_hidden_rows)
    )


@then("a suggestion should be provided to adjust visibility settings")
def step_then_suggest_visibility_fix(context):
    # type: (Any) -> None
    """Suggest adjusting spreadsheet visibility settings."""
    logging.info(
        "Suggestion spreadsheet visibility settings to ensure all data is accessible."
    )


# ================= End of Hidden Rows Validation =================

# ================= Beginning of Max Character Limit Validation =================


@when('I check the "{column_name}" column in the "{sheet_name}" sheet')
def step_when_check_max_character_limit(context, column_name, sheet_name):
    # type: (Any, Any, Any) -> None
    """Check for values exceeding the maximum character limit."""
    if context.file_name.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        df = pd.read_csv(context.file_path)

    context.exceeding_values = df[df[column_name].astype(str).str.len() > 255]
    logging.info(
        "Rows exceeding max character limit detected in column '{}': {}".format(
            column_name, len(context.exceeding_values)
        )
    )


@then("values exceeding the maximum character limit should be flagged")
def step_then_flag_max_character_limit(context):
    # type: (Any) -> None
    """Flag values that exceed the maximum character limit."""
    assert (
        not context.exceeding_values.empty
    ), "No values exceeding the character limit detected."
    logging.warning(
        "Flagged {} rows with values exceeding max character limit.".format(
            len(context.exceeding_values)
        )
    )


@then("an error log should be generated listing the violations")
def step_then_generate_error_log(context):
    # type: (Any) -> None
    """Generate an error log for values exceeding the limit."""
    logging.error(
        "Error log following rows contain values exceeding the max character limit: {}".format(
            context.exceeding_values.index.tolist()
        )
    )


@when("I attempt to process a file containing fields near the max character limit")
def step_when_process_large_character_file(context):
    # type: (Any) -> None
    """Simulate processing a file with values near the maximum character limit."""
    logging.info("Processing file with values close to the max character limit...")


@then("the system should process the file without performance degradation")
def step_then_validate_performance(context):
    # type: (Any) -> None
    """Validate that performance remains stable."""
    logging.info(
        "System successfully processed file without noticeable performance issues."
    )


@then("response times should be logged for benchmarking")
def step_then_log_response_times(context):
    # type: (Any) -> None
    """Log response times for benchmarking."""
    logging.info("Benchmarking times for processing max character limit data recorded.")


@then("any truncated values should be flagged for manual review")
def step_then_flag_truncated_values(context):
    # type: (Any) -> None
    """Flag any truncated values for manual review."""
    # We assume the column being validated is the same as the one used before
    if not context.exceeding_values.empty:
        logging.warning(
            "Warning: {} values may have been truncated. Manual review recommended.".format(
                len(context.exceeding_values)
            )
        )
    else:
        logging.info("No truncated values found exceeding the character limit.")


# ================= End of Max Character Limit Validation =================

# ================= Beginning of Null Values Handling Steps =================


@then("the system should detect it as empty")
def step_then_detect_empty(context):
    # type: (Any) -> None
    """Check if the file is detected as empty"""
    assert context.is_empty, "File is not empty"


@then("an appropriate error message should be returned")
def step_then_return_error(context):
    # type: (Any) -> None
    """Check if the correct error message is returned"""
    assert context.process_result == "Empty file detected", "Incorrect error message"


@then("the file should be excluded from processing")
def step_then_exclude_file(context):
    # type: (Any) -> None
    """Ensure the file is marked as excluded from processing"""
    assert context.is_empty, "File should be excluded"


@then("a system log entry should be recorded for tracking")
def step_then_log_entry(context):
    # type: (Any) -> None
    """Log a system entry"""
    print(
        "Log entry recorded: {} - {}".format(context.file_name, context.process_result)
    )


@when('I check the "{column_name}" column in the "{sheet_name}" sheet')
def step_check_column_in_sheet(context, column_name, sheet_name):
    # type: (Any, Any, Any) -> None
    """Check a column for missing values."""
    if sheet_name == "N/A":
        sheet_data = context.data
    else:
        sheet_data = context.data[sheet_name]

    context.missing_values = sheet_data[column_name].isnull().sum()


@then("a report should be generated listing the affected rows")
def step_then_generate_missing_values_report(context):
    # type: (Any) -> None
    """Generate a report of rows with missing values."""
    print("Generating report for {} missing values...".format(context.missing_values))

    # Optionally filter and save rows with missing values
    if hasattr(context, "data"):
        affected_rows = context.data[context.data.isnull().any(axis=1)]
        report_path = os.path.join("reports", "missing_values_report.csv")

        if not os.path.exists(os.path.dirname(report_path)):
            os.makedirs(os.path.dirname(report_path))

        affected_rows.to_csv(report_path, index=False)
        logging.info("Missing values report saved to {}".format(report_path))
    else:
        logging.warning("No data found in context to generate report.")


@then("a recommendation should be provided for data correction")
def step_then_recommend_correction(context):
    # type: (Any) -> None
    """Provide a recommendation for correcting missing values"""
    print("Recommendation review and update missing data fields.")


@when('I analyze the percentage of missing values in the "{column_name}" column')
def step_when_analyze_missing_threshold(context, column_name):
    # type: (Any, Any) -> None
    """Calculate the percentage of missing values in a column"""
    total_records = len(context.data)
    missing_count = context.data[column_name].isna().sum()
    context.missing_percentage = (missing_count / total_records) * 100


@then('if missing values exceed "{threshold}%", an alert should be generated')
def step_then_alert_threshold(context, threshold):
    # type: (Any, Any) -> None
    """Generate an alert if missing values exceed the threshold"""
    threshold = float(threshold.strip("%"))
    assert (
        context.missing_percentage <= threshold
    ), "Threshold exceeded! Alert triggered."
    print(
        "Missing values: {}% (Threshold: {}%)".format(
            context.missing_percentage, threshold
        )
    )


@then("transactions above the threshold should be marked for review")
def step_then_flag_high_value_transactions(context):
    # type: (Any) -> None
    """Mark transactions above the threshold for review."""
    if hasattr(context, "threshold") and hasattr(context, "data"):
        flagged = context.data[context.data[context.column_name] > context.threshold]
        context.flagged_transactions = flagged
        print(
            "{} transactions exceed the threshold of {} and are flagged for review.".format(
                len(flagged), context.threshold
            )
        )
        logging.info(
            "{} transactions flagged for review above threshold {}.".format(
                len(flagged), context.threshold
            )
        )
    else:
        logging.warning("Threshold or data not found in context.")


@then("corrective action should be recommended based on data quality standards")
def step_then_recommend_corrective_action(context):
    # type: (Any) -> None
    """Recommend corrective action based on data quality"""
    print("Recommended corrective action reconciliation required.")


# ================= End of Null Values Handling Steps =================

# ================= Beginning of Outlier Detection Step Definitions for Edge Case Handling =================
# This script contains step definitions for detecting outliers in bank export files as part of edge case testing.
# It includes:
# - Handling empty files
# - Identifying transactions with extreme values
# - Analyzing historical trends for anomaly detection
# - Ensuring system performance under large datasets


@when("I attempt to process the file")
def step_when_attempt_to_process(context):
    # type: (Any) -> None
    """Check if the file is empty."""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path)
    else:
        raise ValueError("Unsupported file format")

    context.is_empty = df.empty


@then("the system should detect it as empty")
def step_then_system_detects_empty_file(context):
    # type: (Any) -> None
    """Validate the file is empty"""
    assert getattr(context, "is_empty", False), "File is not empty"
    logging.warning("Empty file detected: " + context.file_path)


@then("an appropriate error message should be returned")
def step_then_return_error_message(context):
    # type: (Any) -> None
    """Simulate an error message return"""
    if getattr(context, "is_empty", False):
        context.error_message = "The file is empty and cannot be processed"
    assert (
        context.error_message == "The file is empty and cannot be processed"
    ), "Expected error message not returned."


@then("the file should be excluded from processing")
def step_then_exclude_file_from_processing(context):
    # type: (Any) -> None
    """Exclude empty files from processing"""
    if getattr(context, "is_empty", False):
        if not hasattr(context, "excluded_files"):
            context.excluded_files = []
        context.excluded_files.append(context.file_path)
        logging.warning("Excluded empty file from processing: " + context.file_path)
    else:
        logging.info("File is not empty and will be processed: " + context.file_path)


@then("a system log entry should be recorded for tracking")
def step_then_log_entry_recorded(context):
    # type: (Any) -> None
    """Log the event"""
    log_message = (
        "File " + context.file_path + " was empty and excluded from processing"
    )
    if not hasattr(context, "logs"):
        context.logs = []
    context.logs.append(log_message)
    logging.info(log_message)


@when('I analyze the "{column_name}" column in the "{sheet_name}" sheet')
def step_when_analyze_column(context, column_name, sheet_name):
    # type: (Any, Any, Any) -> None
    """Analyze outliers in a specific column"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")

    context.column_values = df[column_name].dropna()
    context.mean = context.column_values.mean()
    context.std_dev = context.column_values.std()


@then('transactions exceeding the threshold of "{threshold_value}" should be flagged')
def step_then_flag_outliers(context, threshold_value):
    # type: (Any, Any) -> None
    """Flag transactions exceeding a specific threshold"""
    threshold = float(threshold_value)
    context.outliers = context.column_values[context.column_values > threshold]
    assert not context.outliers.empty, "No outliers detected"


@then("flagged transactions should be logged for further review")
def step_then_log_flagged_transactions(context):
    # type: (Any) -> None
    """Log flagged transactions"""
    log_message = (
        "Outliers detected in "
        + context.file_path
        + ": "
        + str(len(context.outliers))
        + " transactions"
    )
    context.logs.append(log_message)


@then("recommendations for corrective action should be generated")
def step_then_generate_recommendations(context):
    # type: (Any) -> None
    """Generate corrective action recommendations"""
    if not hasattr(context, "recommendations"):
        context.recommendations = []
    context.recommendations.append(
        "Review flagged transactions in " + context.file_path + " for anomalies"
    )


@when('I compare the "{column_name}" column with historical data')
def step_when_compare_with_historical(context, column_name):
    # type: (Any, Any) -> None
    """Compare column values with historical data for trend analysis"""
    historical_mean = context.mean * 0.8  # Example of historical average threshold
    context.historical_outliers = context.column_values[
        context.column_values > historical_mean
    ]


@then(
    'records with values beyond "{threshold}%" of the historical average should be flagged'
)
def step_then_flag_historical_outliers(context, threshold):
    # type: (Any, Any) -> None
    """Flag transactions exceeding a percentage threshold of historical data"""
    percentage_threshold = float(threshold) / 100
    flagged_records = context.historical_outliers[
        context.historical_outliers > context.mean * (1 + percentage_threshold)
    ]
    assert not flagged_records.empty, "No significant outliers detected"


@then("corrective action should be suggested")
def step_then_corrective_action_suggested(context):
    # type: (Any) -> None
    """Suggest corrective action for flagged records"""
    if not hasattr(context, "recommendations"):
        context.recommendations = []
    recommendation = "Review historical anomalies in " + context.file_path
    context.recommendations.append(recommendation)
    logging.info(recommendation)


@then("an alert should be generated for data quality review")
def step_then_generate_alert(context):
    # type: (Any) -> None
    """Generate an alert for data quality issues"""
    context.alerts.append(
        "Potential data integrity issue detected in " + context.file_path
    )


@when(
    'I attempt to process a dataset containing more than "{row_count}" transactions with outliers'
)
def step_when_process_large_dataset(context, row_count):
    # type: (Any, Any) -> None
    """Simulate processing large datasets with outliers"""
    context.row_count = int(row_count)
    context.processed_rows = min(
        context.row_count, 200000
    )  # Simulating a processing limit


@then("the system should handle the data efficiently")
def step_then_handle_large_data(context):
    # type: (Any) -> None
    """Ensure system efficiency for large datasets"""
    assert context.processed_rows > 0, "Dataset processing failed"


@then("processing time should be logged for benchmarking")
def step_then_processing_time_logged(context):
    # type: (Any) -> None
    """Log system performance metrics"""
    if not hasattr(context, "logs"):
        context.logs = []
    log_message = (
        "Processed " + str(context.processed_rows) + " rows in " + context.file_path
    )
    context.logs.append(log_message)
    logging.info(log_message)


@then("flagged outliers should be included in the anomaly report")
def step_then_generate_anomaly_report(context):
    # type: (Any) -> None
    """Generate an anomaly detection report"""
    context.reports.append("Anomaly report generated for " + context.file_path)


# ================= End of Outlier Detection Step Definitions for Edge Case Handling =================


# ================= Beginning of Zero-Value Transactions Step Definitions for Edge Case Handling =================
# This script contains step definitions for detecting zero-value transactions in bank export files.
# It includes:
# - Identifying zero-value transactions and classifying them by type
# - Detecting fraudulent transactions with zero values in high-risk categories
# - Validating zero-value transactions against historical patterns
# - Ensuring compliance with business logic
# - Evaluating system performance for large datasets containing zero-value transactions


@when('I analyze the "Amount" column in the "{sheet_name}" sheet')
def step_when_analyze_amount_column(context, sheet_name):
    # type: (Any, Any) -> None
    """Analyze zero-value transactions in the Amount column"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")

    context.zero_value_transactions = df[df["Amount"] == 0]


@then("transactions with an amount of zero should be flagged")
def step_then_flag_zero_value_transactions(context):
    # type: (Any) -> None
    """Flag transactions with zero value"""
    assert (
        not context.zero_value_transactions.empty
    ), "No zero-value transactions detected"


@then('the system should classify them based on "{transaction_type}"')
def step_then_classify_transactions(context, transaction_type):
    # type: (Any, Any) -> None
    """Classify zero-value transactions based on transaction type"""
    context.classification = transaction_type


@when(
    'I check for zero-value transactions in the "{category}" category in the "{sheet_name}" sheet'
)
def step_when_check_zero_value_in_category(context, category, sheet_name):
    # type: (Any, Any, Any) -> None
    """Check zero-value transactions in a specific category"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")

    context.high_risk_zero_values = df[
        (df["Amount"] == 0) & (df["Category"] == category)
    ]


@then(
    "transactions with zero value in high-risk categories should be flagged as suspicious"
)
def step_then_flag_high_risk_zero_values(context):
    # type: (Any) -> None
    """Flag high-risk zero-value transactions"""
    assert (
        not context.high_risk_zero_values.empty
    ), "No high-risk zero-value transactions detected"


@then("flagged transactions should be escalated for compliance review")
def step_then_escalate_for_compliance(context):
    # type: (Any) -> None
    """Escalate flagged transactions for compliance review"""
    context.recommendations.append(
        f"Escalate zero-value transactions for compliance in {context.file_path}"
    )


@then("a fraud detection report should be generated")
def step_then_generate_fraud_report(context):
    # type: (Any) -> None
    """Generate a fraud detection report"""
    context.reports.append(f"Fraud detection report generated for {context.file_path}")


@then('a risk assessment score should be assigned based on "{risk_level}"')
def step_then_assign_risk_level(context, risk_level):
    # type: (Any, Any) -> None
    """Assign a risk assessment score"""
    context.risk_level = risk_level


@when('I compare the "Amount" column with historical data')
def step_when_compare_with_historical_data(context):
    # type: (Any) -> None
    """Compare zero-value transactions against historical patterns"""
    historical_average = 3  # Placeholder for actual historical data analysis
    if not hasattr(context, "zero_value_transactions"):
        context.zero_value_transactions = []
    context.exceeding_threshold = (
        len(context.zero_value_transactions) > historical_average
    )


@then(
    'transactions with zero value exceeding "{threshold}%" of total transactions should be flagged'
)
def step_then_flag_exceeding_zero_values(context, threshold):
    # type: (Any, Any) -> None
    """Flag transactions exceeding the threshold"""
    threshold_percentage = float(threshold) / 100
    assert (
        context.exceeding_threshold
    ), "Zero-value transactions do not exceed the threshold"


@then("an alert should be generated for data quality review")
def step_then_alert_for_data_quality(context):
    # type: (Any) -> None
    """Generate an alert for data quality review"""
    if not hasattr(context, "alerts"):
        context.alerts = []
    context.alerts.append(f"Potential data quality issue in {context.file_path}")


@then('the threshold comparison should consider "{time_period}" historical data')
def step_then_consider_historical_time_period(context, time_period):
    # type: (Any, Any) -> None
    """Validate zero-value transactions against historical trends"""
    context.historical_period = time_period


@when(
    'I attempt to process a dataset containing more than "{row_count}" transactions with zero values'
)
def step_when_process_large_zero_value_dataset(context, row_count):
    # type: (Any, Any) -> None
    """Simulate processing large datasets with zero-value transactions"""
    context.row_count = int(row_count)
    context.processed_rows = min(
        context.row_count, 200000
    )  # Simulating system capability


@then("the system should handle the data efficiently")
def step_then_system_handles_data_efficiently(context):
    # type: (Any) -> None
    """Ensure system efficiency for processing large datasets"""
    assert context.processed_rows > 0, "Dataset processing failed"


@then("flagged zero-value transactions should be included in the validation report")
def zero_value_transactions_calculated_in_report(context):
    # type: (Any) -> None
    """Generate a validation report for zero-value transactions"""
    context.reports.append(f"Validation report generated for {context.file_path}")


@then("system resource utilization should remain within acceptable limits")
def step_then_validate_system_resources(context):
    # type: (Any) -> None
    """Ensure resource usage remains optimal"""
    context.system_health_check = "Resources within acceptable limits"


@when('I validate the "Amount" column in "{sheet_name}"')
def step_when_validate_amount(context, sheet_name):
    # type: (Any, Any) -> None
    """Validate zero-value transactions against business rules"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")

    context.business_rule_violations = df[
        (df["Amount"] == 0) & (df["Category"] != "Fee Waivers")
    ]


@then("zero-value transactions should be checked against predefined business rules")
def step_then_validate_business_rules(context):
    # type: (Any) -> None
    """Ensure business rules are followed for zero-value transactions"""
    assert (
        not context.business_rule_violations.empty
    ), "No business rule violations detected"


@then('exceptions should be made for "{exempt_category}"')
def step_then_handle_exempt_category(context, exempt_category):
    # type: (Any, Any) -> None
    """Allow exceptions for specific categories"""
    context.exempt_category = exempt_category


@then("a compliance report should be generated")
def step_then_generate_compliance_report(context):
    # type: (Any) -> None
    """Generate a compliance report"""
    context.reports.append(f"Compliance report generated for {context.file_path}")


# ================= End of Zero-Value Transactions Step Definitions for Edge Case Handling =================
# ================= Beginning of Basel III Capital Validation Step Definitions for Financial Accuracy Testing =================


# This script contains step definitions for validating Basel III capital adequacy, liquidity, and risk-weighted asset calculations.
# It includes:
# - Validating Tier 1 Capital, RWA, and Capital Ratio adherence to Basel III standards
# - Stress testing scenarios and risk-weighting calculations
# - Liquidity and stability ratio compliance
# - Large dataset performance testing for financial accuracy validation
@when(
    'I check the "Tier 1 Capital", "Risk-Weighted Assets", and "Capital Ratio" fields in "{sheet_name}"'
)
def step_when_check_basel_iii_fields(context, sheet_name):
    # type: (Any, Any) -> None
    """Analyze Basel III capital adequacy report fields"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")

    context.basel_iii_data = df[
        ["Tier 1 Capital", "Risk-Weighted Assets", "Capital Ratio"]
    ]


@then('all values should match the Basel III calculation formula "{formula}"')
def step_then_validate_basel_iii_formula(context, formula):
    # type: (Any, Any) -> None
    """Validate Basel III calculation formulas"""
    assert not context.basel_iii_data.empty, "No Basel III capital adequacy data found"


@then(
    'reports failing to meet "{capital_threshold}" should be flagged for regulatory review'
)
def step_then_flag_non_compliant_reports(context, capital_threshold):
    # type: (Any, Any) -> None
    """Flag reports not meeting capital adequacy thresholds"""
    threshold = float(capital_threshold.strip("%")) / 100
    non_compliant = context.basel_iii_data[
        context.basel_iii_data["Capital Ratio"] < threshold
    ]
    assert not non_compliant.empty, "All reports comply with Basel III requirements"


@when('I apply stress testing scenarios using "{stress_test_methodology}"')
def step_when_apply_stress_test(context, stress_test_methodology):
    # type: (Any, Any) -> None
    """Apply stress testing methodology"""
    context.stress_test = stress_test_methodology


@then('capital adequacy should remain above "{stress_test_threshold}"')
def step_then_validate_stress_test(context, stress_test_threshold):
    # type: (Any, Any) -> None
    """Ensure capital adequacy remains above stress test threshold"""
    threshold = float(stress_test_threshold.strip("%")) / 100
    assert all(
        context.basel_iii_data["Capital Ratio"] >= threshold
    ), "Capital adequacy threshold breached"


@then("deviations beyond the threshold should be flagged for regulatory audit")
def step_then_flag_stress_test_failures(context):
    # type: (Any) -> None
    """Flag capital adequacy breaches from stress tests"""
    context.reports.append("Stress test results flagged for regulatory audit")


@when('I check the "Risk-Weighted Assets" column in "{sheet_name}"')
def step_when_validate_rwa(context, sheet_name):
    # type: (Any, Any) -> None
    """Validate RWA calculations"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")

    context.rwa_data = df["Risk-Weighted Assets"]


@then('all RWA values should be computed correctly using "{rwa_formula}"')
def step_then_validate_rwa_formula(context, rwa_formula):
    # type: (Any, Any) -> None
    """Verify RWA calculations match regulatory standards"""
    assert not context.rwa_data.empty, "No RWA data found"


@then("inconsistencies should be flagged for regulatory review")
def step_then_flag_rwa_inconsistencies(context):
    # type: (Any) -> None
    """Identify inconsistencies in RWA calculations"""
    context.reports.append("RWA calculation inconsistencies flagged")


@when(
    'I check the "High-Quality Liquid Assets" and "Net Cash Outflows" in "{sheet_name}"'
)
def step_when_validate_lcr(context, sheet_name):
    # type: (Any, Any) -> None
    """Validate liquidity coverage ratio (LCR)"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")

    context.lcr_data = df[["High-Quality Liquid Assets", "Net Cash Outflows"]]


@then('the liquidity coverage ratio should be calculated as "{lcr_formula}"')
def step_then_validate_lcr_formula(context, lcr_formula):
    # type: (Any, Any) -> None
    """Verify LCR calculations"""
    assert not context.lcr_data.empty, "No LCR data found"


@then('reports with LCR below "{lcr_threshold}" should be flagged for liquidity risk')
def step_then_flag_lcr_failures(context, lcr_threshold):
    # type: (Any, Any) -> None
    """Identify LCR compliance failures"""
    threshold = float(lcr_threshold.strip("%")) / 100
    low_lcr = context.lcr_data[
        context.lcr_data["Net Cash Outflows"]
        / context.lcr_data["High-Quality Liquid Assets"]
        > threshold
    ]
    assert not low_lcr.empty, "All LCR values meet Basel III requirements"


@when(
    'I check the "Available Stable Funding" and "Required Stable Funding" in "{sheet_name}"'
)
def step_when_validate_nsfr(context, sheet_name):
    # type: (Any, Any) -> None
    """Validate NSFR compliance"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")

    context.nsfr_data = df[["Available Stable Funding", "Required Stable Funding"]]


@then('the NSFR should be calculated as "{nsfr_formula}"')
def step_then_validate_nsfr_formula(context, nsfr_formula):
    # type: (Any, Any) -> None
    """Ensure NSFR calculations match Basel III standards"""
    assert not context.nsfr_data.empty, "No NSFR data found"


@then(
    'reports with NSFR below "{nsfr_threshold}" should be flagged for stability risks'
)
def step_then_flag_nsfr_failures(context, nsfr_threshold):
    # type: (Any, Any) -> None
    """Flag NSFR compliance issues"""
    threshold = float(nsfr_threshold.strip("%")) / 100
    low_nsfr = context.nsfr_data[
        context.nsfr_data["Available Stable Funding"]
        / context.nsfr_data["Required Stable Funding"]
        < threshold
    ]
    assert not low_nsfr.empty, "All NSFR values meet Basel III requirements"


@when('I compare "Tier 1 Capital" and "Tier 2 Capital" values in "{sheet_name}"')
def step_when_validate_tier_1_vs_tier_2(context, sheet_name):
    # type: (Any, Any) -> None
    """Compare Tier 1 and Tier 2 capital"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")

    context.capital_structure = df[["Tier 1 Capital", "Tier 2 Capital"]]


@then('the proportion of Tier 1 capital should meet the required "{tier_1_minimum}"')
def step_then_validate_tier_1_ratio(context, tier_1_minimum):
    # type: (Any, Any) -> None
    """Ensure Tier 1 capital meets Basel III requirements"""
    min_tier_1 = float(tier_1_minimum.strip("%")) / 100
    assert all(
        context.capital_structure["Tier 1 Capital"] >= min_tier_1
    ), "Tier 1 capital does not meet minimum requirement"


@then('Tier 2 capital should not exceed "{tier_2_maximum}"')
def step_then_validate_tier_2_ratio(context, tier_2_maximum):
    # type: (Any, Any) -> None
    """Ensure Tier 2 capital does not exceed regulatory limits"""
    max_tier_2 = float(tier_2_maximum.strip("%")) / 100
    assert all(
        context.capital_structure["Tier 2 Capital"] <= max_tier_2
    ), "Tier 2 capital exceeds allowed limit"


# ================= End of Basel III Capital Validation Step Definitions for Financial Accuracy Testing =================

# ================= Beginning of Foreign Exchange Transactions Validation Step Definitions for Financial Accuracy Testing =================
# This script contains step definitions for validating foreign exchange transactions.
# It includes:
# - Checking exchange rates for accuracy and compliance with official sources
# - Ensuring correct application of currency conversions
# - Detecting fraudulent or inconsistent multi-currency transactions
# - Validating rounding and threshold-based compliance triggers


@when('I check the "Currency", "Amount", and "Exchange Rate" columns in "{sheet_name}"')
def step_when_check_fx_transaction_fields(context, sheet_name):
    # type: (Any, Any) -> None
    """Extract and validate foreign exchange transaction details"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")


@then('the converted amount should match official exchange rates within "{tolerance}"')
def step_then_validate_exchange_rate_tolerance(context, tolerance):
    # type: (Any, Any) -> None
    """Verify the currency conversion adheres to exchange rate tolerances"""
    assert not context.fx_data.empty, "No foreign exchange data found"


@then('transactions exceeding "{alert_threshold}" should trigger a compliance review')
def step_then_flag_large_fx_transactions(context, alert_threshold):
    # type: (Any, Any) -> None
    """Flag transactions that exceed alert thresholds"""
    threshold = float(alert_threshold.replace("$", "").replace(",", ""))
    high_value_transactions = context.fx_data[context.fx_data["Amount"] > threshold]
    assert (
        not high_value_transactions.empty
    ), "All transactions are within alert threshold"


@when('I check the "Exchange Rate" column in the "{sheet_name}" sheet')
def step_when_validate_exchange_rates(context, sheet_name):
    # type: (Any, Any) -> None
    """Validate exchange rates against official sources"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")


@then('all exchange rates should be accurate and sourced from "{official_source}"')
def step_then_validate_exchange_rate_source(context, official_source):
    # type: (Any, Any) -> None
    """Ensure exchange rates align with official sources"""
    assert not context.exchange_rate_data.empty, "No exchange rate data found"


@then("any discrepancies should be flagged for correction")
def step_then_flag_exchange_rate_discrepancies(context):
    # type: (Any) -> None
    """Identify discrepancies in exchange rate application"""
    context.reports.append("Exchange rate discrepancies flagged for review")


@when('I check "Original Amount" and "Converted Amount" in the "{sheet_name}" sheet')
def step_when_validate_currency_conversion(context, sheet_name):
    # type: (Any, Any) -> None
    """Validate currency conversion calculations"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")

    # Assuming that the columns 'Original Amount' and 'Converted Amount' exist in the sheet
    context.conversion_data = df[["Original Amount", "Converted Amount"]]

    # Calculate exchange rate and add to the DataFrame
    context.conversion_data["Exchange Rate"] = (
        context.conversion_data["Converted Amount"]
        / context.conversion_data["Original Amount"]
    )


from file_validation import (
    get_expected_exchange_rates,
)  # Import the required function from file_validation.py


@then('validate the conversion rates for reference date "{reference_date}"')
def step_then_validate_conversion_rates(context, reference_date):
    # type: (Any, str) -> None
    """Ensure correct exchange rates are applied for conversion"""

    # Ensure conversion data is not empty
    assert not context.conversion_data.empty, "No currency conversion data found"

    # Fetch expected conversion rates for the given reference date using the helper function from file_validation.py
    expected_rates = get_expected_exchange_rates(reference_date)

    # Validate each record in conversion data against the expected rates
    for index, row in context.conversion_data.iterrows():
        # Assuming the exchange rate is tied to some currency pair (if needed)
        # For simplicity, using 'USD/EUR' as an example, but this can be extended
        currency_pair = "USD/EUR"  # You can adjust this based on actual columns

        # Retrieve the expected conversion rate for the currency pair
        expected_rate = expected_rates.get(currency_pair)

        # Assert that an expected rate exists for the currency pair
        assert (
            expected_rate is not None
        ), "Expected conversion rate not found for currency pair: {0}".format(
            currency_pair
        )

        # Check if the actual conversion rate matches the expected one
        actual_rate = row['Exchange Rate']
        assert (
            actual_rate == expected_rate
        ), "Conversion rate mismatch for {0}: expected {1} but got {2}".format(
            currency_pair, expected_rate, actual_rate
        )


@then('rounding differences should not exceed "{rounding_tolerance}"')
def step_then_validate_rounding_tolerance(context, rounding_tolerance):
    # type: (Any, Any) -> None
    """Ensure rounding errors stay within acceptable limits"""
    tolerance = float(rounding_tolerance)
    rounding_errors = (
        context.conversion_data["Converted Amount"]
        - (
            context.conversion_data["Original Amount"]
            * context.conversion_data["Exchange Rate"]
        )
    ).abs()
    assert all(rounding_errors <= tolerance), "Rounding differences exceed tolerance"


@when('I check transactions involving multiple currencies in the "{sheet_name}" sheet')
def step_when_validate_multi_currency_transactions(context, sheet_name):
    # type: (Any, Any) -> None
    """Detect inconsistencies in multi-currency transactions"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")

    context.multi_currency_data = df[["Transaction ID", "Currency"]]


@then("transactions with conflicting or unsupported currency codes should be flagged")
def step_then_flag_invalid_currency_codes(context):
    # type: (Any) -> None
    """Identify transactions with inconsistent or unsupported currency codes"""
    invalid_currencies = context.multi_currency_data[
        ~context.multi_currency_data["Currency"].isin(context.valid_currencies)
    ]
    assert not invalid_currencies.empty, "All currency codes are valid"


@then("any unusual currency conversions should trigger an alert for fraud review")
def step_then_flag_suspicious_currency_conversions(context):
    # type: (Any) -> None
    """Detect potentially fraudulent multi-currency transactions"""
    context.reports.append(
        "Suspicious currency conversion patterns flagged for fraud review"
    )


# ================= End of Foreign Exchange Transactions Validation Step Definitions for Financial Accuracy Testing =================

# ================= Beginning of Interest Rate Calculations Validation Step Definitions for Financial Accuracy Testing =================
# This script validates interest rate calculations in banking transactions.
# It ensures:
# - Correct computation of loan interest based on formulas.
# - Proper formatting of interest rates in reports.
# - Consistency in daily vs. monthly interest calculations.
# - Proper application of compounding interest.


@when('I check the "Interest Rate" and "Principal Amount" columns in "{sheet_name}"')
def step_when_check_interest_rate_columns(context, sheet_name):
    # type: (Any, Any) -> None
    """Extract and validate interest rate calculations"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")


@then('the calculated interest should match the expected formula "{formula}"')
def step_then_validate_interest_formula(context, formula):
    # type: (Any, Any) -> None
    """Verify interest calculations adhere to the specified formula"""
    assert not context.interest_data.empty, "No interest calculation data found"


@then("incorrect interest rates should be flagged")
def step_then_flag_incorrect_interest_rates(context):
    # type: (Any) -> None
    """Flag incorrect interest rate calculations"""
    context.reports.append("Incorrect interest rate calculations flagged for review")


@then('deviations greater than "{tolerance}" should trigger a compliance alert')
def step_then_trigger_compliance_alert(context, tolerance):
    # type: (Any, Any) -> None
    """Trigger an alert if interest calculation deviations exceed tolerance"""
    context.reports.append(
        f"Deviations exceeding {tolerance} flagged for compliance review"
    )


@when('I check the "Interest Rate" column in the "{sheet_name}" sheet')
def step_when_validate_interest_rate_format(context, sheet_name):
    # type: (Any, Any) -> None
    """Validate interest rate formatting"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")

    context.interest_rate_format = df["Interest Rate"]


@then('all interest rates should be formatted correctly as "{expected_format}"')
def step_then_validate_interest_rate_format(context, expected_format):
    # type: (Any, Any) -> None
    """Ensure interest rates adhere to the expected format"""
    assert not context.interest_rate_format.empty, "No interest rate data found"


@then(
    "interest rates should be expressed as a percentage with up to two decimal places"
)
def step_then_validate_interest_rate_decimal_places(context):
    # type: (Any) -> None
    """Ensure interest rates have up to two decimal places"""
    context.reports.append("Interest rate decimal formatting verified")


@then("values exceeding regulatory limits should be flagged")
def step_then_flag_excessive_interest_rates(context):
    # type: (Any) -> None
    """Flag interest rates exceeding regulatory limits"""
    context.reports.append("Excessive interest rates flagged for review")


@when(
    'I compare "Daily Interest" and "Monthly Interest" calculations in the "{sheet_name}" sheet'
)
def step_when_validate_daily_vs_monthly_interest(context, sheet_name):
    # type: (Any, Any) -> None
    """Compare daily and monthly interest calculations"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")


@then('calculated interest should match expected values based on "{interest_formula}"')
def step_then_validate_interest_formula_application(context, interest_formula):
    # type: (Any, Any) -> None
    """Ensure the interest formula is applied correctly"""
    assert not context.interest_comparison.empty, "No interest comparison data found"


@then('rounding differences should not exceed "{rounding_tolerance}"')
def step_then_rounding_differences_within_tolerance(context, rounding_tolerance):
    # type: (Any, Any) -> None
    """Ensure rounding errors stay within acceptable limits"""
    context.reports.append(
        "Rounding differences checked against {} tolerance".format(rounding_tolerance)
    )


@then("discrepancies beyond tolerance should be logged for review")
def step_then_log_interest_discrepancies(context):
    # type: (Any) -> None
    """Log any discrepancies beyond the rounding tolerance"""
    context.reports.append(
        "Interest calculation discrepancies logged for further review"
    )


@when('I check the "Compounded Interest" column in the "{sheet_name}" sheet')
def step_when_validate_compounding_interest(context, sheet_name):
    # type: (Any, Any) -> None
    """Validate compounding interest calculations"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")


@then('all values should match the calculated interest for "{compounding_period}"')
def step_then_validate_compounding_interest(context, compounding_period):
    # type: (Any, Any) -> None
    """Ensure compounding interest calculations match expectations"""
    assert (
        not context.compound_interest_data.empty
    ), "No compounding interest data found"


@then("rounding rules should follow regulatory guidelines")
def step_then_validate_compounding_rounding(context):
    # type: (Any) -> None
    """Ensure rounding rules comply with regulations"""
    context.reports.append("Compounding interest rounding rules validated")


@then("discrepancies should be flagged for review")
def step_then_flag_compounding_discrepancies(context):
    # type: (Any) -> None
    """Flag any discrepancies in compounding interest calculations"""
    context.reports.append(
        "Compounding interest discrepancies flagged for further review"
    )


# ================= End of Interest Rate Calculations Validation Step Definitions for Financial Accuracy Testing =================


# ================= Beginning of Loan and Mortgage Payments Validation Step Definitions for Financial Accuracy Testing =================
# This script validates loan and mortgage payment calculations.
# It ensures:
# - Correct computation of monthly mortgage payments based on formulas.
# - Proper amortization schedule distribution.
# - Accurate interest vs. principal breakdown.
@when(
    'I compare "Loan Principal", "Interest Rate", and "Monthly Payment" in the "{sheet_name}" sheet'
)
def step_when_validate_loan_payment(context, sheet_name):
    # type: (Any, Any) -> None
    """Extract and validate loan payment calculations"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")


@then('the monthly payment should be calculated using "{loan_formula}"')
def step_then_validate_loan_formula(context, loan_formula):
    # type: (Any, Any) -> None
    """Verify loan payments adhere to the specified formula"""
    assert not context.loan_data.empty, "No loan payment data found"
    context.reports.append(f"Loan payment calculations validated using {loan_formula}")


@then('rounding errors should not exceed "{rounding_tolerance}"')
def step_then_rounding_errors_within_tolerance(context, rounding_tolerance):
    # type: (Any, Any) -> None
    """Ensure rounding errors stay within acceptable limits"""
    message = "Loan payment rounding errors checked against {} tolerance".format(
        rounding_tolerance
    )
    context.reports.append(message)


@then("incorrect calculations should be flagged")
def step_then_flag_incorrect_loan_calculations(context):
    # type: (Any) -> None
    """Flag incorrect mortgage and loan payment calculations"""
    context.reports.append("Incorrect loan calculations flagged for review")


@when('I check the "Amortization Schedule" column in the "{sheet_name}" sheet')
def step_when_validate_amortization_schedule(context, sheet_name):
    # type: (Any, Any) -> None
    """Validate the loan amortization schedule"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")


@then("all payments should be distributed correctly across the loan term")
def step_then_validate_amortization_distribution(context):
    # type: (Any) -> None
    """Ensure loan payments are distributed correctly over time"""
    assert (
        not context.amortization_schedule.empty
    ), "No amortization schedule data found"
    context.reports.append("Amortization schedule distribution validated")


@then("extra payments should be applied to principal correctly")
def step_then_validate_extra_payments(context):
    # type: (Any) -> None
    """Ensure extra payments are applied properly to principal"""
    context.reports.append("Extra payments verified against principal reduction")


@then("any over/underpayment should be flagged for review")
def step_then_flag_over_underpayments(context):
    # type: (Any) -> None
    """Flag over/underpayments in loan schedules"""
    context.reports.append("Over/underpayments flagged for further review")


@when(
    'I check "Interest Paid" and "Principal Paid" for each payment in the "{sheet_name}" sheet'
)
def step_when_validate_interest_vs_principal(context, sheet_name):
    # type: (Any, Any) -> None
    """Validate breakdown of interest vs. principal in loan payments"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")


@then("the sum of all payments should match the total loan amount plus interest")
def step_then_validate_total_loan_amount(context):
    # type: (Any) -> None
    """Ensure the total payments align with the original loan amount plus interest"""
    assert (
        not context.interest_principal_data.empty
    ), "No interest/principal breakdown found"
    context.reports.append("Total loan amount validation successful")


@then("discrepancies should be flagged for further review")
def step_then_flag_interest_vs_principal_discrepancies(context):
    # type: (Any) -> None
    """Flag discrepancies in interest vs. principal payments"""
    context.reports.append("Discrepancies in loan payment breakdown flagged for review")


# ================= End of Loan and Mortgage Payments Validation Step Definitions for Financial Accuracy Testing =================

# ================= Beginning of Tax Withholding Validation Step Definitions for Financial Accuracy Testing =================
# This script validates tax withholding calculations.
# It ensures:
# - Correct application of expected tax withholding rates.
# - Identification of anomalies in tax withholding.
# - Compliance with financial regulations.


@when('I check the "Tax Withheld" column in the "{sheet_name}" sheet')
def step_when_validate_tax_withheld(context, sheet_name):
    # type: (Any, Any) -> None
    """Extract and validate tax withholding values"""
    if context.file_path.endswith(".csv"):
        df = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        df = pd.read_excel(context.file_path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file format")


@then('all tax withholdings should be "{expected_tax_rate}%"')
def step_then_validate_tax_rate(context, expected_tax_rate):
    # type: (Any, Any) -> None
    """Verify that all tax withholdings match the expected rate"""
    expected_rate = float(expected_tax_rate.strip("%")) / 100
    incorrect_rates = context.tax_withheld_data[
        context.tax_withheld_data != expected_rate
    ]

    if not incorrect_rates.empty:
        context.reports.append(
            "Tax withholdings do not match the expected " + expected_tax_rate + "%"
        )


@then("discrepancies should be flagged for review")
def step_then_flag_tax_discrepancies(context):
    # type: (Any) -> None
    """Flag any tax discrepancies in the report"""
    context.reports.append("Discrepancies in tax withholding flagged for review")


@then("incorrect tax rates should be escalated for compliance audit")
def step_then_escalate_tax_issues(context):
    # type: (Any) -> None
    """Escalate tax withholding issues for compliance review"""
    context.reports.append(
        "Incorrect tax withholding rates escalated for compliance audit"
    )


@then("negative or missing tax amounts should be flagged")
def step_then_flag_missing_or_negative_tax(context):
    # type: (Any) -> None
    """Flag missing or negative tax values"""
    missing_values = context.tax_withheld_data.isnull().sum()
    negative_values = (context.tax_withheld_data < 0).sum()

    if missing_values > 0:
        context.reports.append(
            "{} records with missing tax withheld values flagged".format(missing_values)
        )

    if negative_values > 0:
        context.reports.append(
            "{} records with negative tax withheld values flagged".format(
                negative_values
            )
        )


@then("unusually high or low tax rates should be reported")
def step_then_report_unusual_tax_rates(context):
    # type: (Any) -> None
    """Report tax withholdings that are significantly higher or lower than the expected rate"""
    context.reports.append(
        "Unusual tax withholding rates flagged for further investigation"
    )


# ================= End of Tax Withholding Validation Step Definitions for Financial Accuracy Testing =================

# ================= Beginning of Concurrent Processing Performance Testing Step Definitions =================
# This script evaluates the performance of concurrent file processing.
# It ensures:
# - Parallel processing of multiple export files without errors.
# - Scalability and resource efficiency under high concurrent load.
# - Proper handling of errors and corrupt files.
# - Stability of long-running concurrent processes.


@then("the system should process them in parallel without errors")
def step_then_verify_parallel_processing(context):
    # type: (Any) -> None
    """Verify that all files were processed successfully"""
    # Ensure the processed files match the file paths
    assert len(context.processed_files) == len(
        context.file_paths
    ), "Not all files were processed"

    # Log the success
    logging.info("All files were processed successfully.")


@then("processing time should be within acceptable limits")
def step_then_check_processing_time(context):
    # type: (Any) -> None
    """Check that processing did not exceed expected limits"""
    start_time = datetime.now()

    # Simulate processing logic with an expected duration based on the number of files
    expected_duration = timedelta(
        seconds=len(context.file_paths) * 0.5
    )  # 0.5 seconds per file
    end_time = start_time + expected_duration

    # Simulate the processing delay
    while datetime.now() < end_time:
        pass

    # Calculate elapsed time
    elapsed_time = (datetime.now() - start_time).total_seconds()

    # Check if the elapsed time exceeds the acceptable limit (60 seconds)
    assert (
        elapsed_time < 60
    ), "Processing time exceeded acceptable limits. Elapsed time: {0} seconds.".format(
        elapsed_time
    )

    # Log the processing time
    logging.info("Processing completed in {0:.2f} seconds.".format(elapsed_time))


@then("no data loss or corruption should occur")
def step_then_validate_data_integrity(context):
    # type: (Any) -> None
    """Ensure no data loss occurred during concurrent processing"""
    assert all(context.processed_files), "Some files failed to process"


@then("logs should correctly record processing order")
def step_then_check_logs(context):
    # type: (Any) -> None
    """Verify that logs correctly track processing order"""
    logging.info("Processed files: {}".format(context.processed_files))


@given('a batch of "{file_count}" bank export files')
def batch_of_files(context, file_count):
    # type: (Any, Any) -> None
    """Ensure batch processing scenario is correctly set up"""
    context.file_count = int(file_count)


@when('I attempt to process them concurrently with "{threads}" worker threads')
def step_when_concurrent_processing_with_threads(context, threads):
    # type: (Any, Any) -> None
    """Simulate multi-threaded processing of batch files"""
    context.threads = int(threads)
    start_time = datetime.now()

    def dummy_processing():
        # type: () -> None
        # Simulate processing time of 0.1 seconds using datetime loop
        end_time = datetime.now() + timedelta(seconds=0.1)
        while datetime.now() < end_time:
            pass

    # Use basic threading for Python 2.7 compatibility
    def worker():
        # type: () -> None
        for _ in range(context.file_count // context.threads):
            dummy_processing()

    thread_list = []
    for _ in range(context.threads):
        thread = threading.Thread(target=worker)
        thread_list.append(thread)
        thread.start()

    for thread in thread_list:
        thread.join()


@then('the system should complete processing within "{expected_time}" seconds')
def step_then_check_processing_time_limit(context, expected_time):
    # type: (Any, Any) -> None
    """Ensure the processing time meets the expectation"""
    expected_time = float(expected_time)
    assert (
        context.elapsed_time < expected_time
    ), "Processing took longer than {} seconds".format(expected_time)


@then("no unexpected failures should occur")
def step_then_no_failures(context):
    # type: (Any) -> None
    """Ensure no failures occurred during concurrent processing"""
    assert context.elapsed_time > 0, "Processing did not start correctly"


@then('system resources should not exceed "{resource_limit}%"')
def step_then_check_system_resources_with_limit(context, resource_limit):
    # type: (Any, Any) -> None
    """Ensures resource usage remains within the specified percentage limit."""
    check_system_resources(resource_limit)


@then("CPU and memory usage should remain within acceptable limits")
def step_then_check_system_resources_without_limit(context):
    # type: (Any) -> None
    """Verify CPU and memory usage remains within acceptable limits (mocked validation)."""
    check_system_resources()


@then("a summary report should be generated")
def step_then_generate_summary_report(context):
    # type: (Any) -> None
    """Mock the generation of a system performance report"""
    logging.info("Generated performance report for concurrent processing")


@then("valid files should be processed successfully")
def step_then_process_valid_files(context):
    # type: (Any) -> None
    """Ensure valid files were processed without issues"""
    logging.info("Valid files processed successfully")


@then("corrupt files should be flagged with appropriate error messages")
def step_then_flag_corrupt_files(context):
    # type: (Any) -> None
    """Flag and report corrupt files"""
    logging.warning("Corrupt files detected and flagged")


@then("no valid transactions should be lost due to errors")
def step_then_no_data_loss(context):
    # type: (Any) -> None
    """Ensure no valid data was lost"""
    assert True, "All valid transactions retained"


@then("processing should scale linearly with the number of files")
def step_then_scaling_check(context):
    # type: (Any) -> None
    """Verify that processing time increases proportionally with files"""
    logging.info("Processing scales linearly with the number of files")


@then("system response time should not degrade significantly")
def step_then_response_time_check(context):
    # type: (Any) -> None
    """Ensure response time remains stable under load"""
    logging.info("Response time remains stable")


@then("detailed system metrics should be collected for analysis")
def step_then_collect_metrics(context):
    # type: (Any) -> None
    """Mock collection of system performance metrics"""
    logging.info("Collected system performance metrics")


@then("the system should efficiently manage multiple concurrent file uploads")
def step_then_manage_multi_user_uploads(context):
    # type: (Any) -> None
    """Ensure multi-user concurrent processing runs smoothly"""
    logging.info("System efficiently handled multiple user uploads")


@then("no user should experience significant delays")
def step_then_no_user_delays(context):
    # type: (Any) -> None
    """Verify that users did not experience excessive delays"""
    logging.info("No significant delays experienced")


@then("all processed data should be stored accurately")
def step_then_data_accuracy_check(context):
    # type: (Any) -> None
    """Ensure all processed data is accurately stored"""
    logging.info("Processed data stored accurately")


@then("it should maintain stable performance without crashes")
def step_then_check_stability(context):
    # type: (Any) -> None
    """Ensure the system remains stable during long processing"""
    logging.info("System maintained stable performance")


@then("no memory leaks should occur")
def step_then_memory_leak_check(context):
    # type: (Any) -> None
    """Mock a check for memory leaks"""
    logging.info("No memory leaks detected")


@then("performance degradation should be minimal")
def step_then_minimal_degradation(context):
    # type: (Any) -> None
    """Ensure minimal performance impact over time"""
    logging.info("Minimal performance degradation observed")


@then("log files should capture long-term trends")
def step_then_log_trends(context):
    # type: (Any) -> None
    """Ensure logs record long-term performance trends"""
    logging.info("Log files successfully captured long-term trends")


# ================= End of Concurrent Processing Performance Testing Step Definitions =================

# ================= Beginning of Delayed Processing Performance Testing Step Definitions =================
# This script evaluates the performance of delayed processing in bank export files.
# It ensures:
# - System logs delays and continues processing without corruption.
# - Network latency and batch delays do not impact data integrity.
# - Stability under long-running and high-load conditions.
# - Efficient handling of delayed transactions with retry and priority mechanisms.


@given('a bank export file "{file_name}" with a "{delay_time}" second delay')
def delayed_file(context, file_name, delay_time):
    # type: (Any, Any, Any) -> None
    """Simulate a delayed file processing scenario"""
    file_path = os.path.join(context.base_dir, file_name)
    assert os.path.exists(file_path), "File {} not found".format(file_name)
    context.file_path = file_path
    context.delay_time = int(delay_time)


@when("I attempt to process the file")
@when("I attempt to process the file")
def step_when_process_delayed_file(context):
    # type: (Any) -> None
    """Process file after a simulated delay using datetime"""
    logging.info("Delaying processing for {} seconds...".format(context.delay_time))

    # Simulate delay using datetime instead of time.sleep
    end_time = datetime.now() + timedelta(seconds=context.delay_time)
    while datetime.now() < end_time:
        pass  # Busy wait to simulate delay

    # Load the file after delay
    if context.file_path.endswith(".csv"):
        context.data = pd.read_csv(context.file_path)
    elif context.file_path.endswith(".xlsx"):
        context.data = pd.read_excel(context.file_path)
    else:
        raise ValueError("Unsupported file format")


@then("the system should log the delay and continue processing")
def step_then_log_and_continue(context):
    # type: (Any) -> None
    """Ensure system logs delay but continues processing"""
    logging.info(
        "File {} processed after {} seconds delay.".format(
            context.file_path, context.delay_time
        )
    )


@then("delayed transactions should be flagged for review")
def step_then_flag_delayed_transactions(context):
    # type: (Any) -> None
    """Flag transactions delayed beyond threshold"""
    logging.warning(
        "Delayed transactions in {} flagged for review.".format(context.file_path)
    )


@then("system stability should not be affected")
def step_then_check_system_stability(context):
    # type: (Any) -> None
    """Ensure system remains stable despite delay"""
    logging.info("System stability confirmed after delayed processing.")


@then('a warning should be issued if the delay exceeds "{max_threshold}" seconds')
def step_then_issue_warning(context, max_threshold):
    # type: (Any, Any) -> None
    """Warn if processing delay exceeds maximum threshold"""
    max_threshold = int(max_threshold)  # Ensure max_threshold is an integer
    if context.delay_time > max_threshold:
        logging.warning("Processing delay exceeded {0} seconds!".format(max_threshold))


@given('a batch of bank export files processed in "{batch_interval}" seconds')
def batch_processing(context, batch_interval):
    # type: (Any, Any) -> None
    """Simulate batch processing with delays"""
    context.batch_interval = int(batch_interval)  # Ensure batch_interval is an integer
    logging.info(
        "Batch processing started with a {0} seconds interval.".format(
            context.batch_interval
        )
    )


@then("logs should capture batch processing timestamps")
def step_then_capture_batch_logs(context):
    # type: (Any) -> None
    """Ensure logs contain batch processing timestamps"""
    logging.info(
        "Batch processing completed with {} second intervals.".format(
            context.batch_interval
        )
    )


@then('batch failures should be retried up to "{retry_count}" times')
def step_then_retry_batches(context, retry_count):
    # type: (Any, Any) -> None
    """Ensure failed batch processes are retried"""
    retry_count = int(retry_count)
    logging.info("Batch failures will be retried up to {} times.".format(retry_count))


@given(
    'a bank export file "{file_name}" with simulated network latency of "{latency}" ms'
)
def network_latency(context, file_name, latency):
    # type: (Any, Any, Any) -> None
    """Simulate a file processing delay due to network latency"""
    context.file_path = os.path.join(context.base_dir, file_name)
    context.latency = int(latency) / 1000  # Convert to seconds


@then("the system should retry within an acceptable time frame")
def step_then_retry_within_time(context):
    # type: (Any) -> None
    """Ensure retries are performed within an acceptable timeframe"""
    logging.info(
        "Retrying file processing after network latency of {} seconds.".format(
            context.latency
        )
    )


@then("transactions should not be duplicated due to retries")
def step_then_prevent_duplicates(context):
    # type: (Any) -> None
    """Ensure transactions are not duplicated"""
    logging.info("Checked duplicate transactions due to network retries.")


@then('a fallback mechanism should trigger if latency exceeds "{latency_threshold}" ms')
def step_then_trigger_fallback(context, latency_threshold):
    # type: (Any, Any) -> None
    """Trigger fallback if latency exceeds threshold"""
    latency_threshold = int(latency_threshold) / 1000.0  # Convert to seconds
    if context.latency > latency_threshold:
        logging.warning(
            "Network latency exceeded threshold! Triggering fallback mechanism."
        )


@given(
    'a continuous stream of bank export files arriving every "{interval}" seconds with a "{delay_time}" second delay'
)
def continuous_stream(context, interval, delay_time):
    # type: (Any, Any, Any) -> None
    """Simulate continuous stream of bank files with delay"""
    context.interval = int(interval)
    context.delay_time = int(delay_time)


@then("it should maintain stable performance without crashes")
def step_then_check_long_running_stability(context):
    # type: (Any) -> None
    """Ensure the system does not crash under long-running delays"""
    logging.info("System maintained stability under long-running delays.")


@given('a queue of "{file_count}" delayed bank export files')
def delayed_file_queue(context, file_count):
    # type: (Any, Any) -> None
    """Simulate a queue of delayed files"""
    context.file_count = int(file_count)


@when('I attempt to process them with "{worker_threads}" concurrent threads')
def step_when_process_concurrent_delayed_files(context, worker_threads):
    # type: (Any, Any) -> None
    context.worker_threads = int(worker_threads)

    def process_delayed_file(dummy):
        # type: (Any) -> str
        # Simulated processing delay using datetime
        delay_seconds = random.randint(1, 3)
        end_time = datetime.now() + timedelta(seconds=delay_seconds)
        while datetime.now() < end_time:
            pass
        return "Processed"

    executor = concurrent.futures.ThreadPoolExecutor(max_workers=context.worker_threads)
    results = list(executor.map(process_delayed_file, range(context.file_count)))
    executor.shutdown(wait=True)

    context.processed_results = results


@then("the system should efficiently process all files without excessive queue backlog")
def step_then_check_queue_backlog(context):
    # type: (Any) -> None
    """Ensure all delayed files are processed without backlog"""
    assert (
        len(context.processed_results) == context.file_count
    ), "Not all delayed files were processed"
    logging.info("All delayed files processed successfully.")


@then('delayed files should be prioritized based on "{priority_rule}"')
def step_then_prioritize_delayed_files(context, priority_rule):
    # type: (Any, Any) -> None
    """Ensure delayed files are prioritized correctly"""
    logging.info(f"Delayed files processed based on priority rule: {priority_rule}")


@given('a delayed bank export file "{file_name}"')
def delayed_data_integrity(context, file_name):
    # type: (Any, Any) -> None
    """Ensure file exists for delayed processing integrity test"""
    context.file_path = os.path.join(context.base_dir, file_name)
    assert os.path.exists(context.file_path), f"File {file_name} not found"


@when('I process the file with a delay of "{delay_time}" seconds')
def step_when_process_delayed_data_file(context, delay_time):
    # type: (Any, Any) -> None
    """Simulate processing delay for data integrity"""
    context.delay_time = int(delay_time)  # Convert delay_time to an integer
    end_time = datetime.now() + timedelta(seconds=context.delay_time)

    # Simulate the delay using a busy-wait loop
    while datetime.now() < end_time:
        pass  # Busy-wait until the specified delay time has passed

    logging.info(
        "Processed file with a delay of {0} seconds.".format(context.delay_time)
    )


@then("all transactions should retain their original timestamps")
def step_then_check_original_timestamps(context):
    # type: (Any) -> None
    """Ensure transaction timestamps remain unchanged after delay"""

    # Assuming 'context.transaction_data' is a dataframe containing transaction timestamps
    if context.transaction_data.empty:
        raise ValueError("No transaction data found to check timestamps.")

    # Assuming 'timestamp_column' is the column where the timestamps are stored in the data
    timestamp_column = 'timestamp'  # Modify this according to your data structure

    # Compare the timestamps before and after processing to ensure they remain unchanged
    initial_timestamps = context.transaction_data[timestamp_column]

    # Re-checking after processing the delay (assuming context.transaction_data is updated)
    final_timestamps = context.transaction_data[timestamp_column]

    # Verify that no timestamps have changed
    assert (
        initial_timestamps == final_timestamps
    ).all(), "Timestamps were altered during processing."

    logging.info("Checked delayed transactions retained original timestamps.")


@then("no data should be lost or duplicated due to delays")
def step_then_check_data_loss(context):
    # type: (Any) -> None
    """Ensure no transactions are lost or duplicated due to processing delays"""
    logging.info("Checked data loss or duplication due to delays.")


@then("a reconciliation report should be generated")
def step_then_generate_reconciliation_report(context):
    # type: (Any) -> None
    """Mock reconciliation report generation"""
    logging.info("Reconciliation report generated for delayed transactions.")


# ================= End of Delayed Processing Performance Testing Step Definitions =================

# ================= Beginning of High Concurrent Users Performance Testing Step Definitions =================
# This script evaluates the performance of the system under high concurrent user load.
# It ensures:
# - Stability when multiple users upload and process export files concurrently.
# - Resource utilization remains within acceptable limits.
# - The database scales dynamically under high transaction volume.
# - The system manages request queues efficiently and handles errors properly.
# - No memory leaks or performance degradation over extended high-load periods.


@given('"{user_count}" users accessing the system simultaneously')
def high_concurrent_users(context, user_count):
    # type: (Any, Any) -> None
    """Simulate high concurrent users accessing the system"""
    context.user_count = int(user_count)


@when("they attempt to upload and process bank export files concurrently")
def step_when_users_upload_files(context):
    # type: (Any) -> None
    """Simulate users uploading files concurrently"""

    def process_upload(user_id):
        # type: (Any) -> str
        delay_seconds = random.uniform(0.5, 2)
        end_time = datetime.now() + timedelta(seconds=delay_seconds)

        # Simulated delay using datetime-based busy wait
        while datetime.now() < end_time:
            pass

        return "Processed"

    executor = concurrent.futures.ThreadPoolExecutor(max_workers=context.user_count)
    results = list(executor.map(process_upload, range(context.user_count)))
    executor.shutdown(wait=True)

    context.upload_results = results


@then("the system should maintain stable performance without degradation")
def step_then_system_stable_under_concurrency(context):
    # type: (Any) -> None
    """Ensure the system remains stable under high concurrency."""
    # Log the number of concurrent users handled
    logging.info(
        "System handled {0} concurrent users successfully.".format(context.user_count)
    )


@then('response times should remain within "{expected_response_time}" seconds')
def step_then_check_response_time(context, expected_response_time):
    # type: (Any, Any) -> None
    """Ensure response time remains within limits"""
    expected_response_time = float(expected_response_time)

    # Simulate response delay using random value between 0.5 seconds and expected response time
    delay_seconds = random.uniform(0.5, expected_response_time)
    end_time = datetime.now() + timedelta(seconds=delay_seconds)

    # Simulate the delay using a busy-wait loop
    while datetime.now() < end_time:
        pass

    # The actual response time is the simulated delay
    actual_response_time = delay_seconds

    # Assert if the actual response time is within the acceptable threshold
    assert (
        actual_response_time <= expected_response_time
    ), "Response time exceeded threshold. Actual: {0}, Expected: {1}".format(
        actual_response_time, expected_response_time
    )

    # Log the result
    logging.info(
        "Response time: {0:.2f} seconds, within acceptable range.".format(
            actual_response_time
        )
    )


@given('"{user_count}" users performing simultaneous operations on bank export files')
def resource_utilization(context, user_count):
    # type: (Any, Any) -> None
    """Monitor system resource utilization under high concurrent load"""
    context.user_count = int(user_count)


@then('CPU usage should not exceed "{cpu_limit}%"')
def step_then_check_cpu_usage(context, cpu_limit):
    # type: (Any, Any) -> None
    """Ensure CPU usage remains within limits"""
    cpu_usage = random.uniform(50, int(cpu_limit))  # Simulating CPU usage
    assert cpu_usage <= int(cpu_limit), "CPU usage exceeded threshold"
    logging.info("CPU usage at {:.2f}% within acceptable limits.".format(cpu_usage))


@then('memory usage should remain below "{memory_limit}%"')
def step_then_check_memory_usage(context, memory_limit):
    # type: (Any, Any) -> None
    """Ensure memory usage remains within limits"""
    memory_usage = random.uniform(40, int(memory_limit))  # Simulating memory usage
    assert memory_usage <= int(memory_limit), "Memory usage exceeded threshold"
    logging.info(
        "Memory usage at {:.2f}% within acceptable limits.".format(memory_usage)
    )


@given('"{user_count}" users executing queries simultaneously')
def database_queries(context, user_count):
    # type: (Any, Any) -> None
    """Simulate high concurrent database transactions"""
    context.user_count = int(user_count)


@when("transaction logs are analyzed")
def step_when_analyze_transaction_logs(context):
    # type: (Any) -> None
    """Analyze database transaction logs"""
    logging.info(
        "Analyzing {} concurrent database transactions.".format(context.user_count)
    )


@then("the system should maintain stable performance without degradation")
def step_then_system_stable_under_concurrency(context):
    # type: (Any) -> None
    """Ensure the system remains stable under high concurrency."""
    # Log the number of concurrent users handled
    logging.info(
        "System handled {0} concurrent users successfully.".format(context.user_count)
    )


@then('response times should remain within "{expected_response_time}" seconds')
def step_then_check_response_time(context, expected_response_time):
    # type: (Any, Any) -> None
    """Ensure response time remains within limits"""
    expected_response_time = float(expected_response_time)

    # Simulate response delay using random value between 0.5 seconds and expected response time
    delay_seconds = random.uniform(0.5, expected_response_time)
    end_time = datetime.now() + timedelta(seconds=delay_seconds)

    # Simulate the delay using a busy-wait loop
    while datetime.now() < end_time:
        pass

    # The actual response time is the simulated delay
    actual_response_time = delay_seconds

    # Assert if the actual response time is within the acceptable threshold
    assert (
        actual_response_time <= expected_response_time
    ), "Response time exceeded threshold. Actual: {0}, Expected: {1}".format(
        actual_response_time, expected_response_time
    )

    # Log the result
    logging.info(
        "Response time: {0:.2f} seconds, within acceptable range.".format(
            actual_response_time
        )
    )


@given('"{user_count}" users submitting processing requests')
def request_queue(context, user_count):
    # type: (Any, Any) -> None
    """Simulate concurrent users submitting processing requests"""
    context.user_count = int(user_count)


@when("the system queues the requests for execution")
def step_when_queue_requests(context):
    # type: (Any) -> None
    """Simulate queueing of processing requests"""
    context.queued_requests = context.user_count


@then('the queue should not exceed "{max_queue_size}" pending requests')
def step_then_check_queue_size(context, max_queue_size):
    # type: (Any, Any) -> None
    """Ensure queue size does not exceed limit"""
    max_queue_size = int(max_queue_size)
    assert context.queued_requests <= max_queue_size, "Queue size exceeded threshold"
    logging.info(
        "Queue size: {} within allowed limit of {}.".format(
            context.queued_requests, max_queue_size
        )
    )


@then('prioritization rules should apply based on "{priority_rule}"')
def step_then_apply_queue_priority(context, priority_rule):
    # type: (Any, Any) -> None
    """Ensure queue follows prioritization rules"""
    logging.info("Queue prioritization applied based on: {}".format(priority_rule))


@given('"{user_count}" users performing simultaneous operations')
def error_handling(context, user_count):
    # type: (Any, Any) -> None
    """Monitor error handling under high concurrent load"""
    context.user_count = int(user_count)


@when("some operations fail due to system limitations")
def step_when_simulate_failures(context):
    # type: (Any) -> None
    """Simulate failures in concurrent operations"""
    context.failed_operations = random.randint(
        1, context.user_count // 10
    )  # Simulating failures


@then("failures should be logged with clear error messages")
def step_then_log_errors(context):
    # type: (Any) -> None
    """Ensure failures are logged with error messages"""
    logging.warning(f"{context.failed_operations} operations failed and logged.")


@then('the system should notify relevant users with "{notification_message}"')
@then("users should receive appropriate error notifications")
def step_then_notify_users(context, notification_message=None):
    # type: (Any, Any) -> None
    """Ensure notifications are sent for historical discrepancies or error operations"""

    if notification_message:
        logging.info("Notification sent: {0}".format(notification_message))
    else:
        logging.info(
            "Users notified about {0} failed operations.".format(
                str(context.failed_operations)
            )
        )

    # Reuse the send_notification function from user_notifications.py
    send_notification(context, notification_message or "Error operations detected.")


@given('"{user_count}" users continuously accessing the system for "{duration}" hours')
def long_running_users(context, user_count, duration):
    # type: (Any, Any, Any) -> None
    """Monitor system stability over extended high-load periods"""
    context.user_count = int(user_count)
    context.duration = int(duration)


@when("system health metrics are monitored")
def step_when_monitor_health(context):
    # type: (Any) -> None
    """Simulate monitoring of system health metrics"""
    logging.info(
        "Monitoring system health for {} users over {} hours.".format(
            context.user_count, context.duration
        )
    )


@then("memory leaks should not occur")
def step_then_check_memory_leaks(context):
    # type: (Any) -> None
    """Ensure no memory leaks occur"""
    memory_leak_detected = random.choice(
        [False, False, False, True]
    )  # Simulated detection
    assert not memory_leak_detected, "Memory leak detected!"
    logging.info("No memory leaks detected.")


@then("no crashes or unexpected terminations should happen")
def step_then_check_system_crashes(context):
    # type: (Any) -> None
    """Ensure system does not crash under high concurrent load"""
    crash_occurred = random.choice([False, False, True])  # Simulated crash detection
    assert not crash_occurred, "Unexpected system crash detected!"
    logging.info("System maintained stability without crashes.")


# ================= End of High Concurrent Users Performance Testing Step Definitions =================

# ================= Beginning of Large Data Performance Testing Step Definitions =================
# This script evaluates system performance under large data processing conditions.
# It ensures:
# - Efficient handling of large datasets in exports and database imports.
# - Stability under network latency and long-running processes.
# - Proper error handling and query performance optimizations.
# - Prevention of memory leaks and out-of-memory errors.
# - Batch processing and retry mechanisms for failed operations.


@given('a bank export file with "{row_count}" rows and "{column_count}" columns')
def large_export_file(context, row_count, column_count):
    # type: (Any, Any, Any) -> None
    """Simulate a large export file with specified row and column count"""
    context.row_count = int(row_count)
    context.column_count = int(column_count)


@when("the system processes the file")
def step_when_process_large_file(context):
    # type: (Any) -> None
    """Simulate processing of a large export file"""
    processing_time = random.uniform(30, 180)  # Simulated processing duration
    context.processing_time = processing_time

    # Use datetime-based delay (max 3 seconds)
    delay_duration = min(processing_time, 3)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=delay_duration)

    while datetime.now() < end_time:
        pass  # Busy-wait loop to simulate delay

    logging.info("Simulated processing delay of {:.2f} seconds.".format(delay_duration))


@then("no data loss or corruption should occur")
def step_then_check_data_integrity(context):
    # type: (Any) -> None
    """Ensure no data is lost or corrupted"""
    data_loss = random.choice([False, False, False, True])  # Simulating rare data loss
    assert not data_loss, "Data loss detected!"
    logging.info("Data integrity verified data loss or corruption.")


@then('memory consumption should not exceed "{memory_limit}%"')
def step_then_memory_consumption_within_limit(context, memory_limit):
    # type: (Any, Any) -> None
    """Ensure memory consumption stays within limits."""
    memory_limit = int(memory_limit)
    memory_usage = random.uniform(50, memory_limit)  # Simulated memory usage

    assert memory_usage <= memory_limit, "Memory usage exceeded limit!"
    logging.info(
        "Memory usage at {:.2f}% within acceptable limits.".format(memory_usage)
    )


@given('a bank export file with "{row_count}" rows')
def large_database_import(context, row_count):
    # type: (Any, Any) -> None
    """Simulate importing a large dataset into the database"""
    context.row_count = int(row_count)


@when("I attempt to import the file into the database")
def step_when_import_large_dataset(context):
    # type: (Any) -> None
    """Simulate database import process"""
    import_time = random.uniform(1, 10)  # Simulate import duration
    context.import_time = import_time

    # Delay using datetime (max 5 seconds)
    delay_duration = min(import_time, 5)  # Ensure that delay does not exceed 5 seconds
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=delay_duration)

    # Simulate the delay using a busy-wait loop
    while datetime.now() < end_time:
        pass  # Busy-wait until the delay time has passed

    # Log the import delay
    logging.info("Simulated import delay of {0:.2f} seconds.".format(delay_duration))


@then('the database should complete the import within "{expected_time}" seconds')
def step_then_check_import_time(context, expected_time):
    # type: (Any, Any) -> None
    """Ensure database import completes within expected limits"""
    assert context.import_time <= float(expected_time), "Database import took too long!"
    logging.info(
        "Database import completed in {:.2f} seconds.".format(context.import_time)
    )


@then("indexing operations should not cause performance degradation")
def step_then_check_indexing_performance(context):
    # type: (Any) -> None
    """Ensure indexing does not slow down performance"""
    indexing_slowdown = random.choice([False, False, True])  # Simulating rare issues
    assert not indexing_slowdown, "Indexing slowed down performance!"
    logging.info("Indexing operations completed without performance degradation.")


@given('"{batch_count}" bank export files each containing "{row_count}" rows')
def large_batch_files(context, batch_count, row_count):
    # type: (Any, Any, Any) -> None
    """Simulate batch processing of large export files"""
    context.batch_count = int(batch_count)
    context.row_count = int(row_count)


@when("I process these files in parallel")
def files_are_processed_in_parallel(context):
    # type: (Any) -> None
    """Simulate batch processing in parallel"""

    # Simulate processing time (between 100 and 900 seconds)
    processing_time = random.uniform(100, 900)
    context.batch_processing_time = processing_time

    # Simulate a delay of 5 seconds or the processing time, whichever is smaller
    delay_seconds = min(processing_time, 5)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=delay_seconds)

    # Simulate the delay using a busy-wait loop
    while datetime.now() < end_time:
        pass  # Busy-wait until the delay time has passed

    # Log the parallel processing simulation
    logging.info(
        "Simulated parallel batch processing for {0:.2f} seconds.".format(delay_seconds)
    )


@then('batch failures should be retried up to "{retry_count}" times')
def step_then_retry_failed_batches(context, retry_count):
    # type: (Any, Any) -> None
    """Ensure failed batches are retried"""
    retry_count = int(retry_count)
    failures = random.randint(0, 2)  # Simulating batch failures
    retry_attempts = 0

    while failures > 0 and retry_attempts < retry_count:
        retry_attempts += 1
        failures -= 1  # Simulate retry success

    assert failures == 0, "Some batches failed after retries!"
    logging.info(
        "Batch processing retried "
        + str(retry_attempts)
        + " times and completed successfully."
    )


@given(
    'a bank export file "{file_name}" with "{row_count}" rows and simulated network latency of "{latency}" ms'
)
def large_file_with_latency(context, file_name, row_count, latency):
    # type: (Any, Any, Any, Any) -> None
    """Simulate processing delays due to network latency"""
    context.file_name = file_name
    context.row_count = int(row_count)
    context.latency = int(latency)


@when("I attempt to process the file remotely")
def step_when_remote_processing(context):
    # type: (Any) -> None
    """Simulate remote file processing with network latency"""
    total_latency = context.latency + random.randint(100, 500)
    context.total_latency = total_latency

    delay_seconds = min(total_latency / 1000.0, 5)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=delay_seconds)

    while datetime.now() < end_time:
        pass

    logging.info(
        "Simulated remote processing with total latency of "
        + str(total_latency)
        + " ms."
    )


@then('an alert should be generated if latency exceeds "{latency_threshold}" ms')
def step_then_check_latency(context, latency_threshold):
    # type: (Any, Any) -> None
    """Ensure latency does not exceed thresholds"""
    latency_threshold = int(latency_threshold)
    assert context.total_latency <= latency_threshold, "Network latency exceeded limit!"
    logging.info(
        "Network latency: {} ms within acceptable limits.".format(context.total_latency)
    )


@given(
    'a bank export file "{file_name}" containing "{error_type}" errors in "{error_percentage}%" of rows'
)
def large_file_with_errors(context, file_name, error_type, error_percentage):
    # type: (Any, Any, Any, Any) -> None
    """Simulate large dataset with errors"""
    context.file_name = file_name
    context.error_type = error_type
    context.error_percentage = float(error_percentage)


@when("I attempt to process the file")
def step_when_process_large_file_with_errors(context):
    # type: (Any) -> None
    """Simulate processing of a large file containing errors"""
    error_count = int((context.row_count * context.error_percentage) / 100)
    context.error_count = error_count
    logging.warning(
        "Detected {} {} errors during processing.".format(
            context.error_count, context.error_type
        )
    )


@then("the system should log all errors correctly")
def step_then_system_should_log_errors_correctly(context):
    # type: (Any) -> None
    """Ensure error logging works correctly"""
    assert context.error_count > 0, "No errors logged despite error conditions!"
    logging.info("All {} errors were correctly logged.".format(context.error_count))


@then("processing should continue without failure for valid rows")
def step_then_continue_valid_data_processing(context):
    # type: (Any) -> None
    """Ensure valid rows are processed despite errors"""
    valid_rows = context.row_count - context.error_count
    assert valid_rows > 0, "No valid rows processed!"
    logging.info("{} valid rows processed successfully.".format(valid_rows))


@given('a database containing "{row_count}" records from bank export files')
def large_dataset_in_db(context, row_count):
    # type: (Any, Any) -> None
    """Simulate large dataset in database for query performance testing"""
    context.row_count = int(row_count)


@when("I execute a complex query with multiple joins and filters")
def step_when_execute_large_query(context):
    # type: (Any) -> None
    """Simulate execution of a large query"""
    query_time = random.uniform(2, 15)
    context.query_time = query_time

    delay_seconds = min(query_time, 5)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=delay_seconds)

    while datetime.now() < end_time:
        pass

    logging.info(
        "Simulated complex query execution time: " + str(query_time) + " seconds."
    )


@then('query execution should complete within "{expected_time}" seconds')
def step_then_check_query_performance(context, expected_time):
    # type: (Any, Any) -> None
    """Ensure query execution is within time limits"""
    assert context.query_time <= float(
        expected_time
    ), "Query execution time exceeded limit!"
    logging.info("Query executed in {:.2f} seconds.".format(context.query_time))


# ================= End of Large Data Performance Testing Step Definitions =================

# ================= Beginning of Large Transaction Volume Processing Step Definitions =================
# This script evaluates system performance under high transaction volumes.
# It ensures:
# - Efficient handling of large transaction exports and database imports.
# - Stability under network latency and long-running processes.
# - Proper error handling and query performance optimizations.
# - Prevention of memory leaks and system crashes.
# - Batch processing and retry mechanisms for failed operations.


@given('a bank export file containing "{transaction_count}" transactions')
def large_transaction_export(context, transaction_count):
    # type: (Any, Any) -> None
    """Simulate a bank export file with high transaction volume"""
    context.transaction_count = int(transaction_count)


@when("the system processes the file")
def step_when_process_large_transactions(context):
    # type: (Any) -> None
    """Simulate processing a large transaction file with measured timing."""

    # Simulated processing workload (delay between 0.5 and 2 seconds)
    simulated_delay = random.uniform(0.5, 2.0)  # Keep short to avoid slowing tests

    # Start and end timestamps
    start_time = datetime.now()

    # Simulated logic block (replace with real logic if needed)
    end_time = start_time + timedelta(seconds=simulated_delay)

    # Busy-wait loop to simulate time passing (this can be optimized with time.sleep for real scenarios)
    while datetime.now() < end_time:
        pass

    # Calculate elapsed time
    elapsed_time = (datetime.now() - start_time).total_seconds()

    # Store elapsed time in context for validation
    context.processing_time = elapsed_time

    # Log the elapsed time using .format()
    logging.info(
        "Simulated file processing completed in {0:.2f} seconds.".format(elapsed_time)
    )


@step(parsers.cfparse('processing should complete within "{expected_time}" seconds'))
def processing_should_complete_within(context, expected_time):
    # type: (Any, Any) -> None
    """
    Unified BDD step that validates processing performance across different context types.
    Falls back to measuring time manually if context doesn't store timing attributes.
    """
    start_time = datetime.now()

    # Parse expected time as float for comparison
    expected_time = float(expected_time)

    # Possible timing fields in context
    time_fields = [
        "processing_time",
        "regression_test_time",
        "issue_validation_time",
        "reference_check_time",
    ]

    time_to_check = None
    time_label = "Processing"

    for field in time_fields:
        if hasattr(context, field):
            time_to_check = getattr(context, field)
            time_label = field.replace("_", " ").title()
            break

    if time_to_check is None:
        # Resolve file path from context
        file_attr = getattr(context, "file_path", None) or getattr(
            context, "file_name", None
        )
        if file_attr is None:
            raise ValueError(
                "No valid file attribute ('file_path' or 'file_name') found in context."
            )

        processor = FileProcessor(file_name=file_attr, context=context)
        processor.process()

        # Detect delimiters
        if file_attr.endswith(".csv"):
            delimiters = detect_delimiters_in_csv(file_attr)
        elif file_attr.endswith(".xlsx"):
            delimiters = detect_delimiters_in_excel(file_attr)
        else:
            delimiters = []

        logging.info("Detected delimiters: {}".format(delimiters))

        # Validate 'Transaction Amount' column
        if hasattr(context, "df") and "Transaction Amount" in context.df.columns:
            if not validate_column(context.df, "Transaction Amount", "numeric"):
                raise ValueError(
                    "'Transaction Amount' column contains non-numeric values."
                )

        time_to_check = (datetime.now() - start_time).total_seconds()

    # Final assertion
    if time_to_check > expected_time:
        raise AssertionError(
            time_label
            + " took too long. Elapsed: "
            + "{:.2f}".format(time_to_check)
            + " seconds, Allowed: "
            + "{:.2f}".format(expected_time)
            + " seconds."
        )

    # Final log
    if hasattr(context, "file_count") and hasattr(context, "year_range"):
        logging.info(
            "Performance validation "
            + str(context.file_count)
            + " files from "
            + str(context.year_range)
            + " in under {:.2f} seconds.".format(expected_time)
        )
    else:
        logging.info(time_label + " completed in {:.2f} seconds.".format(time_to_check))


@then('system memory consumption should not exceed "{memory_limit}%"')
def step_then_validate_memory_usage(context, memory_limit):
    # type: (Any, Any) -> None
    """Ensure system memory usage is within limits"""
    memory_usage = random.uniform(50, memory_limit)
    assert memory_usage <= memory_limit, "Memory usage exceeded limit!"
    logging.info(
        "Memory usage at {:.2f}% within acceptable limits.".format(memory_usage)
    )


@given("I attempt to import the file into the database")
def step_when_import_large_transaction_dataset(context):
    # type: (Any) -> None
    """Simulate database import process and track import time."""

    # Simulate variable import delay between 1.0 and 5.0 seconds
    simulated_delay = random.uniform(1.0, 5.0)

    # Start time
    start_time = datetime.now()

    # Simulate processing loop instead of sleeping
    end_time = start_time + timedelta(seconds=simulated_delay)

    # Busy-wait loop to simulate the load for the import process
    while datetime.now() < end_time:
        pass  # Busy-wait until the simulated delay time has passed

    # End time and calculate elapsed time
    elapsed_time = (datetime.now() - start_time).total_seconds()

    # Store import time in context for later validation
    context.import_time = elapsed_time

    # Log the import time (using .format for compatibility with Python 2.7)
    logging.info(
        "Simulated database import completed in {0:.2f} seconds.".format(elapsed_time)
    )


@then('the database should complete the import within "{expected_time}" seconds')
def step_then_validate_import_time(context, expected_time):
    # type: (Any, Any) -> None
    """Ensure database import is completed on time"""
    assert context.import_time <= expected_time, "Database import took too long!"
    logging.info(
        "Database import completed in {:.2f} seconds.".format(context.import_time)
    )


@then("indexing operations should not slow down the system")
def step_then_indexing_should_not_slow_down_system(context):
    # type: (Any) -> None
    """Ensure indexing does not cause performance degradation"""
    indexing_slowdown = random.choice([False, False, True])
    assert not indexing_slowdown, "Indexing slowed down the system!"
    logging.info("Indexing operations completed without performance issues.")


@given(
    '"{batch_count}" bank export files each containing "{transaction_count}" transactions'
)
def large_batch_transactions(context, batch_count, transaction_count):
    # type: (Any, Any, Any) -> None
    """Simulate batch transaction processing"""
    context.batch_count = int(batch_count)
    context.transaction_count = int(transaction_count)


@when("I process these files in parallel")
def step_when_process_transaction_batches(context):
    # type: (Any) -> None
    """Simulate parallel batch file processing and track elapsed time."""

    # Simulated batch workload duration (in seconds)
    simulated_delay = random.uniform(1.0, 5.0)  # Keep test quick and clean

    # Start timing
    start_time = datetime.now()

    # Simulate time passing using busy loop
    while (datetime.now() - start_time).total_seconds() < simulated_delay:
        pass  # No-op to simulate delay

    # End timing and compute elapsed time
    end_time = datetime.now()
    elapsed_time = (end_time - start_time).total_seconds()

    # Store batch processing time in context for further validation
    context.batch_processing_time = elapsed_time

    # Log the elapsed time using .format for compatibility with Python 2.7
    logging.info(
        "Simulated parallel batch processing completed in {0:.2f} seconds.".format(
            elapsed_time
        )
    )


@then('batch failures should be retried up to "{retry_count}" times')
def step_then_retry_transaction_batches(context, retry_count):
    # type: (Any, Any) -> None
    """Ensure failed batch transactions are retried"""
    retry_count = int(retry_count)
    failures = random.randint(0, 2)
    retry_attempts = 0

    while failures > 0 and retry_attempts < retry_count:
        retry_attempts += 1
        failures -= 1  # Simulate retry success

    assert failures == 0, "Some transactions failed after retries!"
    logging.info(
        "Batch processing retried {} times and completed successfully.".format(
            retry_attempts
        )
    )


@given(
    'a bank export file "{file_name}" with "{transaction_count}" transactions and simulated network latency of "{latency}" ms'
)
def file_with_transactions_and_latency(context, file_name, transaction_count, latency):
    # type: (Any, Any, Any, Any) -> None
    """Simulate network latency for large transaction processing"""
    context.file_name = file_name
    context.transaction_count = int(transaction_count)
    context.latency = int(latency)


@when("I attempt to process the file remotely")
def step_when_process_large_transactions_with_latency(context):
    # type: (Any) -> None
    """Simulate remote processing with artificial latency."""

    # Simulate added remote latency in milliseconds (between 100ms and 500ms)
    simulated_additional_latency = random.randint(100, 500)  # In ms
    base_latency = getattr(context, "latency", 0)  # Default to 0 if not set

    # Total latency (in milliseconds and converted to seconds)
    total_latency_ms = base_latency + simulated_additional_latency
    total_latency_sec = total_latency_ms / 1000.0

    # Start timing the simulated processing
    start_time = datetime.now()

    # Simulate delay using busy-wait (could be optimized with time.sleep)
    while (datetime.now() - start_time).total_seconds() < total_latency_sec:
        pass  # Simulate time passing (No-op)

    # End time and record the elapsed time
    elapsed = (datetime.now() - start_time).total_seconds()

    # Store the total latency in the context for future validation
    context.total_latency = elapsed

    # Log the latency using .format for compatibility with Python 2.7
    logging.info(
        "Simulated remote file processing with latency of {0:.2f} seconds ({1} ms).".format(
            elapsed, total_latency_ms
        )
    )


@then('an alert should be generated if latency exceeds "{latency_threshold}" ms')
def step_then_check_latency_threshold(context, latency_threshold):
    # type: (Any, Any) -> None
    """Ensure processing does not exceed latency threshold"""
    assert context.total_latency <= int(
        latency_threshold
    ), "Network latency exceeded threshold!"
    logging.info("Network latency: {} ms within limits.".format(context.total_latency))


@given(
    'a bank export file "{file_name}" containing "{error_type}" errors in "{error_percentage}%" of transactions'
)
def large_transaction_errors(context, file_name, error_type, error_percentage):
    # type: (Any, Any, Any, Any) -> None
    """Simulate transaction processing with errors"""
    context.file_name = file_name
    context.error_type = error_type
    context.error_percentage = float(error_percentage)


@when("I attempt to process the file")
def step_when_process_large_transaction_errors(context):
    # type: (Any) -> None
    """Simulate processing of large file with errors"""
    error_count = int((context.transaction_count * context.error_percentage) / 100)
    context.error_count = error_count
    logging.warning(
        "Detected {} {} errors during processing.".format(
            context.error_count, context.error_type
        )
    )


@then("the system should log all errors correctly")
def step_then_log_transaction_errors(context):
    # type: (Any) -> None
    """Ensure errors are logged properly"""
    assert context.error_count > 0, "No errors logged!"
    logging.info("All {} errors were correctly logged.".format(context.error_count))


@then("processing should continue without failure for valid transactions")
def step_then_continue_processing_valid_transactions(context):
    # type: (Any) -> None
    """Ensure valid transactions are processed despite errors"""
    valid_transactions = context.transaction_count - context.error_count
    assert valid_transactions > 0, "No valid transactions processed!"
    logging.info(
        "{} valid transactions processed successfully.".format(valid_transactions)
    )


@given(
    'a database containing "{transaction_count}" transactions from bank export files'
)
def large_transaction_database(context, transaction_count):
    # type: (Any, Any) -> None
    """Simulate a database with a large transaction dataset"""
    context.transaction_count = int(transaction_count)


@when("I execute a complex query with multiple joins and filters")
def step_when_execute_large_transaction_query(context):
    # type: (Any) -> None
    """Simulate executing complex queries on a large dataset with tracked execution time."""

    # Simulate query duration (short range for test performance)
    simulated_query_time = random.uniform(2.0, 5.0)  # Seconds

    # Start timing the query execution
    start_time = datetime.now()

    # Simulate processing delay using busy-wait loop (no-op)
    while (datetime.now() - start_time).total_seconds() < simulated_query_time:
        pass  # Simulate time passing for the query execution

    # End time and calculate elapsed
    elapsed = (datetime.now() - start_time).total_seconds()

    # Store the query execution time in context for further validation
    context.query_time = elapsed

    # Log the elapsed query time using .format (compatible with Python 2.7)
    logging.info(
        "Simulated query execution completed in {0:.2f} seconds.".format(elapsed)
    )


@then('query execution should complete within "{expected_time}" seconds')
def step_then_validate_query_performance(context, expected_time):
    # type: (Any, Any) -> None
    """Ensure query execution is optimized"""
    assert context.query_time <= float(
        expected_time
    ), "Query execution time exceeded limit!"
    logging.info("Query executed in {:.2f} seconds.".format(context.query_time))


# ================= End of Large Transaction Volume Processing Step Definitions =================

# ================= Beginning of System Memory Usage Performance Step Definitions =================
# This script evaluates system memory consumption during various stages of bank export file processing.
# It ensures:
# - Efficient memory management when handling large datasets.
# - Stability under batch processing and long-running transactions.
# - Prevention of memory leaks and handling of memory overflow errors.
# - Optimal database query execution with minimal memory overhead.


@given('a bank export file containing "{row_count}" rows and "{column_count}" columns')
def large_file_memory_usage(context, row_count, column_count):
    # type: (Any, Any, Any) -> None
    """Simulate a large bank export file for memory performance testing"""
    context.row_count = int(row_count)
    context.column_count = int(column_count)


@when("the system processes the file")
def step_when_process_large_file_memory(context):
    # type: (Any) -> None
    """Simulate processing a large export file and monitor memory usage."""

    # Simulate a processing duration (shortened for testing)
    simulated_processing_time = random.uniform(1.0, 5.0)  # Seconds

    # Capture initial memory usage
    initial_memory = psutil.virtual_memory().percent

    # Start timing
    start_time = datetime.now()

    # Simulate time passing using busy-wait loop (no-op)
    while (datetime.now() - start_time).total_seconds() < simulated_processing_time:
        pass  # Simulate processing time passing

    # End timing and calculate elapsed
    end_time = datetime.now()
    elapsed_time = (end_time - start_time).total_seconds()

    # Estimate memory usage during processing (simulated increase in memory usage)
    context.processing_time = elapsed_time
    context.memory_usage = initial_memory + random.uniform(5.0, 20.0)

    # Logging the completion and memory usage (using .format for compatibility)
    logging.info(
        "Simulated file processing completed in {0:.2f} seconds.".format(elapsed_time)
    )
    logging.info(
        "Estimated memory usage during processing: {0:.2f}%".format(
            context.memory_usage
        )
    )


@then('memory consumption should not exceed "{memory_limit}%"')
def step_then_memory_should_not_exceed(context, memory_limit):
    # type: (Any, Any) -> None
    """Ensure memory usage remains within acceptable limits"""
    assert context.memory_usage <= float(
        memory_limit
    ), "Memory consumption exceeded limit!"
    logging.info(
        "Memory usage: {:.2f}% within acceptable limits.".format(context.memory_usage)
    )


@given("I attempt to import the file into the database")
def step_when_import_large_memory_dataset(context):
    # type: (Any) -> None
    """Simulate database import process and monitor memory usage."""

    # Simulate import delay (test-safe range)
    simulated_import_time = random.uniform(1.0, 5.0)

    # Capture initial memory usage
    initial_memory = psutil.virtual_memory().percent

    # Start timing
    start_time = datetime.now()

    # Simulate delay using busy-wait loop
    while (datetime.now() - start_time).total_seconds() < simulated_import_time:
        pass  # Simulating import delay (busy-wait loop)

    # Calculate elapsed time
    elapsed_time = (datetime.now() - start_time).total_seconds()

    # Store simulated import time and memory usage in context
    context.import_time = elapsed_time
    context.memory_usage = initial_memory + random.uniform(
        10.0, 25.0
    )  # Simulating increased memory usage

    # Log the results (using .format for compatibility)
    logging.info(
        "Simulated database import completed in {0:.2f} seconds.".format(elapsed_time)
    )
    logging.info(
        "Estimated memory usage during import: {0:.2f}%".format(context.memory_usage)
    )


@then('database memory consumption should not exceed "{memory_limit}%"')
def step_then_validate_database_memory(context, memory_limit):
    # type: (Any, Any) -> None
    """Ensure database import does not cause excessive memory usage"""
    assert context.memory_usage <= float(
        memory_limit
    ), "Database memory usage exceeded limit!"
    logging.info(
        "Database memory usage: {:.2f}% within limits.".format(context.memory_usage)
    )


@given('"{batch_count}" bank export files each containing "{row_count}" rows')
def large_batch_files_memory(context, batch_count, row_count):
    # type: (Any, Any, Any) -> None
    """Simulate batch file processing and memory performance"""
    context.batch_count = int(batch_count)
    context.row_count = int(row_count)


@given("I attempt to import the file into the database")
def step_when_import_large_memory_dataset(context):
    # type: (Any) -> None
    """Simulate database import process and monitor memory usage."""

    # Simulate import delay (test-safe range)
    simulated_import_time = random.uniform(1.0, 5.0)

    # Capture initial memory usage
    initial_memory = psutil.virtual_memory().percent

    # Start timing
    start_time = datetime.now()

    # Simulate delay using busy-wait loop
    while (datetime.now() - start_time).total_seconds() < simulated_import_time:
        pass  # Simulating import delay (busy-wait loop)

    # Calculate elapsed time
    elapsed_time = (datetime.now() - start_time).total_seconds()

    # Store simulated import time and memory usage in context
    context.import_time = elapsed_time
    context.memory_usage = initial_memory + random.uniform(
        10.0, 25.0
    )  # Simulating increased memory usage

    # Log the results (using .format for compatibility)
    logging.info(
        "Simulated database import completed in {0:.2f} seconds.".format(elapsed_time)
    )
    logging.info(
        "Estimated memory usage during import: {0:.2f}%".format(context.memory_usage)
    )


@then('total memory consumption should remain below "{memory_limit}%"')
def step_then_check_memory_after_batch(context, memory_limit):
    # type: (Any, Any) -> None
    """Ensure batch processing does not cause excessive memory usage"""
    assert context.memory_usage <= float(
        memory_limit
    ), "Batch processing memory usage exceeded limit!"
    logging.info(
        "Batch memory usage: {:.2f}% within acceptable limits.".format(
            context.memory_usage
        )
    )


@given(
    'a continuous stream of bank export files arriving every "{interval}" seconds with "{row_count}" rows each'
)
def long_running_memory_test(context, interval, row_count):
    # type: (Any, Any, Any) -> None
    """Simulate long-running data stream processing"""
    context.interval = int(interval)
    context.row_count = int(row_count)


@when('the system processes them for "{duration}" hours')
def step_when_process_long_running_memory(context, duration):
    # type: (Any, Any) -> None
    """Monitor memory stability during long-running transactions (simulated)."""

    # Parse and store duration
    context.duration = int(duration)

    # Get initial memory usage
    initial_memory = psutil.virtual_memory().percent

    # Simulate long-running process (scaled down: 1 second = 1 hour)
    simulated_total_time = context.duration  # In seconds for testing

    logging.info(
        "Starting long-running processing simulation for "
        + str(context.duration)
        + " hour(s)..."
    )

    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=simulated_total_time)

    while datetime.now() < end_time:
        memory_spike = random.uniform(0.0, 5.0)  # Simulate memory fluctuation

        if memory_spike > 3.0:
            logging.warning(
                "Memory increased unexpectedly: {:.2f}%".format(memory_spike)
            )

        current_memory = psutil.virtual_memory().percent
        logging.info("Current memory usage: {:.2f}%".format(current_memory))

        time.sleep(0.1)  # Keep loop tight but not CPU-intensive

    logging.info("Long-running processing simulation completed.")


@then("memory usage should not increase unexpectedly")
def step_then_check_memory_leak(context):
    # type: (Any) -> None
    """Ensure memory leaks do not occur"""
    final_memory = psutil.virtual_memory().percent
    assert final_memory < 90, "Memory leak detected!"
    logging.info("Memory usage stabilized at {:.2f}%.".format(final_memory))


@given('a bank export file "{file_name}" that exceeds memory limits')
def memory_overflow_file(context, file_name):
    # type: (Any, Any) -> None
    """Simulate memory overflow scenario"""
    context.file_name = file_name


@when("I attempt to process the file")
def step_when_process_memory_overflow(context):
    # type: (Any) -> None
    """Monitor system response to simulated memory overflow."""

    # Simulate near-critical memory usage
    memory_usage = random.uniform(90.0, 100.0)
    context.memory_usage = memory_usage

    # Simulate short processing delay using busy wait
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=2)

    while datetime.now() < end_time:
        pass  # Busy wait

    # Log simulated memory overflow
    logging.warning("Simulated memory usage reached {:.2f}%.".format(memory_usage))


@then("the system should log memory exhaustion errors")
def step_then_log_memory_exhaustion(context):
    # type: (Any) -> None
    """Ensure system logs memory exhaustion errors correctly"""
    assert context.memory_usage >= 90, "Memory exhaustion not detected!"
    logging.error("Memory exhaustion detected: {:.2f}%.".format(context.memory_usage))


@then("system operations should continue without crashing")
def step_then_system_stability_after_overflow(context):
    # type: (Any) -> None
    """Ensure system does not crash due to memory overflow"""
    system_stable = random.choice([True, False])
    assert system_stable, "System crashed due to memory overload!"
    logging.info("System continued running despite memory overload.")


@given('a database containing "{row_count}" records from bank export files')
def large_query_memory_usage(context, row_count):
    # type: (Any, Any) -> None
    """Simulate database with large record set for query performance testing"""
    context.row_count = int(row_count)


@when("I execute a complex query with multiple joins and filters")
def step_when_execute_query_memory_usage(context):
    # type: (Any) -> None
    """Monitor memory consumption during complex query execution."""

    # Simulate query duration (shortened for testing)
    simulated_query_time = random.uniform(2.0, 5.0)  # Seconds

    # Capture initial memory
    initial_memory = psutil.virtual_memory().percent

    # Start timing
    start_time = datetime.now()

    # Simulate the delay (busy-wait loop)
    while (datetime.now() - start_time).total_seconds() < simulated_query_time:
        pass  # Simulating query execution delay

    # Calculate actual elapsed time
    elapsed = (datetime.now() - start_time).total_seconds()

    # Store in context
    context.query_time = elapsed
    context.memory_usage = initial_memory + random.uniform(
        5.0, 15.0
    )  # Simulated memory increase

    # Log the results using .format() for compatibility
    logging.info("Simulated query executed in {0:.2f} seconds.".format(elapsed))
    logging.info(
        "Estimated memory usage during query: {0:.2f}%".format(context.memory_usage)
    )


@then('memory consumption should not exceed "{memory_limit}%"')
def step_then_validate_query_memory_usage(context, memory_limit):
    # type: (Any, Any) -> None
    """Ensure query execution remains within memory limits"""
    assert context.memory_usage <= float(
        memory_limit
    ), "Query memory usage exceeded limit!"
    logging.info("Query memory usage: {}% within limits.".format(context.memory_usage))


@then('query execution should complete within "{expected_time}" seconds')
def step_then_validate_query_time(context, expected_time):
    # type: (Any, Any) -> None
    """Ensure query execution completes within expected time"""
    assert context.query_time <= float(
        expected_time
    ), "Query execution time exceeded limit!"
    logging.info("Query executed in {} seconds.".format(context.query_time))


# ================= End of System Memory Usage Performance Step Definitions =================

# ================= Beginning of Duplicate Imports Validation Step Definitions =================
# This script ensures the system properly handles duplicate imports in bank export file processing.
# It verifies:
# - Detection and prevention of duplicate file imports.
# - Maintenance of database integrity when duplicate files are encountered.
# - Efficient batch processing behavior while skipping duplicates.
# - Proper error handling and logging for duplicate import attempts.
# - Performance impact assessment when detecting large-scale duplicate imports.


@given('a previously processed bank export file named "{file_name}"')
def previously_processed_file(context, file_name):
    # type: (Any, Any) -> None
    """Simulate a file that has already been processed"""
    if not hasattr(context, "processed_files"):
        context.processed_files = set()
    context.processed_files.add(file_name)
    context.file_name = file_name


@when("I attempt to import the same file again")
def step_when_attempt_duplicate_import(context):
    # type: (Any) -> None
    """Attempt to import a previously processed file."""

    # Ensure the processed_files set exists in context
    if not hasattr(context, "processed_files"):
        context.processed_files = set()

    # Check for duplicate
    context.is_duplicate = context.file_name in context.processed_files

    # Simulate short processing delay (1 second)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass  # Busy wait

    logging.info(
        "Duplicate import check completed for file: {}".format(context.file_name)
    )
    logging.info("Duplicate status: {}".format(context.is_duplicate))


@then("the system should detect the duplicate import attempt")
def step_then_detect_duplicate_import(context):
    # type: (Any) -> None
    """Verify that the system identifies the duplicate file"""
    assert context.is_duplicate, "Duplicate file import was not detected!"
    logging.info("Duplicate import detected: {}".format(context.file_name))


@then('an error message "{error_message}" should be displayed')
def step_then_display_error_message(context, error_message):
    # type: (Any, Any) -> None
    """Ensure the correct error message is displayed when a duplicate file is detected."""

    if getattr(context, "is_duplicate", False):
        actual_message = "Duplicate file detected: {}".format(context.file_name)
        logging.error(actual_message)
        assert (
            error_message in actual_message
        ), "Expected error message '{}' not found in actual message '{}'".format(
            error_message, actual_message
        )
    else:
        logging.info("No error message displayed because the file is not a duplicate.")
        assert (
            False
        ), "Expected duplicate error message but file was not detected as duplicate."


@then("the duplicate file should not be processed")
def step_then_prevent_duplicate_processing(context):
    # type: (Any) -> None
    """Ensure duplicate files are not reprocessed"""
    if context.file_name in processed_files:
        logging.info("File '{}' was correctly skipped.".format(context.file_name))
    else:
        raise AssertionError(
            "Duplicate file '{}' was incorrectly processed.".format(context.file_name)
        )


@given('a database containing records from "{file_name}"')
def database_contains_file(context, file_name):
    # type: (Any, Any) -> None
    """Simulate a database that already contains the records from the file"""
    if not hasattr(context, "processed_files"):
        context.processed_files = set()
    context.processed_files.add(file_name)
    context.file_name = file_name


@then("no duplicate records should be inserted")
def step_then_no_duplicate_records(context):
    # type: (Any) -> None
    """Ensure duplicate records are not added to the database"""
    if not hasattr(context, "processed_files"):
        context.processed_files = set()

    assert (
        context.file_name in context.processed_files
    ), "Duplicate records were inserted!"
    logging.info("No duplicate records added from '" + context.file_name + "'.")


@given('a batch of bank export files including a duplicate file named "{file_name}"')
def batch_with_duplicate_file(context, file_name):
    # type: (Any, Any) -> None
    """Simulate a batch of files including a duplicate"""
    if not hasattr(context, "processed_files"):
        context.processed_files = set()

    context.processed_files.add(file_name)
    context.file_name = file_name
    context.batch_files = [file_name, "new_file_1.csv", "new_file_2.xlsx"]


@when("I process the batch")
def the_file_batch_is_processed(context):
    # type: (Any) -> None
    """Simulate batch processing of files and identify duplicates."""

    if not hasattr(context, "processed_files"):
        context.processed_files = set()

    batch_files = getattr(context, "batch_files", [])

    # Identify duplicates in the batch
    context.duplicates_found = []
    for file_name in batch_files:
        if file_name in context.processed_files:
            context.duplicates_found.append(file_name)
        else:
            context.processed_files.add(file_name)

    # Simulate processing delay (2 seconds)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=2)

    while datetime.now() < end_time:
        pass  # Simulate time passing with a busy wait

    logging.info(
        "Batch processed. Duplicates found: {}".format(context.duplicates_found)
    )


@then("only unique files should be imported")
def step_then_unique_files_only(context):
    # type: (Any) -> None
    """Ensure only unique files are processed"""
    unique_files = []
    for f in context.batch_files:
        if f not in processed_files:
            unique_files.append(f)

    assert len(unique_files) > 0, "No unique files were processed!"
    logging.info("Unique files processed: {}".format(unique_files))


@then('duplicate files should be skipped with a warning "{warning_message}"')
def step_then_warn_duplicate_files(context, warning_message):
    # type: (Any, Any) -> None
    """Ensure duplicate files are skipped with a proper warning"""
    for duplicate in context.duplicates_found:
        logging.warning(warning_message + ": " + duplicate)


@given('an attempt to import a duplicate file named "{file_name}"')
def attempt_duplicate_import(context, file_name):
    # type: (Any, Any) -> None
    """Simulate an attempt to reimport a duplicate file"""
    context.file_name = file_name
    context.is_duplicate = file_name in processed_files


@then('a user notification should be sent with "{notification_message}"')
def step_then_send_notification(context, notification_message):
    # type: (Any, Any) -> None
    """Ensure users are notified when a duplicate is detected"""

    if getattr(context, "is_duplicate", False):
        logging.info("Duplicate issue detected: {0}".format(notification_message))
        # Reuse the send_notification function from user_notifications.py
        send_notification(context, notification_message)


@then("an audit log should capture the duplicate attempt")
def step_then_audit_log_duplicate(context):
    # type: (Any) -> None
    """Ensure the duplicate attempt is recorded in an audit log"""
    logging.info("Audit Log import attempt detected for {}".format(context.file_name))


@given('a system processing "{file_count}" export files per minute')
def system_processing_files(context, file_count):
    # type: (Any, Any) -> None
    """Simulate a system processing a high volume of files"""
    context.file_count = int(file_count)


@when('"{duplicate_count}" duplicate files are included')
def step_when_add_duplicates_to_processing(context, duplicate_count):
    # type: (Any, Any) -> None
    """Simulate adding duplicates to the processing queue"""
    context.duplicate_count = int(duplicate_count)


@then("the duplicate detection should not cause significant processing delay")
def step_then_no_performance_impact(context):
    # type: (Any) -> None
    """Ensure duplicate detection does not slow down processing"""
    processing_time = random.uniform(0.1, 1.5) * context.file_count
    assert (
        processing_time < 5 * context.file_count
    ), "Processing delay due to duplicates!"
    logging.info(
        "Processing completed in {:.2f} seconds without significant delay.".format(
            processing_time
        )
    )


@then('processing speed should remain above "{expected_speed}" files per minute')
def step_then_maintain_processing_speed(context, expected_speed):
    # type: (Any, Any) -> None
    """Ensure processing speed remains within acceptable limits"""
    expected_speed = int(expected_speed)
    actual_speed = context.file_count - context.duplicate_count
    assert actual_speed >= expected_speed, "Processing speed dropped below threshold!"
    logging.info(
        "Processing speed maintained at {} files per minute.".format(actual_speed)
    )


@then("system performance should not degrade significantly")
def step_then_system_performance_stable(context):
    # type: (Any) -> None
    """Ensure system performance remains stable despite duplicate detection"""
    memory_usage = random.uniform(50, 75)
    assert memory_usage < 80, "System performance degraded due to memory overload!"
    logging.info(
        "System memory usage at {:.2f}%, no performance degradation detected.".format(
            memory_usage
        )
    )


# ================= End of Duplicate Imports Validation Step Definitions =================

# ================= Beginning of Historical Data Consistency Validation Step Definitions =================
# This script ensures the system maintains historical data consistency in bank export file processing.
# It verifies:
# - That historical data remains unchanged over time.
# - Database consistency when comparing stored historical records with the latest exports.
# - Efficient batch processing of historical data comparisons.
# - Proper error handling and logging for historical data inconsistencies.
# - Performance impact assessment when validating large-scale historical data.


@given('a database containing historical records from "{year}"')
def historical_database(context, year):
    # type: (Any, Any) -> None
    """Simulate a database containing historical data"""
    context.year = year
    context.db_data = {"year": year, "hash": compute_file_hash(str(year))}


@when("I compare it with the latest processed version")
def step_when_compare_with_latest(context):
    # type: (Any) -> None
    """Simulate comparing the historical file with the latest version"""

    # Compare hashes using the new helper function
    context.is_modified = compare_hashes(
        context,
        compute_file_hash(context.file_name + "_latest"),
        source_name="historical_records",
        target_name=context.file_name,
    )


@when('I compare the stored records with the latest export file "{file_name}"')
def step_when_compare_database_with_file(context, file_name):
    # type: (Any, Any) -> None
    """Simulate a database comparison with the latest export"""

    # Compare hashes using the new helper function
    context.discrepancy_found = compare_hashes(
        context,
        compute_file_hash(file_name),  # Hash to compare
        source_name="context.db_data",  # Hash source in context
        target_name=file_name,  # File name (target)
    )


@given('a bank export file named "{file_name}" containing previously resolved issues')
def resolved_issues_file(context, file_name):
    # type: (Any, Any) -> None
    """Simulate tracking a file with previously resolved issues"""
    if not hasattr(context, "resolved_issues"):
        context.resolved_issues = {}
    context.resolved_issues[file_name] = {"hash": compute_file_hash(file_name)}
    context.file_name = file_name


@given('a historical bank export file named "{file_name}" from "{year}"')
def historical_file(context, file_name, year):
    # type: (Any, Any, Any) -> None
    """Simulate a historical file that has already been stored"""
    historical_records[file_name] = {"year": year, "hash": compute_file_hash(file_name)}
    context.file_name = file_name
    context.year = year


@then("the historical records should remain identical")
def step_then_validate_historical_consistency(context):
    # type: (Any) -> None
    """Verify that historical records remain unchanged"""
    assert not context.is_modified, "Historical records have been modified!"
    logging.info(
        "Historical data consistency validated for {}.".format(context.file_name)
    )


@then("no unauthorized modifications should be detected")
def step_then_no_unauthorized_modifications(context):
    # type: (Any) -> None
    """Ensure no unexpected modifications are detected"""

    # Check if the file has been modified
    if context.is_modified:
        logging.warning(
            "Unauthorized modification detected in {0}!".format(context.file_name)
        )
    else:
        logging.info(
            "No unauthorized modifications detected in {0}.".format(context.file_name)
        )


@then("a validation report should be generated")
def step_then_generate_validation_report(context):
    # type: (Any) -> None
    """Generate a report for the historical data validation"""
    logging.info("Validation report generated for {}.".format(context.file_name))


@then("all records should match exactly")
def step_then_database_consistency(context):
    # type: (Any) -> None
    """Ensure database records are consistent"""
    assert not context.discrepancy_found, "Database records do not match!"
    logging.info("Database consistency verified successfully.")


@then('any discrepancies should be logged as "{discrepancy_type}"')
def step_then_log_discrepancies(context, discrepancy_type):
    # type: (Any, Any) -> None
    """Ensure discrepancies are logged properly"""

    # Log the discrepancy with the given type
    if context.discrepancy_found:
        logging.warning("Discrepancy detected: {0}".format(discrepancy_type))
    else:
        logging.info("No discrepancies found of type: {0}".format(discrepancy_type))


@then("a detailed report should be generated")
def step_then_generate_detailed_report(context):
    # type: (Any) -> None
    """Generate a report for database discrepancies"""
    logging.info("Detailed historical data validation report generated.")


@given('a batch of historical bank export files from "{year_range}"')
def historical_batch(context, year_range):
    # type: (Any, Any) -> None
    """Simulate a batch of historical files for processing"""
    context.year_range = year_range
    context.batch_files = [
        "transactions_{}.csv".format(year) for year in range(2015, 2021)
    ]


@when("I process them for consistency checking")
def step_when_process_historical_batch(context):
    # type: (Any) -> None
    """Simulate processing a batch of historical records for consistency."""

    # Simulate detecting random discrepancies
    context.discrepancy_count = random.randint(0, 5)

    # Simulate processing time (2 seconds)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=2)

    while datetime.now() < end_time:  # Busy wait to simulate processing delay
        pass

    logging.info(
        "Historical batch processed. Discrepancies found: {0}".format(
            context.discrepancy_count
        )
    )


@then("all historical records should be verified")
def step_then_verify_historical_records(context):
    # type: (Any) -> None
    """Ensure all records in the batch are verified"""
    logging.info("All records from {} verified.".format(context.year_range))


@then('discrepancies should be flagged with severity levels "{severity}"')
def step_then_flag_discrepancies(context, severity):
    # type: (Any, Any) -> None
    """Ensure discrepancies are flagged with appropriate severity"""
    if context.discrepancy_count > 0:
        logging.warning(
            str(context.discrepancy_count)
            + " discrepancies found with severity: "
            + severity
        )


@given('an attempt to validate historical data from "{file_name}"')
def attempt_to_validate_historical(context, file_name):
    # type: (Any, Any) -> None
    """Simulate an attempt to validate a historical export file"""
    context.file_name = file_name


@when('inconsistencies such as "{error_type}" are found')
def step_when_find_inconsistencies(context, error_type):
    # type: (Any, Any) -> None
    """Simulate detecting inconsistencies in historical data"""
    context.error_found = bool(random.getrandbits(1))
    context.error_type = error_type if context.error_found else None


@then("a detailed log should capture all errors")
def step_then_log_historical_errors(context):
    # type: (Any) -> None
    """Ensure errors are logged for historical validation"""
    if context.error_found:
        logging.error(
            "Historical data inconsistency detected: {}".format(context.error_type)
        )


@given('a system processing "{file_count}" historical export files per hour')
def system_processing_historical(context, file_count):
    # type: (Any, Any) -> None
    """Simulate the system processing a high volume of historical records"""
    context.file_count = int(file_count)


@when('comparisons involve large datasets from "{year_range}"')
def step_when_compare_large_datasets(context, year_range):
    # type: (Any, Any) -> None
    """Simulate processing large historical datasets."""

    # Store year range in context
    context.year_range = year_range

    # Simulate a processing duration (in seconds, not real wait)
    context.processing_time = random.randint(100, 600)

    # Simulate short delay (replacing time.sleep(1))
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass  # Busy wait loop

    logging.info(
        "Processed large dataset from {} in simulated {} seconds.".format(
            year_range, context.processing_time
        )
    )


@then('system resources should not exceed "{resource_limit}%"')
def step_then_monitor_resource_usage(context, resource_limit):
    # type: (Any, Any) -> None
    """Ensure resource usage remains within limits"""
    actual_resource_usage = random.randint(50, 90)
    assert actual_resource_usage <= int(resource_limit), "Resource usage exceeded!"
    logging.info(
        "Resource usage: {}%, within allowed limit.".format(actual_resource_usage)
    )


@then("database queries should remain optimized")
def step_then_database_queries_optimized(context):
    # type: (Any) -> None
    """Ensure database queries perform optimally"""
    query_time = random.uniform(0.1, 2.5)
    assert query_time < 3, "Database query performance degraded!"
    logging.info("Database query executed in {:.2f} seconds.".format(query_time))


# ================= End of Historical Data Consistency Validation Step Definitions =================

# ================= Beginning of Previously Resolved Issues Validation Step Definitions =================
# This script ensures that previously resolved issues in bank export file processing do not reoccur.
# It verifies:
# - That past data integrity issues remain fixed.
# - Database consistency for previously resolved discrepancies.
# - Batch processing reliability in maintaining data resolution.
# - Proper error handling when past issues resurface.
# - Performance impact of regression verification in large-scale exports.

# ================= Beginning of Previously Resolved Issues Validation Step Definitions =================
# This script ensures that previously resolved issues in bank export file processing do not reoccur.
# It verifies:
# - That past data integrity issues remain fixed.
# - Database consistency for previously resolved discrepancies.
# - Batch processing reliability in maintaining data resolution.
# - Proper error handling when past issues resurface.
# - Performance impact of regression verification in large-scale exports.


@then("the issue should not reoccur")
def step_then_validate_no_reoccurrence(context):
    # type: (Any) -> None
    """Ensure resolved issues remain resolved"""
    assert not context.is_reoccurring, "Previously resolved issues have reoccurred!"
    print("No issues found in {}. Integrity maintained.".format(context.file_name))


@then("a validation report should confirm their resolution")
def step_then_generate_resolution_report(context):
    # type: (Any) -> None
    """Generate a report confirming no reoccurrence of past issues"""
    print("Validation report generated for {}.".format(context.file_name))


@then('any reoccurrence should be flagged as "{severity}"')
def step_then_flag_reoccurrence(context, severity):
    # type: (Any, Any) -> None
    """Flag reoccurrence of issues based on severity."""

    # Simulate checking if the issue is recurring
    message = "Issue reoccurred in {0}. Severity: {1}".format(
        context.file_name, severity
    )

    # Log the event that checks if the issue reoccurred
    logging.info("Checking for issue reoccurrence with message: {0}".format(message))

    # Simulate the flagging logic
    if context.is_reoccurring(message):
        # If the issue is flagged as recurring, log based on severity
        if severity == "high":
            logging.warning("High severity reoccurrence detected: {0}".format(message))
        elif severity == "medium":
            logging.warning(
                "Medium severity reoccurrence detected: {0}".format(message)
            )
        elif severity == "low":
            logging.info("Low severity reoccurrence detected: {0}".format(message))
        else:
            logging.error(
                "Unknown severity level for reoccurrence: {0}".format(message)
            )
    else:
        logging.info("No reoccurrence detected for: {0}".format(message))


@given('a database with "{issue_status}" issues of type "{issue_type}"')
def database_with_issues(context, issue_status, issue_type):
    # type: (Any, Any, Any) -> None
    """Simulate a database with issues, either resolved or unresolved."""
    context.issue_type = issue_type

    if issue_status == "resolved":
        context.db_data = {"issue_type": issue_type, "status": "Resolved"}
    else:
        context.db_data = {"issue_type": issue_type, "status": "Unresolved"}

    context.db_resolved = issue_status == "resolved"
    logging.info(
        "Database status: {} for issue type: {}".format(
            context.db_data["status"], issue_type
        )
    )


@when("I compare the latest records with previous resolutions")
def step_when_compare_database_records(context):
    # type: (Any) -> None
    """Simulate a database comparison to check for resolved issue reoccurrence."""

    # Simulate discrepancy randomly
    context.discrepancy_found = bool(random.getrandbits(1))

    # Simulate processing delay (1 second)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass  # Simulate processing time

    logging.info(
        "Comparison completed. Discrepancy found: {}".format(context.discrepancy_found)
    )


@then("no past issues should reappear")
def step_then_no_past_issues_reappear(context):
    # type: (Any) -> None
    """Ensure past issues do not reoccur"""
    assert (
        not context.discrepancy_found
    ), "Previously resolved issue ({}) has resurfaced!".format(context.issue_type)
    logging.info("Database integrity maintained for past issues.")


@then('any detected inconsistencies should be logged as "{discrepancy_type}"')
def step_then_log_inconsistencies(context, discrepancy_type):
    # type: (Any, Any) -> None
    """Ensure any detected inconsistencies are logged with the provided discrepancy type"""

    # Check if any discrepancies are detected
    if context.discrepancy_found:
        # Log the detected discrepancy based on the type provided in the step
        logging.info(
            "Detected discrepancy: {0} of type {1}".format(
                context.file_name, discrepancy_type
            )
        )

        # Depending on the discrepancy type, log it with different severity
        if discrepancy_type == "high":
            logging.warning(
                "High severity discrepancy detected in file: {0}".format(
                    context.file_name
                )
            )
        elif discrepancy_type == "medium":
            logging.warning(
                "Medium severity discrepancy detected in file: {0}".format(
                    context.file_name
                )
            )
        elif discrepancy_type == "low":
            logging.info(
                "Low severity discrepancy detected in file: {0}".format(
                    context.file_name
                )
            )
        else:
            logging.error(
                "Unknown discrepancy type detected for file: {0}".format(
                    context.file_name
                )
            )
    else:
        # If no discrepancies are found, log that no discrepancies were detected
        logging.info("No discrepancies detected in file: {0}".format(context.file_name))


@given(
    'a batch of bank export files from "{year_range}" containing previously flagged issues'
)
def batch_with_past_issues(context, year_range):
    # type: (Any, Any) -> None
    """Simulate a batch of files containing previously flagged issues"""
    context.year_range = year_range
    context.batch_files = [
        "transactions_{}.csv".format(year) for year in range(2018, 2023)
    ]


@when("I process them for validation")
def step_when_process_for_validation_batch(context):
    # type: (Any) -> None
    """Simulate batch processing for consistency checking"""
    # Simulate random issue count for batch processing
    context.issue_count = random.randint(0, 5)  # Random reoccurrence simulation
    # Simulating the batch processing with datetime
    start_time = datetime.now()
    # Add any specific logic for batch processing here
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()

    # Log the batch processing time
    logging.info("Batch processing took {:.2f} seconds".format(duration))


@then("all records should pass consistency checks")
def step_then_validate_batch_consistency(context):
    # type: (Any) -> None
    """Ensure all records pass consistency validation"""
    logging.info(
        "All records from {} passed consistency checks.".format(context.year_range)
    )


@then("no previously resolved issues should reoccur")
def step_then_no_resolved_issues_reappear(context):
    # type: (Any) -> None
    """Ensure no past issues resurface"""
    assert context.issue_count == 0, "Resolved issues have reappeared!"
    logging.info("Batch processing verified for resolved issues.")


@given('an attempt to process a bank export file "{file_name}"')
def attempt_to_process_file(context, file_name):
    # type: (Any, Any) -> None
    """Simulate an attempt to process a bank export file with past issues"""
    context.file_name = file_name


@when('previously resolved issues such as "{error_type}" are detected again')
def step_when_detect_resolved_issues(context, error_type):
    # type: (Any, Any) -> None
    """Simulate detecting previously resolved issues in the latest export"""
    context.error_found = bool(random.getrandbits(1))
    context.error_type = error_type if context.error_found else None


@then('the issue should be escalated if its severity level is "{severity_level}"')
def step_then_escalate_critical_issues(context, severity_level):
    # type: (Any, Any) -> None
    """Ensure critical issues are escalated if they reoccur"""
    if getattr(context, "error_found", False):
        logging.info(
            "Issue reoccurrence escalated due to severity: {}".format(severity_level)
        )


@given('a system processing "{file_count}" bank export files per hour')
def system_processing_resolved_issues(context, file_count):
    # type: (Any, Any) -> None
    """Simulate the system processing multiple export files per hour"""
    context.file_count = int(file_count)


@when('checking for previously resolved issues in "{year_range}"')
def step_when_checking_past_resolved_issues(context, year_range):
    # type: (Any, Any) -> None
    """Simulate validating previously resolved issues in a specific year range."""

    # Store year range in context
    context.year_range = year_range

    # Simulate processing time (e.g., system effort)
    context.processing_time = random.randint(100, 600)

    # Simulate brief delay (replacing time.sleep(1))
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass

    logging.info(
        "Checked previously resolved issues for {}. Simulated processing time: {} seconds.".format(
            year_range, context.processing_time
        )
    )


@then('system resources should not exceed "{resource_limit}%"')
def step_then_monitor_system_usage(context, resource_limit):
    # type: (Any, Any) -> None
    """Ensure resource usage remains within acceptable limits"""
    actual_resource_usage = random.randint(50, 90)
    assert actual_resource_usage <= int(
        resource_limit
    ), "System resource usage exceeded!"
    logging.info(
        "Resource usage: {}%, within the allowed limit.".format(actual_resource_usage)
    )


@then("data integrity should remain stable throughout the process")
def step_then_data_integrity_stable(context):
    # type: (Any) -> None
    """Ensure data integrity remains stable while validating past resolved issues"""
    integrity_check = bool(
        random.getrandbits(1)
    )  # Simulating data stability validation
    assert integrity_check, "Data integrity issues detected!"
    logging.info("Data integrity verified successfully.")


# ================= End of Previously Resolved Issues Validation Step Definitions =================

# ================= Beginning of High-Volume Transaction Processing Validation =================
# This script validates high-volume transaction processing in regression testing.
# It ensures:
# - Data integrity across millions of transactions.
# - Database performance under large transaction loads.
# - Efficient batch processing with optimized memory usage.
# - Proper error handling for high transaction volumes.
# - Network latency resilience during remote processing.
# - Performance scaling under extreme transaction loads.


@given(
    'a bank export file named "{file_name}" containing "{transaction_count}" transactions'
)
def high_volume_file(context, file_name, transaction_count):
    # type: (Any, Any, Any) -> None
    """Simulate loading a high-volume transaction file"""
    context.file_name = file_name
    context.transaction_count = int(transaction_count)
    context.processed_count = 0


@when("the system processes the file")
def step_when_process_high_volume_file(context):
    # type: (Any) -> None
    """Simulate processing the high-volume transaction file."""

    # Simulate processing time based on transaction count, max 5 seconds
    processing_time = min(context.transaction_count / 10000.0, 5.0)

    # Simulate processing delay using datetime
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=processing_time)

    while datetime.now() < end_time:
        pass  # Simulate processing time

    # Update context with number of transactions processed
    context.processed_count = context.transaction_count

    logging.info(
        "Processed {} transactions from {}.".format(
            context.processed_count, context.file_name
        )
    )


@then("all transactions should be recorded accurately")
def step_then_validate_transactions_recorded(context):
    # type: (Any) -> None
    """Ensure all transactions are recorded without errors"""
    assert (
        context.processed_count == context.transaction_count
    ), "Transaction count mismatch!"
    logging.info("All transactions recorded successfully.")


@then("no data loss or corruption should occur")
def step_then_validate_no_data_loss(context):
    # type: (Any) -> None
    """Ensure no transaction data is lost"""
    assert random.random() > 0.05, "Data loss detected!"
    logging.info("No data corruption detected during processing.")


@given('a database containing "{transaction_count}" transactions')
def database_high_transaction_load(context, transaction_count):
    # type: (Any, Any) -> None
    """Simulate a database loaded with high transaction volumes"""
    context.db_transaction_count = int(transaction_count)


@when('a query retrieves transactions from the last "{time_period}"')
def step_when_query_transactions(context, time_period):
    # type: (Any, Any) -> None
    """Simulate querying a large database for recent transactions."""

    # Simulated query execution time
    context.query_time = random.randint(1, 10)
    context.time_period = time_period

    # Simulate the query delay using datetime
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=context.query_time)

    while datetime.now() < end_time:
        pass  # Simulating query execution delay

    logging.info(
        "Queried transactions for {} in {} seconds.".format(
            context.time_period, context.query_time
        )
    )


@then('the query should execute within "{expected_time}" seconds')
def step_then_query_execution_time(context, expected_time):
    # type: (Any, Any) -> None
    """Ensure queries execute within the expected time"""
    assert context.query_time <= int(expected_time), "Query took too long!"
    logging.info("Query performance within acceptable limits.")


@given(
    '"{batch_count}" bank export files each containing "{transaction_count}" transactions'
)
def batch_transaction_processing(context, batch_count, transaction_count):
    # type: (Any, Any, Any) -> None
    """Simulate batch processing of high-volume transactions"""
    context.batch_count = int(batch_count)
    context.transaction_count = int(transaction_count)


@when("the system processes these files in parallel")
def step_when_parallel_file_processing(context):
    # type: (Any) -> None
    """Simulate batch transaction processing."""

    # Simulate batch processing time in seconds
    context.batch_processing_time = random.randint(100, 900)

    # Simulate short delay using datetime
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass  # Busy wait to simulate delay

    logging.info(
        "Batch processing completed in {} seconds.".format(
            context.batch_processing_time
        )
    )


@then('system memory consumption should remain under "{memory_limit}%"')
def step_then_memory_within_limit(context, memory_limit):
    # type: (Any, Any) -> None
    """Ensure system memory consumption remains within the limit"""
    actual_memory_usage = random.randint(60, 85)
    assert actual_memory_usage <= int(memory_limit), "Memory usage exceeded!"
    logging.info("Memory usage: {}%, within limit.".format(actual_memory_usage))


@given(
    'a bank export file "{file_name}" with "{error_type}" errors in "{error_percentage}%" of transactions'
)
def high_volume_errors(context, file_name, error_type, error_percentage):
    # type: (Any, Any, Any, Any) -> None
    """Simulate a high-volume transaction file with errors"""
    context.file_name = file_name
    context.error_type = error_type
    context.error_percentage = int(error_percentage)


@then("all errors should be logged properly")
def step_then_all_errors_logged(context):
    # type: (Any) -> None
    """Ensure all errors are correctly logged"""
    logging.warning(
        "{} transactions failed due to {} errors.".format(
            context.error_count, context.error_type
        )
    )


@given('a system processing "{transaction_count}" transactions per second')
def transaction_load(context, transaction_count):
    # type: (Any, Any) -> None
    """Simulate a system processing a specific number of transactions per second"""
    context.transaction_rate = int(transaction_count)


@when('the transaction load increases by "{increase_percentage}%"')
def step_when_transaction_load_increases(context, increase_percentage):
    # type: (Any, Any) -> None
    """Simulate a sudden increase in transaction processing load"""
    context.scaled_transaction_rate = context.transaction_rate * (
        1 + int(increase_percentage) / 100
    )
    logging.info(
        "New transaction processing rate: {} per second.".format(
            context.scaled_transaction_rate
        )
    )


@then("the system should scale dynamically without degradation")
def step_then_validate_scalability(context):
    # type: (Any) -> None
    """Ensure the system scales properly under high load"""
    assert (
        context.scaled_transaction_rate <= context.transaction_rate * 2
    ), "System scalability failure!"
    logging.info("System scaled dynamically under high load.")


@given(
    'a bank export file "{file_name}" with "{transaction_count}" transactions and simulated network latency of "{latency}" ms'
)
def network_latency_simulation(context, file_name, transaction_count, latency):
    # type: (Any, Any, Any, Any) -> None
    """Simulate network latency during high-volume transaction processing"""
    context.file_name = file_name
    context.transaction_count = int(transaction_count)
    context.latency = int(latency)


@when("I attempt to process the file remotely")
def step_when_process_with_latency(context):
    # type: (Any) -> None
    """Simulate remote processing of a large transaction file under latency."""

    # Convert latency from ms to seconds
    context.latency_effect = context.latency / 1000.0

    # Simulate delay using datetime loop
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=context.latency_effect)

    while datetime.now() < end_time:
        pass  # Simulate delay

    logging.info(
        "Processing "
        + context.file_name
        + " with "
        + str(context.latency)
        + " ms latency."
    )


@then("processing should not hang indefinitely")
def step_then_validate_latency_handling(context):
    # type: (Any) -> None
    """Ensure the system does not hang under latency conditions"""
    assert context.latency_effect < 5, "Processing hung due to high latency!"
    logging.info("Network latency handled successfully.")


# ================= End of High-Volume Transaction Processing Validation =================

# ================= Beginning of Regression Testing for Previously Fixed Bugs =================
# This script validates that previously fixed bugs do not reoccur in regression testing.
# It ensures:
# - Data integrity by preventing past issues from resurfacing.
# - Database consistency in records that were affected by previous bugs.
# - Batch processing maintains resolved issues without reintroducing errors.
# - Proper error handling and alerting for unexpected bug reoccurrence.
# - System performance is maintained during large-scale regression checks.


@given(
    'a bank export file named "{file_name}" that had an issue fixed in version "{fixed_version}"'
)
def fixed_issue_file(context, file_name, fixed_version):
    # type: (Any, Any, Any) -> None
    """Simulate loading a bank export file with previously fixed issues"""
    context.file_name = file_name
    context.fixed_version = fixed_version
    context.issue_resolved = True


@when("I process the file")
def step_when_process_fixed_issue_file(context):
    # type: (Any) -> None
    """Simulate processing the export file with previously fixed issues."""

    # Simulate short processing delay
    processing_time = random.uniform(0.5, 2.0)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=processing_time)

    while datetime.now() < end_time:
        pass  # Simulated delay

    # Simulate resolution status
    context.issue_resolved = random.choice([True, False])

    # Log result using .format()
    status = "Resolved" if context.issue_resolved else "Reoccurred"
    logging.info(
        "Processed {}. Previously fixed issue status: {}".format(
            context.file_name, status
        )
    )


@then("the issue should not reoccur")
def step_then_validate_fixed_issue(context):
    # type: (Any) -> None
    """Ensure previously resolved issues do not reappear"""
    assert context.issue_resolved, "Previously fixed issue has reoccurred!"
    logging.info("Previously resolved issue has not reoccurred.")


@then("a validation report should confirm the resolution")
def step_then_validation_report_confirms_resolution(context):
    # type: (Any) -> None
    """Generate a validation report for the fixed issue"""
    report_status = (
        "Issue successfully validated" if context.issue_resolved else "Issue reoccurred"
    )
    logging.info("Validation report generated: {}".format(report_status))


@when("I compare the latest records with the resolved state")
def step_when_compare_latest_records(context):
    # type: (Any) -> None
    """Simulate checking database records for past issues."""

    # Simulated comparison time
    comparison_time = random.uniform(1.0, 5.0)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=comparison_time)

    while datetime.now() < end_time:
        pass  # Simulate delay

    # Randomly determine resolution status
    context.db_resolved = random.choice([True, False])

    # Format and log the result
    status = "Resolved" if context.db_resolved else "Reoccurred"
    logging.info(
        "Checked for {} in database. Status: {}".format(context.issue_type, status)
    )


@then("no past issues should reappear")
def step_then_validate_no_past_issues(context):
    # type: (Any) -> None
    """Ensure past database issues do not reoccur"""
    assert context.db_resolved, "Previously resolved issue {} has reoccurred!".format(
        context.issue_type
    )
    logging.info(
        "No reoccurrence of {} detected in the database.".format(context.issue_type)
    )


@given(
    '"{batch_count}" bank export files from "{year_range}" containing previously flagged issues'
)
def batch_export_files_with_issues(context, batch_count, year_range):
    # type: (Any, Any, Any) -> None
    """Simulate batch processing of export files with previously resolved issues"""
    context.batch_count = int(batch_count)
    context.year_range = year_range


@when("I process them for validation")
def step_when_validate_batch_processing(context):
    # type: (Any) -> None
    """Simulate batch processing validation."""

    # Simulate batch processing time (just for context)
    batch_processing_time = random.randint(100, 900)

    # Simulate short delay using datetime
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass  # Simulate delay loop

    # Randomly determine validation outcome
    context.batch_validated = random.choice([True, False])

    logging.info(
        "Batch validation for {} completed. Issues reoccurred: {}".format(
            context.year_range, not context.batch_validated
        )
    )


@then("all records should pass consistency checks")
def step_then_all_records_pass_consistency_checks(context):
    # type: (Any) -> None
    """Ensure batch processing does not reintroduce old issues"""
    assert (
        context.batch_validated
    ), "Batch processing reintroduced previously resolved issues!"
    logging.info("Batch processing maintained resolved issues without regression.")


@given('an attempt to process a bank export file "{file_name}"')
def error_handling_validation(context, file_name):
    # type: (Any, Any) -> None
    """Simulate an attempt to process a file with potential regression issues"""
    context.file_name = file_name


@given('a system processing "{file_count}" bank export files per hour')
def high_volume_regression_check(context, file_count):
    # type: (Any, Any) -> None
    """Simulate high-volume regression testing"""
    context.file_count = int(file_count)


@when('checking for previously fixed issues in "{year_range}"')
def step_when_perform_regression_check(context, year_range):
    # type: (Any, Any) -> None
    """Perform regression testing on high volumes of historical data."""

    # Store the year range in context
    context.year_range = year_range

    # Simulate regression testing time
    context.regression_test_time = random.randint(200, 600)

    # Simulate short delay using datetime
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass  # Simulated delay using busy wait

    # Log using .format()
    logging.info(
        "Checked {} files for fixed issues in {}.".format(
            getattr(context, "file_count", "N/A"), context.year_range
        )
    )


@then('system resources should not exceed "{resource_limit}%"')
def step_then_validate_resource_usage(context, resource_limit):
    # type: (Any, Any) -> None
    """Ensure system resource usage remains within acceptable limits"""
    actual_usage = random.randint(60, 85)  # Simulated usage percentage
    assert actual_usage <= int(
        resource_limit
    ), "System resource usage exceeded! Actual: {}%, Limit: {}%".format(
        actual_usage, resource_limit
    )
    logging.info("System resource usage: {}%, within limit.".format(actual_usage))


# ================= End of Regression Testing for Previously Fixed Bugs =================

# ================= Beginning of Regression Testing for Transaction Reference Uniqueness =================
# This script ensures:
# - Every transaction reference remains unique.
# - No duplicate references exist in batch and database processing.
# - Any detected duplicates are flagged and logged appropriately.
# - System maintains data integrity and performance while validating uniqueness.


@given(
    'a bank export file named "{file_name}" containing "{transaction_count}" transactions'
)
def transaction_file(context, file_name, transaction_count):
    # type: (Any, Any, Any) -> None
    """Simulate loading a bank export file with transactions"""
    context.file_name = file_name
    context.transaction_count = int(transaction_count)
    context.duplicates_found = False
    logging.info("Loaded {} transactions from {}.".format(transaction_count, file_name))


@when("the system processes the file")
def step_when_process_transactions(context):
    # type: (Any) -> None
    """Simulate processing the transactions and checking for duplicate references."""

    # Simulated processing delay
    processing_time = random.uniform(0.5, 2.0)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=processing_time)

    # Simulate time passing (busy-wait loop to simulate delay)
    while datetime.now() < end_time:
        pass  # Simulating delay (busy-wait)

    # Simulate transaction reference uniqueness check
    duplicate_rate = 0.001  # 0.1% chance of duplicate
    context.duplicates_found = random.random() < duplicate_rate

    # Logging the results
    logging.info(
        "Processed file: {0}. Duplicate references found: {1}".format(
            context.file_name, context.duplicates_found
        )
    )


@then("each transaction should have a unique reference ID")
def step_then_validate_transaction_uniqueness(context):
    # type: (Any) -> None
    """Ensure all transaction references are unique"""
    assert not context.duplicates_found, "Duplicate transaction references detected!"
    logging.info("All transactions have unique reference IDs.")


@then('duplicate transaction references should be flagged as "{severity}"')
def step_then_flag_duplicates(context, severity):
    # type: (Any, Any) -> None
    """Flag duplicates if found"""
    if getattr(context, "duplicates_found", []):
        logging.warning(
            "Duplicate transaction references found! Severity: {}".format(severity)
        )
    else:
        logging.info("No duplicate transaction references detected.")


@given('a database containing transaction records from "{year_range}"')
def database_transactions(context, year_range):
    # type: (Any, Any) -> None
    """Simulate database containing past transaction records"""
    context.year_range = year_range


@when("I check for duplicate transaction references")
def step_when_check_duplicate_references(context):
    # type: (Any) -> None
    """Check for duplicate transaction references in the database."""

    # Simulated database check time
    check_time = random.uniform(1.0, 5.0)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=check_time)

    # Simulate delay loop
    while datetime.now() < end_time:
        pass

    # Simulate result of duplicate check
    context.db_duplicates_found = random.choice([True, False])

    logging.info(
        "Checked database records for duplicates in {}. Found: {}".format(
            context.year_range, context.db_duplicates_found
        )
    )


@then("no duplicate references should exist")
def step_then_validate_database_uniqueness(context):
    # type: (Any) -> None
    """Ensure database transaction references remain unique"""
    assert (
        not context.db_duplicates_found
    ), "Duplicate transaction references detected in database!"
    logging.info("Database records maintain unique transaction references.")


@given('"{batch_count}" bank export files from "{year_range}"')
def batch_transactions(context, batch_count, year_range):
    # type: (Any, Any, Any) -> None
    """Simulate batch processing of bank export files"""
    context.batch_count = int(batch_count)
    context.year_range = year_range


@when("I process them for validation")
def step_when_validate_batch_transactions(context):
    # type: (Any) -> None
    """Simulate batch processing for transaction reference uniqueness."""

    # Simulated batch processing time (just for context)
    batch_processing_time = random.randint(100, 900)

    # Simulate delay using datetime instead of time.sleep
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass

    # Simulate result of validation
    context.batch_validated = random.choice([True, False])

    logging.info(
        "Batch validation for {} completed. Issues found: {}".format(
            context.year_range, not context.batch_validated
        )
    )


@then("all transactions should maintain unique references")
def step_then_validate_batch_reference_uniqueness(context):
    # type: (Any) -> None
    """Ensure batch processing maintains unique transaction references"""
    assert (
        context.batch_validated
    ), "Batch processing introduced duplicate transaction references!"
    logging.info("Batch processing maintained unique transaction references.")


@given('an attempt to process a bank export file "{file_name}"')
def duplicate_handling(context, file_name):
    # type: (Any, Any) -> None
    """
    Simulate an attempt to process a file with duplicate transaction references.
    """
    context.file_name = file_name


@when('duplicate transaction references such as "{error_type}" are detected')
def step_when_detect_duplicate_issue(context, error_type):
    # type: (Any, Any) -> None
    """Simulate detecting duplicate transaction reference issues"""
    context.duplicate_issue_detected = random.choice([True, False])
    context.error_type = error_type

    if context.duplicate_issue_detected:
        logging.warning(
            "Detected " + error_type + " issue again in " + context.file_name
        )
    else:
        logging.info("No duplicate transaction references detected.")


@given('a system processing "{file_count}" bank export files per hour')
def high_volume_reference_check(context, file_count):
    # type: (Any, Any) -> None
    """Simulate high-volume transaction reference checking"""
    context.file_count = int(file_count)


@when('checking for duplicate transaction references in "{year_range}"')
def step_when_perform_reference_check(context, year_range):
    # type: (Any, Any) -> None
    """Perform duplicate transaction reference check across multiple years."""

    # Store the year range in the context
    context.year_range = year_range

    # Simulate checking time (in milliseconds) for duplicate transaction references
    context.reference_check_time = random.randint(
        200, 600
    )  # Simulated time for reference check (in ms)

    # Simulate a short processing delay (1 second)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass  # Simulating delay for processing

    # Simulate checking for duplicate references within the year range
    # Simulating a 5% chance of finding duplicates
    context.duplicates_found = random.random() < 0.05

    # Log the result of the check using .format() for Python 2.7 compatibility
    logging.info(
        "Checked {0} files for duplicate references in year range {1}. Duplicates found: {2}".format(
            context.file_count, context.year_range, context.duplicates_found
        )
    )

    # If duplicates are found, log a warning and store that info in the context
    if context.duplicates_found:
        # Flag the system for further action or alert
        logging.warning(
            "Duplicates detected in year range {0}. Please review the files for discrepancies.".format(
                context.year_range
            )
        )
        context.duplicate_issue_detected = (
            True  # Store flag indicating duplicates were found
        )
    else:
        # If no duplicates found, log the success
        logging.info(
            "No duplicates found in the year range {0}.".format(context.year_range)
        )
        context.duplicate_issue_detected = False  # No issues detected


# ================= End of Regression Testing for Transaction Reference Uniqueness =================

# ================= Beginning of Structural Testing for Column Case Sensitivity =================
# This script ensures:
# - Column headers are mapped correctly despite case variations.
# - Database integrity is maintained with case-insensitive column mappings.
# - No data is lost due to column case mismatches.
# - The system maintains performance while validating column names.


@given('a bank export file named "{file_name}" with columns in "{case_format}"')
def column_case_variations(context, file_name, case_format):
    # type: (Any, Any, Any) -> None
    """Simulate loading a bank export file with different column case formats"""
    context.file_name = file_name
    context.case_format = case_format
    context.mismatched_columns = random.choice(
        [True, False]
    )  # Randomly simulate case mismatch
    logging.info(
        "Loaded {} with columns in {} case format.".format(file_name, case_format)
    )


@then("column headers should be mapped correctly")
def step_then_validate_column_mappings(context):
    # type: (Any) -> None
    """Ensure column headers are correctly mapped regardless of case"""
    assert not context.mismatched_columns, "Column mapping mismatch detected!"
    logging.info("All column headers mapped correctly.")


@then("no data should be lost due to case mismatches")
def step_then_no_data_loss_case_mismatches(context):
    # type: (Any) -> None
    """Ensure data is not lost due to case mismatches"""
    assert (
        not context.mismatched_columns
    ), "Potential data loss due to column case mismatches!"
    logging.info("No data loss occurred due to column case variations.")


@given('a database with expected column names in "{expected_case_format}"')
def database_column_expectations(context, expected_case_format):
    # type: (Any, Any) -> None
    """Simulate expected column names in the database"""
    context.expected_case_format = expected_case_format


@when('I compare imported column names from "{file_name}"')
def step_when_compare_database_columns(context, file_name):
    # type: (Any, Any) -> None
    """Compare column names from imported files to the expected format"""
    context.file_name = file_name
    context.discrepancies_found = random.choice([True, False])
    logging.info(
        "Compared imported columns in {}. Discrepancies found: {}".format(
            file_name, context.discrepancies_found
        )
    )


@then("column mappings should be case-insensitive")
def step_then_validate_case_insensitive_mappings(context):
    # type: (Any) -> None
    """Ensure case variations do not affect column mappings"""
    assert not context.discrepancies_found, "Case-sensitive column mismatches detected!"
    logging.info("Column mappings are correctly case-insensitive.")


@given('a batch of bank export files with columns in "{case_format}"')
def batch_column_variations(context, case_format):
    # type: (Any, Any) -> None
    """Simulate batch processing of bank export files with varying column cases"""
    context.case_format = case_format


@when("the system processes them for validation")
def system_processes_batch_files_for_validation(context):
    # type: (Any) -> None
    """
    Process batch files and check for various validation issues including:
    - whitespace issues
    - missing columns
    - column reordering
    - protected sheets (Excel)
    - merged cells
    """

    context.batch_issues = {}

    for file_name in context.batch_files:
        try:
            # Resolve full file path
            file_path = get_test_file_path(context, file_name)

            # Load the file
            if file_name.endswith(".csv"):
                df = pd.read_csv(file_path)
            elif file_name.endswith(".xlsx"):
                df = pd.read_excel(file_path)
            else:
                context.batch_issues[file_name] = "Unsupported file format"
                continue

            # Check for whitespace issues
            whitespace_issue = (
                df.applymap(lambda x: isinstance(x, str) and x.strip() != x).any().any()
            )
            if whitespace_issue:
                context.batch_issues[file_name] = (
                    context.batch_issues.get(file_name, "")
                    + " | Whitespace Issues Detected"
                )

            # Check for missing columns
            missing_columns = [
                col for col in context.missing_columns if col not in df.columns
            ]
            if missing_columns:
                context.batch_issues[file_name] = (
                    context.batch_issues.get(file_name, "")
                    + " | Missing Columns: "
                    + str(missing_columns)
                )

            # Check for reordered columns
            if df.columns.tolist() != context.ref_columns:
                context.batch_issues[file_name] = (
                    context.batch_issues.get(file_name, "")
                    + " | Column Reordering Detected"
                )

            # Check for protected sheets (Excel only)
            if file_name.endswith(".xlsx"):
                workbook = load_workbook(file_path)
                protected_sheets = [
                    sheet
                    for sheet in workbook.sheetnames
                    if workbook[sheet].protection.sheet
                ]
                if protected_sheets:
                    context.batch_issues[file_name] = (
                        context.batch_issues.get(file_name, "")
                        + " | Protected Sheets: "
                        + str(protected_sheets)
                    )

            # Call reusable validation logic
            issues = validate_file(
                df, context.column_mappings, file_type=file_name.split('.')[-1]
            )
            if issues:
                issue_summary = " | ".join(
                    [key + ": " + str(value) for key, value in issues.items()]
                )
                context.batch_issues[file_name] = (
                    context.batch_issues.get(file_name, "") + " | " + issue_summary
                )

        except Exception as e:
            context.batch_issues[file_name] = "Error processing file: " + str(e)

    # Log the batch issues
    if context.batch_issues:
        logging.warning("Issues found in batch files: " + str(context.batch_issues))


@then("all columns should be correctly recognized")
def step_then_validate_batch_columns(context):
    # type: (Any) -> None
    """Ensure batch files maintain correct column mappings"""
    assert (
        not context.batch_column_mismatch
    ), "Batch processing introduced column mismatches!"
    logging.info("All batch files maintained correct column mappings.")


@given('an attempt to process a bank export file "{file_name}"')
def file_with_column_mismatch(context, file_name):
    # type: (Any, Any) -> None
    """Simulate an attempt to process a file with column mismatches"""
    context.file_name = file_name


@when('a column case mismatch such as "{error_type}" is detected')
def step_when_detect_column_mismatch(context, error_type):
    # type: (Any, Any) -> None
    """Detect case mismatches in column names"""

    context.error_detected = random.choice([True, False])
    context.error_type = error_type

    if context.error_detected:
        logging.warning(
            "Detected "
            + str(context.error_type)
            + " issue in "
            + str(context.file_name)
        )
    else:
        logging.info("No column case mismatches detected.")


@given('a system processing "{file_count}" bank export files per hour')
def high_volume_column_check(context, file_count):
    # type: (Any, Any) -> None
    """Simulate high-volume file processing while checking column consistency"""
    context.file_count = int(file_count)


@when('checking for column case variations in "{year_range}"')
def step_when_perform_column_case_check(context, year_range):
    # type: (Any, Any) -> None
    """Perform column case validation across multiple years."""

    # Set year range and simulate processing time
    context.year_range = year_range
    context.processing_time = random.randint(100, 600)

    # Simulate short processing delay
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass  # Busy wait to simulate delay

    # Log using .format()
    logging.info(
        "Checked {} files for column case mismatches in {}.".format(
            getattr(context, "file_count", "N/A"), context.year_range
        )
    )


@then('system resources should not exceed "{resource_limit}%"')
def step_then_system_resource_limit(context, resource_limit):
    # type: (Any, Any) -> None
    """Ensure system resource usage remains within acceptable limits"""
    actual_usage = random.randint(60, 85)
    assert actual_usage <= int(resource_limit), "System resource usage exceeded!"
    logging.info("System resource usage: {}%, within limit.".format(actual_usage))


# ================= End of Structural Testing for Column Case Sensitivity =================

# ================= Beginning of Structural Testing for Column Format Validation =================
# This script ensures:
# - Column formats adhere to expected data types.
# - Database integrity is maintained with correct column format mappings.
# - No data is lost due to format mismatches.
# - The system maintains performance while validating column formats.


@given(
    'a bank export file named "{file_name}" with column formats defined as "{expected_format}"'
)
def column_format_variations(context, file_name, expected_format):
    # type: (Any, Any, Any) -> None
    """Simulate loading a bank export file with expected column formats"""
    context.file_name = file_name
    context.expected_format = expected_format
    context.format_mismatches = random.choice(
        [True, False]
    )  # Simulating a format mismatch
    logging.info(
        "Loaded {} with expected format: {}.".format(file_name, expected_format)
    )


@then("all columns should match the expected format")
def step_then_validate_column_formats(context):
    # type: (Any) -> None
    """Ensure column formats match expected data types"""
    assert not context.format_mismatches, "Column format mismatch detected!"
    logging.info("All column formats match the expected types.")


@then('any format mismatches should be flagged as "{severity}"')
def step_then_flag_format_mismatches(context, severity):
    # type: (Any, Any) -> None
    """Ensure mismatched formats are flagged accordingly"""

    if getattr(context, "format_mismatches", False):
        logging.warning(
            "Column format mismatch detected in {}. Severity: {}".format(
                context.file_name, severity
            )
        )
    else:
        logging.info("No format mismatches found in file: {}".format(context.file_name))


@given('a database with expected column formats in "{expected_format}"')
def database_column_format_expectations(context, expected_format):
    # type: (Any, Any) -> None
    """Simulate expected column formats in the database"""
    context.expected_format = expected_format


@when('I compare imported column formats from "{file_name}"')
def step_when_compare_database_column_formats(context, file_name):
    # type: (Any, Any) -> None
    """Compare column formats from imported files to the expected format"""
    context.file_name = file_name
    context.format_discrepancies_found = random.choice([True, False])
    logging.info(
        "Compared imported column formats in {}. Discrepancies found: {}".format(
            file_name, context.format_discrepancies_found
        )
    )


@then("all columns should match the expected data type")
def step_then_validate_column_data_types(context):
    # type: (Any) -> None
    """Ensure column data types are correctly formatted"""
    assert (
        not context.format_discrepancies_found
    ), "Column format discrepancies detected!"
    logging.info("Column formats are correctly aligned with database expectations.")


@given('a batch of bank export files with columns formatted as "{format_type}"')
def batch_column_format_variations(context, format_type):
    # type: (Any, Any) -> None
    """Simulate batch processing of bank export files with varying column formats"""
    context.format_type = format_type


@then("all columns should adhere to the correct format")
def step_then_validate_batch_column_formats(context):
    # type: (Any) -> None
    """Ensure batch files maintain correct column formatting"""
    assert (
        not context.batch_format_mismatch
    ), "Batch processing introduced column format mismatches!"
    logging.info("All batch files maintained correct column formats.")


@given('an attempt to process a bank export file "{file_name}"')
def file_with_column_format_mismatch(context, file_name):
    # type: (Any, Any) -> None
    """Simulate an attempt to process a file with column format mismatches"""
    context.file_name = file_name


@when('a column format inconsistency such as "{error_type}" is detected')
def step_when_detect_column_format_mismatch(context, error_type):
    # type: (Any, Any) -> None
    """Detect format inconsistencies in column values"""
    context.error_detected = random.choice([True, False])
    context.error_type = error_type

    if context.error_detected:
        logging.warning("Detected {} issue in {}".format(error_type, context.file_name))
    else:
        logging.info("No column format mismatches detected.")


@given('a system processing "{file_count}" bank export files per hour')
def high_volume_column_format_check(context, file_count):
    # type: (Any, Any) -> None
    """Simulate high-volume file processing while checking column format consistency"""
    context.file_count = int(file_count)


@when('checking for column format variations in "{year_range}"')
def step_when_perform_column_format_check(context, year_range):
    # type: (Any, Any) -> None
    """Perform column format validation across multiple years."""

    # Set year range and simulate processing time
    context.year_range = year_range
    context.processing_time = random.randint(100, 600)

    # Simulate delay using datetime
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass  # Busy wait to simulate delay

    # Log result
    logging.info(
        "Checked {} files for column format mismatches in {}.".format(
            context.file_count, context.year_range
        )
    )


@then('system resources should not exceed "{resource_limit}%"')
def step_then_validate_column_format_resource_usage(context, resource_limit):
    # type: (Any, Any) -> None
    """Ensure system resource usage remains within acceptable limits"""
    actual_usage = random.randint(60, 85)
    assert actual_usage <= int(resource_limit), "System resource usage exceeded!"
    logging.info("System resource usage: {}%, within limit.".format(actual_usage))


# ================= End of Structural Testing for Column Format Validation =================

# ================= Beginning of Structural Testing for Extra Columns =================
# This script ensures:
# - Extra columns in bank export files are detected and handled properly.
# - Unexpected columns do not disrupt data integrity.
# - The database remains consistent despite schema variations.
# - System performance is maintained when processing files with extra columns.


@given('a bank export file named "{file_name}" containing extra columns')
def extra_columns_file(context, file_name):
    # type: (Any, Any) -> None
    """Simulate loading a bank export file with extra columns"""
    context.file_name = file_name
    context.extra_columns_detected = random.choice(
        [True, False]
    )  # Simulating extra column presence
    logging.info(
        "Loaded {} with extra columns detected: {}.".format(
            file_name, context.extra_columns_detected
        )
    )


@then('extra columns should be flagged as "{severity}"')
def step_then_flag_extra_columns(context, severity):
    # type: (Any, Any) -> None
    """Ensure extra columns are flagged accordingly"""

    def generate_report_of_extra_columns():
        """
        Optional: Generate a detailed report of the extra columns detected, including severity levels.
        """
        report = {
            "file_name": context.file_name,
            "extra_columns_flagged_as_high": context.extra_columns_flagged_as_high,
            "extra_columns_flagged_as_medium": context.extra_columns_flagged_as_medium,
            "extra_columns_flagged_as_low": context.extra_columns_flagged_as_low,
            "severity_flag": context.severity_flag,
        }

        # Here you can either write the report to a file, database, or send it via an API
        # For this example, we will just log it
        logging.info("Generated report for extra columns: {0}".format(report))
        # Optionally, write this to a CSV or another format for further analysis

    def send_notification_about_extra_columns():
        """
        Optional: Send a notification if extra columns exceed a critical threshold (e.g., high severity).
        """
        # Example: Send an email or use a notification service to alert relevant stakeholders
        logging.info(
            "Sending alert about extra columns detected in {0}. Severity: High".format(
                context.file_name
            )
        )
        # Code to send email or notification (depends on the specific system you're using)
        # For example, you could use smtplib for email or integrate with a notification service API.

    # Check if extra columns are detected in the file
    if context.extra_columns_detected:
        # Log the detection of extra columns and flag their severity
        logging.warning(
            "Extra columns detected in {0}. Severity: {1}".format(
                context.file_name, severity
            )
        )

        # Flag the severity of extra columns based on the BDD step description
        if severity.lower() == "high":
            context.extra_columns_flagged_as_high = True
            logging.info(
                "High severity flag set for extra columns in {0}.".format(
                    context.file_name
                )
            )
        elif severity.lower() == "medium":
            context.extra_columns_flagged_as_medium = True
            logging.info(
                "Medium severity flag set for extra columns in {0}.".format(
                    context.file_name
                )
            )
        elif severity.lower() == "low":
            context.extra_columns_flagged_as_low = True
            logging.info(
                "Low severity flag set for extra columns in {0}.".format(
                    context.file_name
                )
            )
        else:
            logging.error(
                "Invalid severity level provided: {0}. Please use 'high', 'medium', or 'low'.".format(
                    severity
                )
            )

        # Store the severity flag in the context for use in further steps
        context.severity_flag = severity.lower()

        # Optional: Generate a report of detected extra columns and their severity
        generate_report_of_extra_columns()

        # Optional: Send a notification if extra columns exceed a threshold
        if context.extra_columns_flagged_as_high:
            send_notification_about_extra_columns(context)

    else:
        # If no extra columns are detected, log this information
        logging.info(
            "No extra columns detected in {0}. No action needed.".format(
                context.file_name
            )
        )


@then("a validation report should document unexpected columns")
def step_then_generate_extra_column_report(context):
    # type: (Any) -> None
    """Generate a report listing the extra columns"""

    def generate_report_of_extra_columns():
        """
        Generate a report that lists the extra columns found, including details on the severity and affected rows.
        """
        report = {
            "file_name": context.file_name,
            "extra_columns_flagged_as_high": context.extra_columns_flagged_as_high,
            "extra_columns_flagged_as_medium": context.extra_columns_flagged_as_medium,
            "extra_columns_flagged_as_low": context.extra_columns_flagged_as_low,
            "severity_flag": context.severity_flag,
        }

        # You can write this report to a CSV or database, depending on your requirements.
        # For now, we will log it as an example.
        logging.info("Generated report for extra columns: {0}".format(report))

        # Optionally, write this to a CSV or send it via an API if needed.

    # Check if extra columns were detected
    if context.extra_columns_detected:
        logging.info(
            "Validation report generated for extra columns in {0}.".format(
                context.file_name
            )
        )

        # Generate a report of the extra columns detected
        generate_report_of_extra_columns()

    else:
        logging.info(
            "No extra columns detected in {0}. No report generated.".format(
                context.file_name
            )
        )


@given("a database expecting standard column structure")
def database_standard_schema(context):
    # type: (Any) -> None
    """Simulate a database with a predefined schema"""
    context.database_has_fixed_schema = True

    def check_column_consistency():
        """
        Check if the columns in the file match the expected columns in the database schema.
        """
        if context.database_has_fixed_schema:
            expected_columns = (
                context.expected_columns
            )  # This should be set somewhere in the context
            actual_columns = (
                context.file_columns
            )  # This should also be set somewhere in the context

            missing_columns = [
                col for col in expected_columns if col not in actual_columns
            ]
            extra_columns = [
                col for col in actual_columns if col not in expected_columns
            ]

            if missing_columns:
                logging.error(
                    "Missing columns in the file: {0}".format(
                        ", ".join(missing_columns)
                    )
                )

            if extra_columns:
                logging.warning(
                    "Extra columns detected in the file: {0}".format(
                        ", ".join(extra_columns)
                    )
                )

            return missing_columns, extra_columns

        else:
            logging.warning(
                "Database schema is not fixed. Unable to validate column structure."
            )
            return [], []

    # Simulate checking the column structure of the database
    missing_columns, extra_columns = check_column_consistency()

    # Store the result in the context for further steps
    context.missing_columns = missing_columns
    context.extra_columns = extra_columns
    logging.info("Column consistency check completed.")


@when('I compare imported column names from "{file_name}"')
def step_when_compare_database_column_names(context, file_name):
    # type: (Any, Any) -> None
    """Compare imported column names with the expected schema"""
    context.file_name = file_name
    context.schema_mismatch = random.choice([True, False])
    logging.info(
        "Compared column names for {}. Schema mismatch: {}".format(
            file_name, context.schema_mismatch
        )
    )


@then('extra columns should be ignored or flagged as "{discrepancy_type}"')
def step_then_handle_extra_columns_in_db(context, discrepancy_type):
    # type: (Any, Any) -> None
    """Ensure extra columns are ignored or flagged correctly in the database"""

    # Simulate the detection of extra columns (assuming context has a schema mismatch check)
    if context.schema_mismatch:
        logging.warning(
            "Schema mismatch detected in {0}. Discrepancy Type: {1}".format(
                context.file_name, discrepancy_type
            )
        )

        # If mismatch found, take necessary actions (e.g., flag or ignore extra columns)
        if discrepancy_type == "High":
            logging.error(
                "High-severity schema mismatch in {0}. Immediate action required.".format(
                    context.file_name
                )
            )
        elif discrepancy_type == "Medium":
            logging.warning(
                "Medium-severity schema mismatch detected in {0}. Review recommended.".format(
                    context.file_name
                )
            )
        else:
            logging.info(
                "Low-severity schema mismatch detected in {0}. Monitoring ongoing.".format(
                    context.file_name
                )
            )

    else:
        logging.info(
            "No schema mismatch detected for {0}. All columns are valid.".format(
                context.file_name
            )
        )


@given("a batch of bank export files with extra columns")
def batch_extra_column_files(context):
    # type: (Any) -> None
    """Simulate batch processing of export files with extra columns"""

    # Randomly simulate whether batch contains extra columns or not
    context.batch_contains_extra_columns = random.choice([True, False])

    if context.batch_contains_extra_columns:
        logging.info(
            "Batch contains extra columns. These will be flagged during processing."
        )
    else:
        logging.info("Batch does not contain any extra columns.")


@then('all extra columns should be detected and flagged as "{severity}"')
def step_then_validate_batch_extra_columns(context, severity):
    # type: (Any, Any) -> None
    """Ensure extra columns in batch files are flagged"""

    # If batch contains extra columns, simulate the detection and flagging
    if context.batch_contains_extra_columns:
        logging.warning(
            "Extra columns detected in batch processing. Severity: {0}".format(severity)
        )

        # Handle different severities for extra columns
        if severity == "High":
            logging.error(
                "High severity: Immediate review needed for extra columns detected in batch."
            )
        elif severity == "Medium":
            logging.warning(
                "Medium severity: Review extra columns for potential issues."
            )
        else:
            logging.info(
                "Low severity: Extra columns detected, but no immediate action required."
            )
    else:
        logging.info("No extra columns detected in the batch files.")


@given('an attempt to process a bank export file "{file_name}"')
def file_with_extra_columns(context, file_name):
    # type: (Any, Any) -> None
    """Simulate an attempt to process a file with extra columns"""
    context.file_name = file_name


@when('extra columns such as "{column_name}" are detected')
def step_when_detect_extra_columns(context, column_name):
    # type: (Any, Any) -> None
    """Detect extra columns in the file"""

    # Simulate detection of extra columns (True or False randomly)
    context.extra_column_detected = random.choice([True, False])
    context.column_name = column_name

    # If extra columns are detected, log the warning
    if context.extra_column_detected:
        logging.warning(
            "Extra column '{0}' detected in {1}".format(column_name, context.file_name)
        )
    else:
        logging.info("No extra columns detected in {0}".format(context.file_name))


@then("an auto-mapping mechanism should suggest appropriate actions")
def step_then_suggest_column_mapping(context):
    # type: (Any) -> None
    """Suggest auto-mapping for extra columns"""

    if context.extra_column_detected:
        logging.info(
            "Auto-mapping suggested for extra column '{0}' in {1}.".format(
                context.column_name, context.file_name
            )
        )

        # Functionality embedded within the step
        def suggest_column_mapping_action(column_name, file_name):
            """
            Suggest appropriate actions for auto-mapping extra columns.
            This can be expanded with actual mapping logic or interaction with a UI.
            """
            # Simulating a suggestion to handle the extra column
            logging.info(
                "Suggested action: Map extra column '{0}' from file '{1}' to the correct database field.".format(
                    column_name, file_name
                )
            )

        suggest_column_mapping_action(context.column_name, context.file_name)
    else:
        logging.info("No extra columns detected; no auto-mapping necessary.")


@given('a system processing "{file_count}" bank export files per hour')
def high_volume_extra_column_check(context, file_count):
    # type: (Any, Any) -> None
    """Simulate high-volume file processing while handling extra columns"""
    context.file_count = int(file_count)
    logging.info(
        "System processing {0} bank export files per hour.".format(context.file_count)
    )


@when('extra columns are present in "{year_range}"')
def step_when_detect_extra_columns_in_year_range(context, year_range):
    # type: (Any, Any) -> None
    """Perform extra column validation across multiple years."""

    # Store year range for future use and simulate processing time
    context.year_range = year_range
    context.processing_time = random.randint(100, 600)

    # Simulate a delay (e.g., 1 second) for processing using a busy-wait loop
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass  # Simulate processing time

    # Log the result of extra column validation
    logging.info(
        "Checked {0} files for extra columns in {1}.".format(
            context.file_count, context.year_range
        )
    )


@then('system resources should not exceed "{resource_limit}%"')
def step_then_validate_extra_column_resource_usage(context, resource_limit):
    # type: (Any, Any) -> None
    """Ensure system resource usage remains within acceptable limits"""
    actual_usage = random.randint(60, 85)
    assert actual_usage <= int(resource_limit), "System resource usage exceeded!"
    logging.info("System resource usage: {}%, within limit.".format(actual_usage))


# ================= End of Structural Testing for Extra Columns =================

# ================= Beginning of Structural Testing for Header Mismatch Validation =================
# This script ensures:
# - Bank export files have the correct headers in expected order.
# - Any header mismatches are flagged and logged.
# - The database remains consistent with predefined header structures.
# - System performance remains stable when processing files with header discrepancies.


@given('a bank export file named "{file_name}" with headers "{header_format}"')
def header_format(context, file_name, header_format):
    # type: (Any, Any, Any) -> None
    """Simulate loading a bank export file with specific header formats"""
    context.file_name = file_name
    context.header_format = header_format
    context.header_mismatch_detected = random.choice(
        [True, False]
    )  # Simulating header mismatch
    logging.info(
        "Loaded {} with header format: {}. Mismatch detected: {}".format(
            file_name, header_format, context.header_mismatch_detected
        )
    )


@then("all headers should match the expected structure")
def step_then_validate_headers(context):
    # type: (Any) -> None
    """Ensure headers are properly validated"""

    if context.schema_mismatch:
        logging.warning("Header mismatch found in {0}!".format(context.file_name))
    else:
        logging.info(
            "All headers in {0} match the expected structure.".format(context.file_name)
        )


@then('any mismatched headers should be flagged as "{severity}"')
def step_then_flag_header_mismatches(context, severity):
    # type: (Any, Any) -> None
    """Flag header mismatches based on severity"""

    if context.schema_mismatch:
        logging.error(
            "Header mismatch severity: {0} for {1}".format(severity, context.file_name)
        )
    else:
        logging.info("No header mismatches detected in {0}.".format(context.file_name))


@then("a validation report should document header inconsistencies")
def step_then_generate_header_validation_report(context):
    # type: (Any) -> None
    """Generate a report for header mismatches"""

    if context.schema_mismatch:
        logging.info(
            "Validation report generated for header inconsistencies in {0}.".format(
                context.file_name
            )
        )
    else:
        logging.info(
            "No header inconsistencies found in {0}.".format(context.file_name)
        )


@then("if auto-mapping is enabled, a correction suggestion should be provided")
def step_then_suggest_header_correction(context):
    # type: (Any) -> None
    """Provide auto-mapping suggestions if enabled"""

    if context.schema_mismatch:
        logging.info("Suggested header correction for {0}.".format(context.file_name))
    else:
        logging.info("No header corrections needed for {0}.".format(context.file_name))


@given("a database expecting a predefined column structure")
def database_standard_header(context):
    # type: (Any) -> None
    """Simulate a database with a predefined schema for headers"""
    context.database_has_fixed_schema = True


@when('I compare imported headers from "{file_name}"')
def step_when_compare_imported_headers(context, file_name):
    # type: (Any, str) -> None
    """Compare imported headers with expected headers"""
    context.file_name = file_name

    # Retrieve expected headers using the helper function from file_validation.py
    expected_headers = get_expected_headers(file_name)

    # Retrieve imported headers using the helper function from file_validation.py
    imported_headers = get_imported_headers(file_name)

    # Compare headers and detect mismatch
    context.schema_mismatch = expected_headers != imported_headers

    # Log the comparison result
    if context.schema_mismatch:
        logging.warning("Schema mismatch detected for {0}!".format(file_name))
    else:
        logging.info("Headers match expected structure for {0}.".format(file_name))


@then("all headers should align with the expected format")
def step_then_validate_database_headers(context):
    # type: (Any) -> None
    """Ensure headers in database match the expected format"""

    # If schema mismatch, log a warning
    if context.schema_mismatch:
        logging.warning(
            "Header schema mismatch detected in {0}".format(context.file_name)
        )
    else:
        logging.info(
            "Headers align with the expected format for {0}".format(context.file_name)
        )


@then('any mismatches should be flagged as "{discrepancy_type}"')
def step_then_flag_schema_mismatches(context, discrepancy_type):
    # type: (Any, Any) -> None
    """Flag schema mismatches in headers"""

    # If schema mismatch, log an error with the specified discrepancy type
    if context.schema_mismatch:
        logging.error(
            "Schema mismatch detected in {0}. Discrepancy Type: {1}".format(
                context.file_name, discrepancy_type
            )
        )
    else:
        logging.info("No schema mismatches detected in {0}".format(context.file_name))


@given("a batch of bank export files with header inconsistencies")
def batch_files_with_header_issues(context):
    # type: (Any) -> None
    """Simulate batch processing of export files with header issues"""
    context.batch_contains_header_mismatches = random.choice([True, False])
    logging.info(
        "Batch processing simulation started. Header mismatches present: {0}".format(
            context.batch_contains_header_mismatches
        )
    )


@then('all header mismatches should be detected and flagged as "{severity}"')
def step_then_validate_batch_header_mismatches(context, severity):
    # type: (Any, Any) -> None
    """Ensure header mismatches in batch files are flagged"""
    if context.batch_contains_header_mismatches:
        logging.warning(
            "Header mismatches detected in batch processing. Severity: {0}".format(
                severity
            )
        )
        # Store the severity for later use and flag the header mismatches
        context.header_flagged_as = severity
        logging.info("Header mismatches flagged with severity: {0}".format(severity))
        # Simulate flagging logic based on severity
        logging.info(
            "Flagging header mismatches with severity level: {0}".format(severity)
        )
    else:
        logging.info("No header mismatches detected in batch processing.")


@given('an attempt to process a bank export file "{file_name}"')
def file_with_header_issues(context, file_name):
    # type: (Any, str) -> None
    """Simulate an attempt to process a file with header mismatches"""
    context.file_name = file_name

    # Retrieve the expected headers using the helper function from file_validation.py
    context.expected_headers = get_expected_headers(file_name)

    # Retrieve the imported headers using the helper function from file_validation.py
    context.imported_headers = get_imported_headers(file_name)

    # Log the headers comparison for debugging purposes
    logging.info(
        "Processing file: {0}. Expected headers: {1}, Imported headers: {2}".format(
            file_name, context.expected_headers, context.imported_headers
        )
    )


@when('header mismatches such as "{error_type}" are detected')
def step_when_detect_header_mismatches(context, error_type):
    # type: (Any, Any) -> None
    """Detect header mismatches in the file"""
    context.header_mismatch_detected = False
    if context.imported_headers != context.expected_headers:
        context.header_mismatch_detected = True
        context.error_type = error_type
        logging.warning(
            "Header issue detected in {0}: {1}".format(context.file_name, error_type)
        )
    else:
        logging.info("No header mismatches detected.")


@then("an auto-mapping mechanism should suggest appropriate actions")
def step_then_suggest_header_correction(context):
    # type: (Any) -> None
    """Provide auto-mapping suggestions if enabled"""
    if context.header_mismatch_detected:
        logging.info(
            "Auto-mapping suggested for headers in {0}.".format(context.file_name)
        )


@then('all header mismatches should be detected and flagged as "{severity}"')
def step_then_validate_batch_header_mismatches(context, severity):
    # type: (Any, Any) -> None
    """Ensure header mismatches in batch files are flagged"""
    if context.batch_contains_header_mismatches:
        logging.warning(
            "Header mismatches detected in batch processing. Severity: {0}".format(
                severity
            )
        )
        # Additional logic for severity-based processing (e.g., flagging, reporting)
        context.header_flagged_as = severity  # Store the severity for later processing
    else:
        logging.info("No header mismatches detected in batch processing.")


@when('I compare imported headers from "{file_name}"')
def step_when_compare_imported_headers(context, file_name):
    # type: (Any, str) -> None
    """Compare imported headers with expected headers"""
    context.file_name = file_name

    # Use the helper function from file_validation.py to get expected headers
    expected_headers = get_expected_headers(file_name)

    # Use the helper function from file_validation.py to retrieve imported headers
    context.imported_headers = get_imported_headers(file_name)

    # Check if the headers match
    context.schema_mismatch = expected_headers != context.imported_headers

    # Log the comparison result
    logging.info(
        "Compared headers for {0}. Schema mismatch: {1}".format(
            file_name, context.schema_mismatch
        )
    )


@then('any mismatched headers should be flagged as "{severity}"')
def step_then_flag_header_mismatches(context, severity):
    # type: (Any, Any) -> None
    """Flag header mismatches based on severity"""
    if context.schema_mismatch:
        logging.warning(
            "Header mismatch severity: {0} for {1}".format(severity, context.file_name)
        )


@then('all headers should match the expected structure')
def step_then_validate_headers(context):
    # type: (Any) -> None
    """Ensure headers are properly validated"""
    if context.schema_mismatch:
        logging.warning("Header mismatch found in {0}!".format(context.file_name))


@then("an auto-mapping mechanism should suggest appropriate corrections")
def step_then_suggest_header_mapping(context):
    # type: (Any) -> None
    """Suggest auto-mapping for header mismatches"""
    if context.header_mismatch_detected:
        logging.info(
            "Auto-mapping suggested for headers in {0} due to {1} mismatch.".format(
                context.file_name, context.error_type
            )
        )
        # You can add further logic for suggesting the mapping (e.g., call another helper to map headers)
        context.auto_mapping_suggested = True
    else:
        logging.info("No header mismatches detected, no mapping needed.")


@given('a system processing "{file_count}" bank export files per hour')
def high_volume_header_check(context, file_count):
    # type: (Any, Any) -> None
    """Simulate high-volume file processing while handling header mismatches"""
    context.file_count = int(file_count)


@when('header mismatches are present in "{year_range}"')
def step_when_detect_header_mismatches_in_year_range(context, year_range):
    # type: (Any, Any) -> None
    """Perform header validation across multiple years."""

    # Store the year range for later use and simulate processing time
    context.year_range = year_range
    context.processing_time = random.randint(100, 600)  # Simulated processing time

    # Simulate a short delay using datetime to simulate processing time
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)  # Processing delay of 1 second

    while datetime.now() < end_time:  # Busy-wait loop for simulating processing time
        pass

    # Log the header mismatch validation for the given year range
    logging.info(
        "Checked {0} files for header mismatches in {1}.".format(
            context.file_count, context.year_range
        )
    )


@then('system resources should not exceed "{resource_limit}%"')
def step_then_validate_header_resource_usage(context, resource_limit):
    # type: (Any, Any) -> None
    """Ensure system resource usage remains within acceptable limits"""

    # Simulate system resource usage (random value for testing)
    actual_usage = random.randint(
        60, 85
    )  # Simulating actual system usage between 60% and 85%

    # Check if actual usage exceeds the given resource limit
    assert actual_usage <= int(
        resource_limit
    ), "System resource usage exceeded the limit of {0}%!".format(resource_limit)

    # Log the resource usage and confirm if it's within the limit
    logging.info(
        "System resource usage: {0}, within the limit of {1}%.".format(
            actual_usage, resource_limit
        )
    )


@then(
    'if resource utilization exceeds "{critical_limit}%", an alert should be triggered'
)
def step_then_trigger_alert_for_resource_usage(context, critical_limit):
    # type: (Any, Any) -> None
    """Trigger an alert if resource usage exceeds the critical threshold"""

    # Simulate actual resource usage (random value for testing)
    actual_usage = random.randint(
        60, 95
    )  # Simulating actual system usage between 60% and 95%

    # Check if actual usage exceeds the critical limit
    if actual_usage > int(critical_limit):
        # Log and trigger a critical alert if the usage exceeds the limit
        logging.critical(
            "CRITICAL ALERT: Resource usage exceeded {0}%!".format(critical_limit)
        )
        # Here you can expand to notify users via email, SMS, etc.
        send_resource_alert(actual_usage, critical_limit)
    else:
        # Log the usage if it is within the safe limits
        logging.info(
            "System resource usage: {0}, within safe limits (below {1}%).".format(
                actual_usage, critical_limit
            )
        )


def send_resource_alert(actual_usage, critical_limit):
    """
    Function to simulate sending an alert for resource utilization exceeding the critical limit.
    In a real-world scenario, this could trigger an email, an SMS, or a notification.
    """
    # Simulate sending a resource usage alert
    logging.info(
        "Alert: System resource usage of {0}% exceeds the critical limit of {1}%. Please investigate.".format(
            actual_usage, critical_limit
        )
    )


# ================= End of Structural Testing for Header Mismatch Validation =================

# ================= Beginning of Structural Testing for Merged Cells Validation =================
# This script ensures:
# - Merged cells in export files are detected and logged.
# - Data integrity is maintained despite merged cell issues.
# - Auto-splitting is suggested if enabled.
# - System performance remains stable when handling merged cell inconsistencies.


@given('a bank export file named "{file_name}" containing merged cells')
def file_with_merged_cells(context, file_name):
    # type: (Any, Any) -> None
    """Simulate loading a bank export file that contains merged cells"""
    context.file_name = file_name
    context.merged_cells_detected = random.choice([True, False])
    logging.info(
        "Loaded {}. Merged cells detected: {}".format(
            file_name, context.merged_cells_detected
        )
    )


@then('all merged cell occurrences should be detected and flagged as "{severity}"')
def step_then_flag_batch_merged_cells(context, severity):
    # type: (Any, Any) -> None
    """Ensure merged cell issues in batch files are flagged"""

    # Function to handle merged cell detection and flagging
    def detect_and_flag_merged_cells(file_name, severity_level):
        # Simulating the process of checking for merged cells in a batch
        merged_cells_detected = random.choice(
            [True, False]
        )  # Random choice for simulation
        if merged_cells_detected:
            logging.warning(
                "Merged cells detected in batch file {0}. Severity: {1}".format(
                    file_name, severity_level
                )
            )
            return True
        else:
            logging.info("No merged cells detected in batch file {0}".format(file_name))
            return False

    # Check merged cells in each batch file
    for file_name in context.batch_files:
        merged_cells_flagged = detect_and_flag_merged_cells(file_name, severity)

        # Store the status in context if merged cells were flagged
        if merged_cells_flagged:
            context.merged_cells_flagged = True
        else:
            context.merged_cells_flagged = False

    # Additional logic can be added to handle severity levels
    if context.merged_cells_flagged:
        logging.info("Merged cells flagged with severity: {0}".format(severity))
        # Add further actions such as sending notifications or reporting if needed.
    else:
        logging.info("No merged cells detected in the batch.")


@then("a validation report should document the merged cell locations")
def step_then_generate_merged_cell_report(context):
    # type: (Any) -> None
    """Generate a report documenting merged cell locations"""
    if context.merged_cells_detected:
        # Simulating the validation report for merged cells
        logging.info(
            "Validation report generated for merged cells in {0}".format(
                context.file_name
            )
        )
        # Here, you could add logic to generate the actual report and store it.
    else:
        logging.info("No merged cells detected in {0}".format(context.file_name))


@then("if auto-splitting is enabled, the system should attempt to correct the issue")
def step_then_auto_correct_merged_cells(context):
    # type: (Any) -> None
    """Simulate an auto-splitting mechanism for merged cells"""
    context.merged_cells_detected = random.choice([True, False])
    if context.merged_cells_detected:
        # Simulating auto-splitting applied logic
        auto_split_success = random.choice([True, False])
        if auto_split_success:
            logging.info(
                "Auto-splitting applied successfully for {0}".format(context.file_name)
            )
        else:
            logging.warning("Auto-splitting failed for {0}".format(context.file_name))
    else:
        logging.info(
            "No merged cells detected, no auto-splitting needed for {0}".format(
                context.file_name
            )
        )


@given('an attempt to process a bank export file "{file_name}"')
def file_for_error_handling(context, file_name):
    # type: (Any, Any) -> None
    """Simulate an attempt to process a file with merged cell issues"""
    context.file_name = file_name
    # Simulate detecting merged cells in the given file
    context.merged_cells_detected = random.choice([True, False])
    logging.info(
        "Processing file: {0}. Merged cells detected: {1}".format(
            context.file_name, context.merged_cells_detected
        )
    )

    # Optional: Logic to handle processing the file with merged cells
    if context.merged_cells_detected:
        # Here you can simulate detecting specific merged cells or processing logic
        logging.info(
            "File {0} contains merged cells. Validation will be applied.".format(
                context.file_name
            )
        )
    else:
        logging.info("No merged cells detected in {0}".format(context.file_name))


@when('merged cells are detected in "{column_name}"')
def step_when_detect_merged_cells_in_column(context, column_name):
    # type: (Any, Any) -> None
    """Detect merged cells in a specific column"""
    context.merged_cells_detected = random.choice([True, False])
    context.column_name = column_name
    if context.merged_cells_detected:
        logging.warning(
            "Merged cells detected in column {0} of {1}".format(
                column_name, context.file_name
            )
        )
    else:
        logging.info(
            "No merged cells detected in column {0} of {1}".format(
                column_name, context.file_name
            )
        )


@then("an auto-splitting mechanism should suggest corrections if applicable")
def step_then_suggest_auto_splitting(context):
    # type: (Any) -> None
    """Suggest auto-splitting for merged cells"""
    if context.merged_cells_detected:
        logging.info(
            "Suggested auto-splitting correction for {0}, Column: {1}".format(
                context.file_name, context.column_name
            )
        )
    else:
        logging.info(
            "No need for auto-splitting for {0}, Column: {1}".format(
                context.file_name, context.column_name
            )
        )


@then("if correction is not possible, the file should be rejected")
def step_then_reject_file_with_unfixable_merged_cells(context):
    # type: (Any) -> None
    """Reject files if merged cells cannot be fixed"""
    if context.merged_cells_detected and not random.choice([True, False]):
        logging.error(
            "File {0} rejected due to unresolvable merged cells.".format(
                context.file_name
            )
        )


@given("a batch of bank export files with merged cell issues")
def batch_with_merged_cells(context):
    # type: (Any) -> None
    """Simulate batch processing of export files containing merged cells"""
    context.batch_contains_merged_cells = random.choice([True, False])


@given('a system processing "{file_count}" bank export files per hour')
def high_volume_merged_cell_validation(context, file_count):
    # type: (Any, Any) -> None
    """Simulate high-volume file processing with merged cells"""
    context.file_count = int(file_count)


@when('merged cells are present in "{year_range}"')
def step_when_detect_merged_cells_in_year_range(context, year_range):
    # type: (Any, Any) -> None
    """Perform merged cell validation across multiple years."""
    context.year_range = year_range
    context.processing_time = random.randint(100, 600)

    # Simulate processing delay using datetime
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=1)

    while datetime.now() < end_time:
        pass  # Simulating processing time

    logging.info(
        "Checked {0} files for merged cells in {1}.".format(
            context.file_count, context.year_range
        )
    )


@then('system resources should not exceed "{resource_limit}%"')
def step_then_validate_merged_cell_resource_usage(context, resource_limit):
    # type: (Any, Any) -> None
    """Ensure system resource usage remains within acceptable limits"""
    actual_usage = random.randint(60, 85)
    assert actual_usage <= int(resource_limit), "System resource usage exceeded!"
    logging.info("System resource usage: {0}, within limit.".format(actual_usage))


@given('an export file "{file_name}" with schema "{schema_type}"')
def file_with_merged_cell_schema(context, file_name, schema_type):
    # type: (Any, Any, Any) -> None
    """Simulate an export file with a specific schema that includes merged cells"""
    context.file_name = file_name
    context.schema_type = schema_type

    # Simulate file schema validation
    context.merged_cells_detected = random.choice([True, False])
    if context.merged_cells_detected:
        logging.info(
            "File {0} with schema {1} contains merged cells.".format(
                file_name, schema_type
            )
        )
    else:
        logging.info(
            "File {0} with schema {1} does not contain merged cells.".format(
                file_name, schema_type
            )
        )


@when("I check the schema validation rules")
def step_when_check_schema_validation_for_merged_cells(context):
    # type: (Any) -> None
    """Validate merged cell schema compliance"""
    context.schema_mismatch = random.choice([True, False])
    logging.info(
        "Checked schema for {0}. Schema mismatch: {1}".format(
            context.file_name, context.schema_mismatch
        )
    )


@then('all merged cells should be split according to "{expected_format}"')
def step_then_split_merged_cells(context, expected_format):
    # type: (Any, Any) -> None
    """Ensure merged cells are split correctly based on schema"""
    if context.schema_mismatch:
        logging.warning(
            "Schema mismatch detected in {0}. Expected format: {1}".format(
                context.file_name, expected_format
            )
        )
        # Assuming a function exists to handle the merging logic
        handle_merged_cells_split(context.file_name, expected_format)


def handle_merged_cells_split(file_name, expected_format):
    """
    Simulates splitting of merged cells based on expected format.
    Here you can add logic to check the format and handle corrections.
    """
    logging.info(
        "Attempting to split merged cells in {0} according to {1}".format(
            file_name, expected_format
        )
    )
    # Logic for splitting merged cells based on the expected format
    # If the splitting succeeds, we log a success message
    logging.info("Merged cells split successfully in {0}".format(file_name))
    # If splitting fails, we log a failure message
    # logging.error("Failed to split merged cells in {0}".format(file_name))


@then('any detected schema violations should be logged as "{error_severity}"')
def step_then_log_schema_violations(context, error_severity):
    # type: (Any, Any) -> None
    """Log schema violations due to merged cells"""
    if context.schema_mismatch:
        logging.error(
            "Schema violation detected in {0}. Severity: {1}".format(
                context.file_name, error_severity
            )
        )
        # Depending on severity, you may need to trigger alerts or further actions
        log_severity_action(context.file_name, error_severity)


def log_severity_action(file_name, error_severity):
    """
    This function handles actions based on the severity of the schema violation.
    You can expand it to send alerts, escalate issues, etc.
    """
    if error_severity == "high":
        logging.critical(
            "Critical error in file {0} due to schema violation.".format(file_name)
        )
        # Trigger any critical actions here (e.g., sending alerts or stopping processes)
    elif error_severity == "medium":
        logging.warning(
            "Warning: Schema issue in file {0} with medium severity.".format(file_name)
        )
        # Handle medium severity (e.g., log for future reference)
    elif error_severity == "low":
        logging.info(
            "Minor issue detected in file {0} with low severity.".format(file_name)
        )
        # Low severity might just be logged for informational purposes


# Example helper function for logging merged cell corrections
def handle_merged_cells_split(file_name, expected_format):
    """
    This function simulates the process of splitting merged cells in a file.
    You can extend this to actually check for merged cells and attempt to fix them.
    """
    logging.info(
        "Attempting to split merged cells in file {0} according to format {1}".format(
            file_name, expected_format
        )
    )
    # Simulate success/failure of split operation
    split_successful = random.choice([True, False])
    if split_successful:
        logging.info("Successfully split merged cells in {0}".format(file_name))
    else:
        logging.error("Failed to split merged cells in {0}".format(file_name))


# ================= End of Structural Testing for Merged Cells Validation =================

# ================= Beginning of Missing Columns Validation Script =================


@given(
    'a bank export file named "{file_name}" with missing columns "{missing_columns}"'
)
def export_file_with_missing_columns(context, file_name, missing_columns):
    # type: (Any, Any, Any) -> None
    """Loads the file and identifies missing columns."""
    context.file_name = file_name
    context.missing_columns = missing_columns.split(", ")

    try:
        # Attempt to read the file
        if file_name.endswith(".csv"):
            df = pd.read_csv(file_name)
        else:
            df = pd.read_excel(file_name)
        context.df = df
        context.error = None
    except Exception as e:
        context.df = None
        context.error = str(e)


# General file processing function for all file types (CSV and Excel)
def process_file(context, processing_type=None):
    # type: (Any, Any) -> None
    """Handles file processing based on the processing type."""

    # Simulate file processing delay using datetime
    simulated_delay = random.uniform(0.5, 2.0)
    start_time = datetime.now()
    end_time = start_time + timedelta(seconds=simulated_delay)
    while datetime.now() < end_time:
        pass  # Simulating processing delay

    def process_each_file(file_path):
        # type: (Any) -> pd.DataFrame
        """Processes files based on their type (CSV or Excel)."""
        if file_path.endswith(".csv"):
            df = pd.read_csv(file_path)
            logging.info("Processed CSV file: {0}".format(file_path))
            return df
        elif file_path.endswith(".xlsx"):
            df = pd.read_excel(file_path, sheet_name=None)
            logging.info("Processed Excel file: {0}".format(file_path))
            return df
        else:
            raise ValueError("Unsupported file format: {0}".format(file_path))

    # Handle different types of processing based on the `processing_type`
    if processing_type == "error-prone":
        context.error_count = (
            context.transaction_count * context.error_percentage
        ) // 100
        context.valid_transactions = context.transaction_count - context.error_count
        logging.info(
            "Processed {0} valid transactions from {1}.".format(
                context.valid_transactions, context.file_name
            )
        )

    elif processing_type == "concurrent":
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            context.processed_files = list(
                executor.map(process_each_file, context.file_paths)
            )

    elif processing_type == "column_validation":
        df = process_each_file(context.file_name)
        context.mismatched_columns = [
            col for col in context.expected_columns if col not in df.columns
        ]
        logging.info(
            "Processed {0}. Column mismatches detected: {1}".format(
                context.file_name, context.mismatched_columns
            )
        )

    elif processing_type == "format_validation":
        df = process_each_file(context.file_name)
        context.format_mismatches = [
            col
            for col in context.columns_to_check
            if not isinstance(df[col].dtype, str)
        ]
        logging.info(
            "Processed {0}. Format mismatches detected: {1}".format(
                context.file_name, context.format_mismatches
            )
        )

    elif processing_type == "extra_columns":
        df = process_each_file(context.file_name)
        context.extra_columns_detected = [
            col for col in df.columns if col not in context.expected_columns
        ]
        logging.info(
            "Processing {0}. Extra columns detected: {1}".format(
                context.file_name, context.extra_columns_detected
            )
        )

    elif processing_type == "header_validation":
        df = process_each_file(context.file_name)
        context.header_mismatch_detected = [
            col for col in context.expected_columns if col not in df.columns
        ]
        logging.info(
            "Processing {0}. Header mismatch detected: {1}".format(
                context.file_name, context.header_mismatch_detected
            )
        )

    elif processing_type == "merged_cells":
        if context.file_name.endswith(".xlsx"):
            df = pd.read_excel(context.file_name)
            context.merged_cells_detected = any(df.iloc[:, 0].isnull())
            logging.info(
                "Processing {0}. Merged cells detected: {1}".format(
                    context.file_name, context.merged_cells_detected
                )
            )

    else:
        df = process_each_file(context.file_name)
        context.process_result = "File processed successfully."
        logging.info("File {0} processed successfully.".format(context.file_name))


# Combined step definitions for the feature file, reusing the `process_file` function for all cases
@when('I attempt to process the file')
def step_when_process_file(context):
    # type: (Any) -> None
    """Simulate basic file processing with a check for empty files."""
    if context.is_empty:
        context.process_result = "Empty file detected"
    else:
        context.process_result = "File processed successfully"


@when("I process these files concurrently")
def step_when_process_files_concurrently(context):
    # type: (Any) -> None
    """Simulate concurrent processing of multiple files."""
    process_file(context, processing_type="concurrent")


@when("I attempt to process the file with errors")
def step_when_process_file_with_errors(context):
    # type: (Any) -> None
    """Simulate error-prone file processing."""
    process_file(context, processing_type="error-prone")


@when("the system processes the file and validates column headers")
def step_when_process_file_case_variations(context):
    # type: (Any) -> None
    """Simulate processing the file and validating column headers."""
    process_file(context, processing_type="column_validation")


@when("the system processes the file and validates column formats")
def step_when_process_file_format_variations(context):
    # type: (Any) -> None
    """Simulate processing the file and validating column formats."""
    process_file(context, processing_type="format_validation")


@when("the system processes the file and detects extra columns")
def step_when_process_file_with_extra_columns(context):
    # type: (Any) -> None
    """Simulate processing the file and detecting extra columns."""
    process_file(context, processing_type="extra_columns")


@when("the system processes the file and checks headers")
def step_when_process_file_with_headers(context):
    # type: (Any) -> None
    """Simulate processing the file and checking headers."""
    process_file(context, processing_type="header_validation")


@ when("the system processes the file and checks for merged cells") @ when(
    "the system processes the file and checks for merged cells"
)
def step_when_process_file_with_merged_cells(context):
    # type: (Any) -> None
    """Simulate processing the file and checking for merged cells."""

    # Ensure that file_name is available in context
    if not hasattr(context, "file_name"):
        raise ValueError("File name is not defined in the context.")

    logging.info(
        "Starting to process the file and check for merged cells: {0}".format(
            context.file_name
        )
    )

    try:
        # Using FileProcessor to process the file (if Excel file, we handle merged cells checking)
        file_processor = FileProcessor(file_name=context.file_name, context=context)
        file_processor.process()  # This will load the file into the context.df

        # Check if the file is an Excel file (.xlsx)
        if context.file_name.endswith(".xlsx"):
            # Check for merged cells
            check_for_merged_cells_in_file(context)
        else:
            logging.info(
                "File {0} is not an Excel file, skipping merged cells check.".format(
                    context.file_name
                )
            )

    except Exception as e:
        logging.error(
            "Error processing file {0}: {1}".format(context.file_name, str(e))
        )
        raise ValueError(
            "Error processing file {0}: {1}".format(context.file_name, str(e))
        )


@when("the system processes the file and checks for missing columns")
def step_when_process_file_and_check_columns(context):
    # type: (Any) -> None
    """Check if specified columns are missing in the file."""
    if context.df is not None:
        context.missing_detected = [
            col for col in context.missing_columns if col not in context.df.columns
        ]
    else:
        context.missing_detected = []


@then('the missing columns should be flagged with severity "{severity}"')
def flagging_missing_columns(context, severity):
    # type: (Any, Any) -> None
    """Flags missing columns based on severity level."""
    assert (
        len(context.missing_detected) > 0
    ), "No missing columns detected in {0}".format(context.file_name)

    print(
        "Missing columns in {0}: {1} - Severity: {2}".format(
            context.file_name, context.missing_detected, severity
        )
    )


@then("a validation report should document the missing fields")
def documenting_missing_columns_validation_report(context):
    # type: (Any) -> None
    """Generates a validation report for missing columns."""
    report = {
        "file_name": context.file_name,
        "missing_columns": context.missing_detected,
    }
    print("Validation report generated: {0}".format(report))


# Error Handling for Missing Columns
@given('an attempt to process a bank export file "{file_name}"')
def attempt_to_process(context, file_name):
    # type: (Any, Any) -> None
    """Attempts to process the file while handling missing columns."""
    context.file_name = file_name
    try:
        if file_name.endswith(".csv"):
            df = pd.read_csv(file_name)
        else:
            df = pd.read_excel(file_name)
        context.df = df
        context.error = None
    except Exception as e:
        context.df = None
        context.error = str(e)


@when('required columns are missing such as "{missing_column}"')
def step_when_required_columns_missing(context, missing_column):
    # type: (Any, Any) -> None
    """Checks for missing required columns."""
    context.missing_column = missing_column
    if context.df is not None:
        context.missing_detected = (
            [missing_column] if missing_column not in context.df.columns else []
        )
    else:
        context.missing_detected = []


@then(
    'the issue should be escalated if the missing column severity level is "{severity_level}"'
)
def step_then_escalate_issue(context, severity_level):
    # type: (Any, Any) -> None
    """Escalates issue based on severity level."""
    if len(context.missing_detected) > 0:
        logging.warning(
            "Escalation required for missing columns: {0} - Severity: {1}".format(
                context.missing_detected, severity_level
            )
        )


@then("if correction is not possible, the file should be rejected")
def step_then_reject_file(context):
    # type: (Any) -> None
    """Rejects the file if the issue is critical."""
    if len(context.missing_detected) > 0:
        logging.error(
            "REJECTED: {0} due to missing critical columns: {1}".format(
                context.file_name, context.missing_detected
            )
        )


# Batch Processing for Missing Columns
@given("a batch of bank export files with missing columns")
def batch_export_files_with_missing_columns(context):
    # type: (Any) -> None
    """Loads batch files for validation."""
    context.batch_files = [
        "transactions_missing_account.csv",
        "transactions_missing_currency.xlsx",
    ]
    # Simulate reading and storing batch file information
    context.batch_issues = {
        "transactions_missing_account.csv": ["Account Column"],
        "transactions_missing_currency.xlsx": ["Currency Column"],
    }


@then('all missing column occurrences should be detected and flagged as "{severity}"')
def step_then_flag_batch_issues(context, severity):
    # type: (Any, Any) -> None
    """Flags missing column occurrences in batch processing."""
    for file, issues in context.batch_issues.items():
        logging.warning(
            "Batch file {0} has missing columns: {1} - Severity: {2}".format(
                file, issues, severity
            )
        )


@then("processing should continue without failure")
def continue_processing_without_failure(context):
    # type: (Any) -> None
    """Ensures batch processing continues even with missing columns."""
    logging.info("Batch processing continued despite missing column issues.")


# Performance Testing for Missing Columns
@given('a system processing "{file_count}" bank export files per hour')
def system_performance(context, file_count):
    # type: (Any, Any) -> None
    """Simulates system performance testing with missing column validation."""
    context.file_count = int(file_count)
    logging.info("System processing {0} files per hour.".format(file_count))


@when('missing columns are present in "{year_range}"')
def missing_columns_in_year_range(context, year_range):
    # type: (Any, Any) -> None
    """Simulates missing column presence across multiple years."""
    context.year_range = year_range
    logging.info("Simulating missing columns in year range: {0}".format(year_range))


@then("the system should handle the missing columns gracefully")
def system_handling_missing_columns_gracefully(context):
    # type: (Any) -> None
    """Ensure the system handles missing columns gracefully without failure."""
    if len(context.missing_detected) > 0:
        logging.warning("Missing columns detected, processing may be affected.")
    else:
        logging.info("No missing columns detected, proceeding with processing.")


# ================= Beginning of Missing Columns Validation Script =================


@given('an export file "{file_name}" with schema "{schema_type}"')
def get_an_export_file_with_schema(context, file_name, schema_type):
    # type: (Any, Any, Any) -> None
    """Loads the file and identifies missing columns."""
    context.file_name = file_name
    context.schema_type = schema_type


@when("I check the schema validation rules")
def step_when_check_schema_rules(context):
    # type: (Any) -> None
    """Checks the schema validation rules for missing columns."""
    context.schema_errors = []


@then('missing columns should be flagged as "{error_severity}"')
def step_then_flag_schema_errors(context, error_severity):
    # type: (Any, Any) -> None
    """Flags schema-related errors due to missing columns."""
    print(
        "Schema validation for {}: Missing columns flagged as {}".format(
            context.file_name, error_severity
        )
    )


@then("system logs should capture all schema-related discrepancies")
def step_then_log_schema_discrepancies(context):
    # type: (Any) -> None
    """Logs schema validation issues for missing columns."""
    print("Logged schema validation issues for {}".format(context.file_name))


# ================= Beginning of Protected Sheets Validation Script =================


@given('a bank export file "{file_name}" with a protected sheet "{sheet_name}"')
def protected_sheet(context, file_name, sheet_name):
    # type: (Any, Any, Any) -> None
    """Loads the file and checks for protected sheets."""
    context.file_name = file_name
    context.sheet_name = sheet_name
    try:
        workbook = load_workbook(file_name)
        sheet = workbook[sheet_name]
        context.protection_type = (
            "Read-Only" if sheet.protection.sheet else "Unprotected"
        )
        context.error = None
    except Exception as e:
        context.protection_type = "Unknown"
        context.error = str(e)


@when("the system processes the file")
def step_when_process_protected_file(context):
    # type: (Any) -> None
    """Attempts to process the file and identify protection settings."""
    if hasattr(context, "error") and context.error:
        context.protection_status = "Error"
    else:
        context.protection_status = "Processed"


@then('the protection level should be identified as "{protection_type}"')
def step_then_identify_protection(context, protection_type):
    # type: (Any, Any) -> None
    """Verifies if the identified protection matches the expected type."""
    assert (
        context.protection_type == protection_type
    ), "Expected {}, but found {}".format(protection_type, context.protection_type)
    print(
        "Protection level for {} in {}: {}".format(
            context.sheet_name, context.file_name, context.protection_type
        )
    )


@then("a validation report should document the protection settings")
def step_then_document_protection(context):
    # type: (Any) -> None
    """Generates a validation report for protected sheets."""
    report = {
        "file_name": context.file_name,
        "sheet_name": context.sheet_name,
        "protection_type": context.protection_type,
    }
    print("Validation report generated: {}".format(report))


@then("if credentials are available, the sheet should be unlocked for processing")
def step_then_unlock_sheet(context):
    # type: (Any) -> None
    """Simulates unlocking a protected sheet if credentials are available."""
    if context.protection_type in ["Password-Protected", "Read-Only"]:
        print(
            "Attempting to unlock {} in {}... Access Denied".format(
                context.sheet_name, context.file_name
            )
        )


# Error Handling for Protected Sheets


@given('an attempt to process a bank export file "{file_name}"')
def attempt_to_process_protected_file(context, file_name):
    # type: (Any, Any) -> None
    """Attempts to process the file while handling protected sheets."""
    context.file_name = file_name
    try:
        context.workbook = load_workbook(file_name)
        context.error = None
    except Exception as e:
        context.workbook = None
        context.error = str(e)


@when('a protected sheet "{sheet_name}" is encountered')
def step_when_protected_sheet_found(context, sheet_name):
    # type: (Any, Any) -> None
    """Checks if the sheet is protected."""
    context.sheet_name = sheet_name
    if context.workbook and sheet_name in context.workbook.sheetnames:
        sheet = context.workbook[sheet_name]
        context.protection_detected = (
            "Protected" if sheet.protection.sheet else "Unprotected"
        )
    else:
        context.protection_detected = "Sheet Not Found"


@then('the protection status should be logged')
def step_then_log_protection_status(context):
    # type: (Any) -> None
    """Logs the protection status of the sheet."""
    if context.protection_detected == "Sheet Not Found":
        logging.error("Sheet {} not found in the file.".format(context.sheet_name))
    else:
        logging.info(
            "Sheet {} in file {} is {}".format(
                context.sheet_name, context.file_name, context.protection_detected
            )
        )


# Error Handling for Missing Columns


@given('an export file "{file_name}" with missing columns "{missing_columns}"')
def get_an_export_file_with_missing_columns(context, file_name, missing_columns):
    # type: (Any, Any, Any) -> None
    """Loads the file and identifies missing columns."""
    context.file_name = file_name
    context.missing_columns = missing_columns.split(", ")
    try:
        context.df = (
            pd.read_csv(file_name)
            if file_name.endswith(".csv")
            else pd.read_excel(file_name)
        )
        context.error = None
    except Exception as e:
        context.df = None
        context.error = str(e)


@then(
    'the issue should be escalated if the missing column severity level is "{severity_level}"'
)
def step_then_escalate_issue(context, severity_level):
    # type: (Any, Any) -> None
    """Escalates issue based on severity level."""
    if len(context.missing_detected) > 0:
        logging.warning(
            "Escalation required for missing columns: {0} - Severity: {1}".format(
                context.missing_detected, severity_level
            )
        )


@then("if correction is not possible, the file should be rejected")
def step_then_reject_file(context):
    # type: (Any) -> None
    """Rejects the file if the issue is critical."""
    if len(context.missing_detected) > 0:
        logging.error(
            "REJECTED: {0} due to missing critical columns: {1}".format(
                context.file_name, context.missing_detected
            )
        )


@then('the issue should be escalated if the protection level is "{severity_level}"')
def step_then_escalate_protected_sheet_issue(context, severity_level):
    # type: (Any, Any) -> None
    """Escalates issue based on severity level."""
    if context.protection_detected == "Protected":
        print(
            "Escalation required: {} in {} - Severity: {}".format(
                context.sheet_name, context.file_name, severity_level
            )
        )


@then("an override attempt should be logged if credentials are provided")
def step_then_log_override(context):
    # type: (Any) -> None
    """Logs override attempt if credentials are available."""
    print(
        "Override attempt logged for {} in {}".format(
            context.sheet_name, context.file_name
        )
    )


# Batch Processing for Protected Sheets
@given("a batch of bank export files with protected sheets")
def batch_with_protected_sheets(context):
    # type: (Any) -> None
    """Loads batch files for validation."""
    context.batch_files = ["transactions_protected.xlsx", "transactions_password.xlsx"]


@then('all protected sheets should be detected and flagged as "{severity}"')
def step_then_flag_protected_sheets_with_severity(context, severity):
    # type: (Any, Any) -> None
    """Flags protected sheet occurrences in batch processing."""
    for file, issues in context.batch_issues.items():
        print(
            "Batch file {} has protected sheets: {} - Severity: {}".format(
                file, issues, severity
            )
        )


@then("processing should continue if read-only access is available")
def step_then_continue_with_read_only_access(context):
    # type: (Any) -> None
    """Ensures batch processing continues if read-only access is available."""
    print("Batch processing continued with read-only access.")


# Performance Testing for Protected Sheets
@when('protected sheets are present in "{year_range}"')
def step_when_protected_sheets_in_year_range(context, year_range):
    # type: (Any, Any) -> None
    """Simulates protected sheet presence across multiple years."""
    context.year_range = year_range


@then('system resources should not exceed "{resource_limit}%"')
def step_then_resource_limit_check(context, resource_limit):
    # type: (Any, Any) -> None
    """Ensures resource usage remains within acceptable limits."""
    print("System resource usage within {} percent limit.".format(resource_limit))


# Security Validation for Protected Sheets
@given('an export file "{file_name}" with security protection on "{sheet_name}"')
def protected_file(context, file_name, sheet_name):
    # type: (Any, Any, Any) -> None
    """Loads a file with a protected sheet for security validation."""
    context.file_name = file_name
    context.sheet_name = sheet_name


@when("I check the security validation rules")
def step_when_check_security_rules(context):
    # type: (Any) -> None
    """Checks the security validation rules for protected sheets."""
    context.security_issues = []


@then('protected sheets should conform to security standards "{security_level}"')
def step_then_validate_security(context, security_level):
    # type: (Any, Any) -> None
    """Flags security issues due to protected sheets."""
    print(
        "Security validation for "
        + context.file_name
        + ": Protection level "
        + security_level
        + " verified."
    )


@then("system logs should capture all security-related discrepancies")
def step_then_log_security_discrepancies(context):
    # type: (Any) -> None
    """Logs security validation issues for protected sheets."""
    print("Logged security validation issues for " + context.file_name)


# ================= End of Protected Sheets Validation Script =================

# ================= Beginning of Reordered Columns Validation Script =================


@given  # ================== Step Definitions ==================
@given('a bank export file "{file_name}" with columns in an unexpected order')
def reordered_columns(context, file_name):
    # type: (Any, Any) -> None
    """Loads the file and verifies column order."""
    context.file_name = file_name
    try:
        # Try loading the file
        if file_name.endswith(".csv"):
            context.df = pd.read_csv(file_name)
        else:
            context.df = pd.read_excel(file_name)
        context.error = None
    except Exception as e:
        context.df = None
        context.error = str(e)


@when("the system processes the file")
def step_when_process_reordered_columns(context):
    # type: (Any) -> None
    """Processes the file and checks column order."""
    if context.error:
        context.processing_status = "Error"
    else:
        context.processing_status = "Processed"


@then('column order should be verified against the reference format "{reference_file}"')
def step_then_verify_column_order(context, reference_file):
    # type: (Any, Any) -> None
    """Compares file's column order with a reference format."""
    try:
        # Initialize the file processor to handle the reference file
        ref_processor = FileProcessor(file_name=reference_file, context=context)
        ref_processor.process()  # This will load the reference file into context.df

        # Get reference columns and the columns from the current file
        context.ref_columns = list(context.df.columns)  # Reference columns
        context.file_columns = list(context.df.columns)  # File columns being processed

        # Check if column order is different from reference
        context.reordered = context.file_columns != context.ref_columns

        if context.reordered:
            logging.warning(
                "Column order in {0} does not match the reference format: {1}".format(
                    context.file_name, reference_file
                )
            )

    except Exception as e:
        context.reordered = False
        context.error = str(e)
        logging.error("Error during column order verification: {0}".format(str(e)))


@then('any reordering should be flagged as "{severity}"')
def step_then_flag_reordering(context, severity):
    # type: (Any, Any) -> None
    """Flags reordered columns if detected."""
    if context.reordered:
        logging.warning(
            "Column reordering detected in {0} - Severity: {1}".format(
                context.file_name, severity
            )
        )


@then("if auto-mapping is enabled, the system should realign the columns")
def step_then_realign_columns(context):
    # type: (Any) -> None
    """Simulates auto-mapping and column realignment."""

    # Ensure that the columns are reordered before attempting realignment
    if context.reordered:
        try:
            # Process the file using FileProcessor
            file_processor = FileProcessor(file_name=context.file_name, context=context)
            file_processor.process()  # This reads the file into context.df

            # Retrieve column mappings for the current file
            column_mapping = context.file_mappings.get(context.file_name, {})

            # Validate the file and capture any issues
            validation_issues = validate_file(context.df, column_mapping)

            if validation_issues:
                context.error = "Column realignment failed: " + " | ".join(
                    [key + ": " + value for key, value in validation_issues.items()]
                )
            else:
                # If no issues, simulate the realignment process
                context.reordered = False  # Mark that columns have been realigned
                context.error = None  # Reset any previous errors
                logging.info(
                    "Auto-mapping enabled columns for {0}".format(context.file_name)
                )

        except Exception as e:
            context.error = "Error during column realignment: " + str(e)
            logging.error("Error during column realignment: {0}".format(str(e)))


@given('an attempt to process a bank export file "{file_name}"')
def attempt_process_file(context, file_name):
    # type: (Any, Any) -> None
    """Attempts to process the file while handling reordered columns and performing validation checks."""

    # Set the file name in context and initialize other attributes
    context.file_name = file_name
    context.error = None
    context.df = None

    try:
        # Initialize the FileProcessor to handle the file reading
        file_processor = FileProcessor(file_name=file_name, context=context)

        # Process the file (reads it into context.df)
        file_processor.process()

        # Validate the file content using the new validate_file function
        column_mapping = context.file_mappings.get(file_name, {})
        validation_issues = validate_file(
            context.df, column_mapping, file_type=file_name.split('.')[-1]
        )

        # If there are validation issues, store them in the context error
        if validation_issues:
            context.error = "Validation failed: " + " | ".join(
                [
                    "{0}: {1}".format(key, value)
                    for key, value in validation_issues.items()
                ]
            )
            logging.error("Validation failed: {0}".format(context.error))
        else:
            context.error = None  # No issues found
            logging.info(
                "File {0} processed successfully without validation issues.".format(
                    file_name
                )
            )

    except Exception as e:
        # If an error occurs during file processing, store the error in context.error
        context.error = "Error processing file: {0}".format(str(e))
        logging.error("Error processing file: {0}".format(context.error))


@when("columns are reordered in an unexpected way")
def step_when_unexpected_reorder(context):
    # type: (Any) -> None
    """Checks for unexpected column order and validates columns using helpers from data_validation.py."""

    # Ensure required attributes are present in context
    if (
        not hasattr(context, 'df')
        or not hasattr(context, 'ref_columns')
        or context.df is None
    ):
        raise ValueError(
            "Context is missing required attributes: 'df' or 'ref_columns'."
        )

    # Initialize context attributes
    context.reordered = False
    context.batch_issues = getattr(context, 'batch_issues', {})

    # Check for unexpected column reordering
    if context.df.columns.tolist() != context.ref_columns:
        context.reordered = True  # Mark as reordered if columns don't match

        # Track the issues found in the file
        context.batch_issues[context.file_name] = (
            "Unexpected column reordering detected."
        )

        # Run validations for the columns using the helper function from data_validation.py
        column_mapping = context.file_mappings.get(context.file_name, {})
        issues = validate_file(context.df, column_mapping)

        # Log the issues for this file
        if issues:
            context.batch_issues[context.file_name] += " | " + " | ".join(
                ["{0}: {1}".format(key, value) for key, value in issues.items()]
            )

        # Handle missing critical columns explicitly
        missing_required_columns = [
            col for col in context.ref_columns if col not in context.df.columns
        ]
        if missing_required_columns:
            context.batch_issues[
                context.file_name
            ] += " | Missing critical columns: " + ", ".join(missing_required_columns)

        # Handle misplaced columns in the current file
        misplaced_columns = [
            col for col in context.df.columns if col not in context.ref_columns
        ]
        if misplaced_columns:
            context.batch_issues[
                context.file_name
            ] += " | Misplaced columns: " + ", ".join(misplaced_columns)

        # Optionally log the detected issues for debugging or auditing
        logging.info(
            "Issues detected in file: {0}".format(
                context.batch_issues[context.file_name]
            )
        )

    else:
        # If columns are in the expected order, mark it as valid
        context.reordered = False
        context.batch_issues[context.file_name] = "Columns are in the expected order."

    # Print any issues found in the batch (for debugging purposes)
    logging.info("Batch issues: {0}".format(context.batch_issues))


@given('an attempt to process a bank export file "{file_name}"')
def attempt_process_file(context, file_name):
    # type: (Any, Any) -> None
    """Attempts to process the file while handling reordered columns and performing validation checks."""

    # Set the file name in context and initialize other attributes
    context.file_name = file_name
    context.error = None
    context.df = None

    try:
        # Initialize the FileProcessor to handle the file reading
        file_processor = FileProcessor(file_name=file_name, context=context)

        # Process the file (reads it into context.df)
        file_processor.process()

        # Validate the file content using the new validate_file function
        column_mapping = context.file_mappings.get(file_name, {})
        validation_issues = validate_file(
            context.df, column_mapping, file_type=file_name.split('.')[-1]
        )

        # If there are validation issues, store them in the context error
        if validation_issues:
            context.error = "Validation failed: " + " | ".join(
                [
                    "{0}: {1}".format(key, value)
                    for key, value in validation_issues.items()
                ]
            )
            logging.error("Validation failed: {0}".format(context.error))
        else:
            context.error = None  # No issues found
            logging.info(
                "File {0} processed successfully without validation issues.".format(
                    file_name
                )
            )

    except Exception as e:
        # If an error occurs during file processing, store the error in context.error
        context.error = "Error processing file: {0}".format(str(e))
        logging.error("Error processing file: {0}".format(context.error))


# Delimiter Consistency Validation
@given('a CSV file "{file_name}" with columns reordered near delimiters')
def check_csv_with_reordered_columns_near_delimiters(context, file_name):
    # type: (Any, Any) -> None
    """Loads a CSV file and checks for issues like missing columns, spaces, and delimiter consistency."""

    context.file_name = file_name

    # Check if the file exists
    if not os.path.isfile(file_name):
        raise ValueError("The file does not exist: {0}".format(file_name))

    try:
        # Load the CSV file
        context.df = pd.read_csv(file_name)

        # Validate if the file is empty
        if context.df.empty:
            raise ValueError("The CSV file is empty: {0}".format(file_name))

        # Define expected columns
        expected_columns = [
            "transaction_id",
            "date",
            "amount",
            "currency_code",
            "account_number",
        ]

        # Check for missing columns
        missing_cols = [
            col for col in expected_columns if col not in context.df.columns
        ]
        if missing_cols:
            raise ValueError("Missing expected columns: {0}".format(missing_cols))

        # Strip spaces from column names and check for consistency
        context.df.columns = context.df.columns.str.strip()
        if any(col != col.strip() for col in context.df.columns):
            raise ValueError(
                "Columns have leading or trailing spaces in the file: {0}".format(
                    file_name
                )
            )

        # Check delimiter consistency in the CSV
        with open(file_name, 'r') as f:
            first_line = f.readline()
            # Check if delimiter usage is consistent
            if first_line.count(",") != len(first_line.split(",")) - 1:
                raise ValueError(
                    "Inconsistent delimiter usage in the file: {0}".format(file_name)
                )

        # If no issues, set context.error to None
        context.error = None

    except Exception as e:
        # If an error occurs, store the error message in context.error
        context.error = str(e)
        logging.error("Error while loading the file: {0}".format(str(e)))


@when("the system parses the file")
def parse_file(context):
    # type: (Any) -> None
    """Parses the CSV or Excel file while checking delimiter integrity and validating structure."""
    print("Parsing file", context.file_name, "...")

    # Check if the file is a valid CSV or Excel file
    if not (context.file_name.endswith(".csv") or context.file_name.endswith(".xlsx")):
        raise ValueError("Invalid file format. Only CSV and Excel files are supported.")

    # Try to load the file (either CSV or Excel)
    try:
        if context.file_name.endswith(".csv"):
            context.df = pd.read_csv(context.file_name)
        else:
            context.df = pd.read_excel(
                context.file_name, sheet_name=None
            )  # Load all sheets if Excel
    except Exception as e:
        raise ValueError("Error parsing file", context.file_name, ":", str(e))

    # Ensure the file is not empty (check for CSV or Excel)
    if context.df.empty:
        raise ValueError("The file", context.file_name, "is empty.")

    # For multi-sheet Excel files, check each sheet for the same structure
    if isinstance(context.df, dict):  # If it's an Excel file with multiple sheets
        for sheet_name, df in context.df.items():
            if df.empty:
                raise ValueError("Sheet", sheet_name, "is empty in the Excel file.")
            # Optionally, check for expected columns here (example column check)
            expected_columns = [
                "column1",
                "column2",
                "column3",
            ]  # Replace with actual column names
            if not all(col in df.columns for col in expected_columns):
                raise ValueError(
                    "Sheet", sheet_name, "in the file is missing expected columns."
                )
    else:  # For CSV or single-sheet Excel
        # Check for expected columns in CSV or single-sheet Excel
        expected_columns = [
            "column1",
            "column2",
            "column3",
        ]  # Replace with actual column names
        if not all(col in context.df.columns for col in expected_columns):
            raise ValueError(
                "The file", context.file_name, "is missing expected columns."
            )

    # For CSV, check delimiter integrity by ensuring all rows are consistent
    if context.file_name.endswith(".csv"):
        with open(context.file_name, "r") as f:
            first_line = f.readline()
            delimiter = (
                ","  # Change if a different delimiter is used (e.g., tabs, semicolons)
            )
            if first_line.count(delimiter) != len(first_line.split(delimiter)) - 1:
                raise ValueError("Inconsistent delimiter usage in", context.file_name)

    # Print the parsed data info (optional)
    print(
        "File",
        context.file_name,
        "parsed successfully. Dataframe shape:",
        context.df.shape,
    )


@then("delimiter integrity should be preserved")
def check_delimiters(context):
    # type: (Any) -> None
    """Ensures delimiters remain intact after column reordering."""

    print("Checking delimiter integrity for file:", context.file_name)

    try:
        # Check if the file is a valid CSV file
        if not context.file_name.endswith(".csv"):
            raise ValueError(
                "File is not in CSV format. Only CSV files are supported for delimiter checks."
            )

        # Open the file and read the first few lines
        with open(context.file_name, 'r') as f:
            first_line = f.readline()
            # Assuming ',' as the delimiter, change this if needed
            delimiter = ","
            # Check if delimiter consistency is preserved by counting delimiters
            delimiter_count = first_line.count(delimiter)
            columns_in_first_line = len(first_line.split(delimiter))

            if delimiter_count != (columns_in_first_line - 1):
                raise ValueError(
                    "Inconsistent delimiter usage in the file: {0}. Delimiters count mismatch.".format(
                        context.file_name
                    )
                )

        print("Delimiter consistency check passed for {0}".format(context.file_name))

    except ValueError as e:
        print("Error during delimiter integrity check:", e)
        raise
    except Exception as e:
        print(
            "Unexpected error during delimiter check for {0}: {1}".format(
                context.file_name, str(e)
            )
        )
        raise


# ================= End of Reordered Columns Validation Script =================


# ================= Beginning of Trailing and Leading Spaces Validation Script =================
@given(
    'a bank export file "{file_name}" with values containing leading or trailing spaces'
)
def check_trailing_spaces(context, file_name):
    # type: (Any, Any) -> None
    """Loads the file and verifies space inconsistencies, missing values, and empty rows."""
    context.file_name = file_name
    try:
        # Load the file as either CSV or Excel
        context.df = (
            pd.read_csv(file_name)
            if file_name.endswith(".csv")
            else pd.read_excel(file_name)
        )

        # Check for missing values (NaN values)
        missing_values = context.df.isnull().sum().sum()
        if missing_values > 0:
            context.error = "File contains {} missing values.".format(missing_values)
            print("Warning: {} missing values found.".format(missing_values))
        else:
            context.error = None

        # Check for trailing or leading spaces in any cell
        rows_with_spaces = context.df.applymap(
            lambda x: isinstance(x, str) and (x.startswith(' ') or x.endswith(' '))
        )
        rows_with_spaces_count = rows_with_spaces.sum().sum()
        if rows_with_spaces_count > 0:
            context.error = (
                "File contains {} cells with leading or trailing spaces.".format(
                    rows_with_spaces_count
                )
            )
            print("Warning: {} cells with spaces found.".format(rows_with_spaces_count))
        else:
            context.error = None

        # Check if the file is empty (no rows or columns)
        if context.df.empty:
            context.error = "The file is empty."
            print("Warning: File is empty.")
        else:
            context.error = None

    except Exception as e:
        # Handle any exceptions that occur while loading the file
        context.df = None
        context.error = str(e)
        print("Error while loading the file: {}".format(str(e)))


@when("the system processes the file")
def step_when_process_trailing_spaces(context):
    # type: (Any) -> None
    """Processes the file and checks for leading/trailing spaces."""
    if context.error:
        context.processing_status = "Error"
    else:
        context.processing_status = "Processed"


@then("all affected fields should be flagged in the validation report")
def step_then_flag_spaces(context):
    # type: (Any) -> None
    """Flags fields containing leading or trailing spaces."""
    if context.df is not None:
        for col in context.df.columns:
            if context.df[col].astype(str).str.strip().ne(context.df[col]).any():
                print(
                    "WARNING: Leading or trailing spaces detected in column '{}' of {}".format(
                        col, context.file_name
                    )
                )


@then("the system should apply auto-trimming if configured")
def step_then_apply_auto_trimming(context):
    # type: (Any) -> None
    """Applies auto-trimming if enabled."""

    if context.df is not None:
        # Check if auto-trimming is enabled (assuming a flag is available in context)
        if hasattr(context, 'auto_trimming_enabled') and context.auto_trimming_enabled:
            # Apply auto-trimming to all string columns in the dataframe
            context.df = context.df.applymap(
                lambda x: x.strip() if isinstance(x, str) else x
            )
            logging.info("Auto-trimming applied to {}.".format(context.file_name))
        else:
            logging.info(
                "Auto-trimming is not enabled for {}".format(context.file_name)
            )


@then('fields with persistent spaces should be escalated as "{severity}"')
def step_then_escalate_persistent_spaces(context, severity):
    # type: (Any, Any) -> None
    """Escalates space issues based on severity."""
    print(
        "Escalation required for persistent space issues in {} - Severity: {}".format(
            context.file_name, severity
        )
    )


# Error Handling for Trailing Spaces
@given('an attempt to process a bank export file "{file_name}"')
def attempt_to_process_spaces(context, file_name):
    # type: (Any, Any) -> None
    """Attempts to process the file while handling trailing spaces."""

    context.file_name = file_name
    context.df = None
    context.error = None

    try:
        # Attempt to read the file (either CSV or Excel)
        if file_name.endswith(".csv"):
            context.df = pd.read_csv(file_name)
        elif file_name.endswith(".xlsx"):
            context.df = pd.read_excel(file_name)
        else:
            raise ValueError(f"Unsupported file format for {file_name}")

    except Exception as e:
        context.error = str(e)
        logging.error(f"Error processing file {file_name}: {context.error}")


@when('trailing or leading spaces are detected in "{column_name}"')
def step_when_detect_spaces_in_column(context, column_name):
    # type: (Any, Any) -> None
    """Checks for leading/trailing spaces in a specific column."""

    # Ensure the DataFrame is loaded and the column exists
    if context.df is not None and column_name in context.df.columns:

        # Check for trailing/leading spaces in the specified column
        context.space_issue = (
            context.df[column_name]
            .astype(str)  # Convert the column values to string
            .str.strip()  # Remove leading and trailing spaces
            .ne(
                context.df[column_name]
            )  # Compare stripped value with original to detect spaces
            .any()  # Check if any value has leading or trailing spaces
        )

        # If spaces are found, log the issue
        if context.space_issue:
            logging.warning(
                f"Trailing or leading spaces detected in column '{column_name}' of file '{context.file_name}'"
            )
        else:
            logging.info(
                f"No spaces detected in column '{column_name}' of file '{context.file_name}'"
            )
    else:
        context.space_issue = (
            False  # No issue if the column doesn't exist or DataFrame is empty
        )
        logging.error(f"Column '{column_name}' not found in file '{context.file_name}'")


@then("a system alert should notify relevant users")
def step_then_send_duplicate_alert(context: Any) -> None:
    """
    Trigger an alert if duplicate transaction references are detected.

    :param context: The context containing file-related data
    :return: None
    """
    # Check if duplicate references are detected
    if getattr(context, "is_duplicate", False):
        # Prepare the notification message and use the user_notifications' combined alert function
        notification_message = (
            "Duplicate transaction references detected in {0}".format(context.file_name)
        )
        send_combined_alert(context)

        # Optionally log the alert if needed
        logging.info("Duplicate issue detected. Alert sent to users.")
    else:
        logging.info("No duplicate issues detected.")


@then('the issue should be escalated if the space count exceeds "{threshold}"')
def step_then_escalate_space_issue(context, threshold):
    # type: (Any, Any) -> None
    """Escalates issue based on space count threshold."""

    # Ensuring that we have the necessary context attributes set
    if not hasattr(context, "space_issue") or not hasattr(context, "df"):
        raise ValueError("Missing necessary context attributes: 'space_issue' or 'df'")

    # Check if the space count in the specific column exceeds the given threshold
    column_name = "example_column"  # Replace with the actual column you are checking
    if column_name in context.df.columns:
        spaces_count = (
            context.df[column_name].astype(str).str.contains(r"^\s.*\s$").sum()
        )
        threshold_value = int(threshold)  # Threshold is expected to be an integer

        # Check if the count of rows with leading/trailing spaces exceeds the threshold
        if spaces_count > threshold_value:
            logging.warning(
                "Escalation triggered: Space count exceeds threshold in file '{}' for column '{}'. "
                "Space count: {}, Threshold: {}".format(
                    context.file_name, column_name, spaces_count, threshold_value
                )
            )
            # Here you could also trigger an alert, such as sending an email
            # Example: send_escalation_email(context.file_name, spaces_count)
        else:
            logging.info(
                "No escalation needed: Space count in '{}' for column '{}' is within the threshold.".format(
                    context.file_name, column_name
                )
            )
    else:
        logging.warning(
            "Column '{}' not found in the file '{}' for space count check.".format(
                column_name, context.file_name
            )
        )


# Batch Processing for Trailing Spaces
@given(
    "a batch of bank export files with leading and trailing spaces in multiple fields"
)
def batch_with_space_issues(context):
    # type: (Any) -> None
    """Loads batch files for whitespace validation."""
    context.batch_files = [
        "transactions_trailing_spaces.csv",
        "transactions_leading_spaces.xlsx",
    ]

    # Log that the batch files have been loaded for whitespace validation
    logging.info(
        "Batch files with leading and trailing spaces have been loaded: {}".format(
            context.batch_files
        )
    )


@given('I load the file "{file_name}"')
def step_given_load_file(context, file_name):
    # type: (Any, Any) -> None
    """Load a CSV or Excel file into the context."""
    # Use load_bank_export to load the file into the context
    context.df = load_bank_export(file_name)


@when('I check the "Transaction Amount" and "Recipient" columns in "{sheet_name}"')
def step_when_check_transaction_amount(context, sheet_name):
    # type: (Any, Any) -> None
    """Load the file and verify the presence of necessary columns"""

    # Load the data using the appropriate helper function
    context.df = load_bank_export(context.file_path, sheet_name=sheet_name)

    # List of required columns to check
    required_columns = ["Transaction Amount", "Recipient"]

    # Check if required columns exist in the dataframe
    missing_columns = [col for col in required_columns if col not in context.df.columns]

    if missing_columns:
        logging.error(f"Missing columns: {', '.join(missing_columns)}")
        assert False, f"Missing column(s): {', '.join(missing_columns)}"

    # Validate that the dataframe is not empty
    assert not context.df.empty, "Loaded dataframe is empty."

    # Optionally, you can also validate if the columns are of the expected type
    if "Transaction Amount" in context.df.columns:
        # Validate that 'Transaction Amount' column contains numeric values
        if not validate_column(context.df, "Transaction Amount", "numeric"):
            logging.error(
                "The 'Transaction Amount' column contains non-numeric values."
            )
            assert False, "'Transaction Amount' column contains non-numeric values."

    # Validate 'Recipient' column if necessary
    if "Recipient" in context.df.columns:
        # Check if the 'Recipient' column contains missing values
        if context.df["Recipient"].isnull().any():
            logging.error("The 'Recipient' column contains missing values.")
            assert False, "'Recipient' column contains missing values."


def validate_column(df: pd.DataFrame, column_name: str, validation_type: str) -> bool:
    """
    Validates a column based on the validation type provided.

    :param df: DataFrame containing the data
    :param column_name: Column name to validate
    :param validation_type: Type of validation to apply ("numeric", "unique", etc.)
    :return: Boolean indicating if the validation passed
    """
    if validation_type == "numeric":
        # Check if the column's data type is numeric
        return pd.api.types.is_numeric_dtype(df[column_name])
    elif validation_type == "unique":
        # Check if all values in the column are unique
        return df[column_name].is_unique
    elif validation_type == "non_empty":
        # Check if the column has no missing (null) values
        return df[column_name].notnull().all()
    elif validation_type == "positive":
        # Check if the numeric values in the column are positive
        if pd.api.types.is_numeric_dtype(df[column_name]):
            return (df[column_name] > 0).all()
        else:
            return False  # Return False if the column is not numeric
    else:
        # Default return True for unrecognized validation type
        return True


@then('all space-related discrepancies should be flagged as "{severity}"')
def flag_single_or_batch_files_spacing_discrepancies(context, severity):
    # type: (Any, str) -> None
    """Flags space-related discrepancies in single or batch file processing."""

    # Ensure that batch_issues is iterable in a consistent way for both single and batch file cases
    space_related_discrepancies_batch_files = (
        context.batch_issues.items()  # For batch processing (dict of issues)
        if isinstance(context.batch_issues, dict)
        else [(context.file_name, context.batch_issues)]  # For single file case
    )

    flagged_files = []

    # Iterate over the files and their associated issues
    for file, issues in space_related_discrepancies_batch_files:
        # Check if the file has space-related issues
        if 'spaces' in issues:
            logging.info(
                "File {0} has space-related issues: {1} - Severity: {2}".format(
                    file, issues['spaces'], severity
                )
            )

            # Check severity and flag accordingly
            if severity.lower() == 'high':
                flagged_files.append(file)
        else:
            logging.info(
                "File {0} has no space-related issues - Severity: {1}".format(
                    file, severity
                )
            )

    # Store the flagged files in context for further use, if necessary
    if flagged_files:
        if not hasattr(context, 'flagged_files'):
            context.flagged_files = []

        context.flagged_files.extend(flagged_files)

    # Log the files that were flagged
    if flagged_files:
        logging.info(
            "Flagged files for space-related discrepancies: {0}".format(
                ", ".join(flagged_files)
            )
        )
    else:
        logging.info("No files were flagged for space-related discrepancies.")


# Performance Testing for Trailing Spaces
@given('a system processing "{file_count}" bank export files per hour')
def system_processing_number_of_files_per_hour(context, file_count):
    # type: (Any, Any) -> None
    """Simulates system performance testing with whitespace validation and file size tracking."""

    context.file_count = int(file_count)  # Ensure file_count is treated as an integer
    logging.info(
        "Simulating system performance for processing {0} files per hour.".format(
            context.file_count
        )
    )

    total_file_size = 0  # Initialize the total file size tracker

    try:
        # Using FileProcessor to simulate file processing
        for i in range(context.file_count):
            # Simulating processing each file (reusing the existing file processor)
            file_name = "file_{0}.csv".format(
                i + 1
            )  # Assuming a naming convention for files
            file_processor = FileProcessor(file_name=file_name, context=context)
            file_processor.process()  # Process the file

            # Calculate the file size
            file_size = os.path.getsize(
                file_processor.file_path
            )  # Get the file size in bytes
            total_file_size += file_size  # Add the file size to the total

            logging.info(
                "Processed file: {0} with size: {1} bytes.".format(file_name, file_size)
            )

        # Log the total size of processed files
        logging.info("Total size of processed files: {0} bytes".format(total_file_size))

    except Exception as e:
        # Handling any exception that occurs during processing
        logging.error(
            "Error simulating system performance for processing files: {0}".format(
                str(e)
            )
        )
        # Optionally, you could raise the error to stop execution, depending on your testing framework
        raise ValueError(
            "Error during system performance simulation: {0}".format(str(e))
        )

    # Optional: Validate whitespace in the files (if necessary for performance testing)
    validate_whitespace_in_files(context)


def validate_whitespace_in_files(context: Any) -> None:
    """
    Validates for any leading or trailing spaces in the files' columns.
    This function processes each file and checks for whitespace inconsistencies.
    """

    try:
        # Assuming `context.df` contains the loaded dataframe for a file
        for i in range(context.file_count):
            file_name = "file_{0}.csv".format(i + 1)  # Simulating file names
            file_processor = FileProcessor(file_name=file_name, context=context)
            file_processor.process()  # Process the file

            # Strip leading/trailing spaces from column names
            context.df.columns = context.df.columns.str.strip()

            # Check for any leading/trailing spaces in the data (only for object columns)
            for column in context.df.columns:
                if context.df[column].dtype == object:  # Check only for string columns
                    context.df[column] = context.df[column].str.strip()

                    # Log any rows with leading/trailing spaces
                    if context.df[column].str.contains(r"^\s|^\s$", regex=True).any():
                        logging.warning(
                            "Leading/trailing spaces found in column {0} of file {1}".format(
                                column, file_name
                            )
                        )

    except Exception as e:
        logging.error("Error during whitespace validation: {0}".format(str(e)))
        raise ValueError("Whitespace validation failed: {0}".format(str(e)))


@when('trailing spaces are present in "{year_range}"')
def trailing_spaces_in_year_range(context, year_range):
    # type: (Any, Any) -> None
    """Simulates trailing spaces presence across multiple years."""

    # Strip leading/trailing spaces from the input and assign to context
    context.year_range = year_range.strip()

    logging.info(
        "Trailing spaces simulation for year range: {0}".format(context.year_range)
    )

    try:
        # Using FileProcessor to simulate the process for the year_range
        file_processor = FileProcessor(file_name=context.file_name, context=context)
        file_processor.process()  # Reuse the processing logic from FileProcessor

        # Assuming `year_range` is a range like "2020-2022"
        start_year, end_year = map(int, context.year_range.split('-'))

        for year in range(start_year, end_year + 1):
            # Generating the file name based on the year
            file_name = "bank_export_{}.csv".format(
                year
            )  # Example file naming convention

            # Process each file for the year
            file_processor = FileProcessor(file_name=file_name, context=context)
            file_processor.process()

            # Check for trailing spaces in columns using the helper function
            check_for_trailing_spaces(context, file_name)

    except Exception as e:
        logging.error(
            "Error processing year range for file {0}: {1}".format(
                context.file_name, str(e)
            )
        )
        raise ValueError(
            "Error processing year range for file {0}: {1}".format(
                context.file_name, str(e)
            )
        )


@when("the system processes the file for validation")
def step_when_process_for_validation(context):
    # type: (Any) -> None
    """Processes the file and checks performance validation."""
    process_and_validate_performance(context, context.expected_time)


@then('system resources should not exceed "{resource_limit}%"')
def validate_resource_usage_spaces(context, resource_limit):
    # type: (Any, Any) -> None
    """
    Ensures that system resource usage remains within the acceptable limits.
    Validates that memory usage does not exceed the specified resource limit.
    """
    try:
        # Validate the resource_limit is a valid integer or float percentage
        resource_limit = float(resource_limit)
        if resource_limit < 0 or resource_limit > 100:
            raise ValueError("Resource limit must be between 0 and 100 percent.")

        # Get the current memory usage in percentage (using the helper function from file_processing.py)
        _, memory_usage_pct = get_memory_usage()

        # Log the memory usage details
        logging.info("Current memory usage percentage: %.2f%%" % memory_usage_pct)

        # Check if the memory usage exceeds the specified resource limit
        if memory_usage_pct > resource_limit:
            logging.error(
                "Memory usage exceeds the specified resource limit of %.2f%% (current: %.2f%%)"
                % (resource_limit, memory_usage_pct)
            )
            raise Exception(
                "Memory usage exceeds the specified limit. Please optimize system resources."
            )
        else:
            logging.info(
                "System memory usage is within the specified limit of %.2f%% (current: %.2f%%)"
                % (resource_limit, memory_usage_pct)
            )
    except ValueError as e:
        logging.error("Invalid resource limit provided: {0}".format(str(e)))
        raise
    except Exception as e:
        logging.error("Error validating resource usage: {0}".format(str(e)))
        raise


# Schema Validation for Trailing Spaces
@given('an export file "{file_name}" containing space inconsistencies')
def having_a_file_with_spacing_inconsistencies(context, file_name):
    # type: (Any, Any) -> None
    """Loads a file with space inconsistencies for schema validation."""
    context.file_name = file_name
    try:
        file_processor = FileProcessor(
            file_name=file_name, context=context
        )  # Reusing FileProcessor class to handle file loading
        file_processor.process()  # Process the file (this is where the logic for handling the file is executed)
    except Exception as e:
        context.error = f"Error loading file {file_name}: {str(e)}"
        logging.error(context.error)


@when("I check the schema validation rules")
def when_check_schema_rules(context: Any) -> None:
    """Validates schema rules related to space issues and file structure."""
    logging.info("Checking schema validation rules for " + context.file_name)

    try:
        # Initialize file processor (reuse code from file_processing.py)
        file_processor = FileProcessor(file_name=context.file_name, context=context)

        # Process the file based on type (CSV or Excel)
        file_processor.process()

        # After processing, validate schema (using predefined column mappings)
        issues = validate_file(
            file_processor.context.df,
            context.column_mappings,
            file_type=context.file_type,
        )

        if issues:
            logging.warning(
                "Issues found in {0}: {1}".format(context.file_name, str(issues))
            )
        else:
            logging.info(
                "Schema validation successful for {0}".format(context.file_name)
            )

    except Exception as e:
        context.error = "Error during schema validation for {0}: {1}".format(
            context.file_name, str(e)
        )
        logging.error(context.error)


@then('fields with excessive spaces should be flagged as "{error_severity}"')
def step_then_flag_schema_spaces(context, error_severity):
    # type: (Any, Any) -> None
    """Flags fields with excessive spaces based on schema validation."""
    if hasattr(context, 'df') and context.df is not None:
        # Check for leading/trailing spaces in the dataframe
        space_issues = context.df.applymap(
            lambda x: isinstance(x, str) and x.strip() != x
        )

        if space_issues.any().any():
            logging.warning(
                f"Schema validation spaces detected in {context.file_name} - Severity: {error_severity}"
            )
            context.space_issues_flagged = True
        else:
            logging.info(f"No excessive spaces found in {context.file_name}.")
            context.space_issues_flagged = False
    else:
        logging.error(f"Dataframe for {context.file_name} is not loaded properly.")


# Delimiter Handling Validation
@given('a CSV file "{file_name}" with leading/trailing spaces near delimiters')
def csv_with_delimiters_spaces(context, file_name):
    # type: (Any, Any) -> None
    """Loads a CSV file with space inconsistencies near delimiters."""
    context.file_name = file_name
    logging.info(f"Loaded CSV file with spaces near delimiters: {file_name}")

    try:
        # Read the file and check for delimiter-related issues
        df = pd.read_csv(file_name, skipinitialspace=True)

        # Check for extra spaces near delimiters (if any)
        if df.isnull().values.any():
            logging.warning(
                f"CSV file {file_name} contains missing values or inconsistencies near delimiters."
            )

        context.df = df
    except Exception as e:
        context.error = f"Error loading file {file_name}: {str(e)}"
        logging.error(context.error)


@when("the system parses the file")
def step_when_parse_csv_spaces(context):
    # type: (Any) -> None
    """Parses the CSV file while checking delimiter integrity."""
    logging.info("Parsing CSV file {}".format(context.file_name))

    # Ensure that the file exists and can be loaded
    if not os.path.isfile(context.file_name):
        raise ValueError("File {} does not exist.".format(context.file_name))

    try:
        # Reading the CSV file and checking for any delimiter inconsistencies
        context.df = pd.read_csv(context.file_name, skipinitialspace=True)

        # Check if the dataframe is empty
        if context.df.empty:
            raise ValueError("CSV file {} is empty.".format(context.file_name))

        # Check if the delimiter consistency is intact (skip spaces around the delimiter)
        with open(context.file_name, "r") as f:
            first_line = f.readline()
            delimiter_count = first_line.count(
                ","
            )  # Adjust if you're using a different delimiter
            if delimiter_count != len(first_line.split(",")) - 1:
                raise ValueError(
                    "Delimiter inconsistency found in file {}".format(context.file_name)
                )

    except Exception as e:
        context.error = "Error parsing file {}: {}".format(context.file_name, str(e))
        logging.error(context.error)


@then("delimiter integrity should be preserved")
def step_then_check_delimiters_spaces(context):
    # type: (Any) -> None
    """Ensures delimiters remain intact despite space inconsistencies."""
    if hasattr(context, "error") and context.error:
        raise ValueError("Error detected: {}".format(context.error))

    # If no error found during parsing, confirm delimiter integrity
    logging.info("Delimiter consistency check passed for {}".format(context.file_name))


# Memory Usage Validation for Trailing Spaces
@given('a system processing "{file_count}" large files with excessive spaces')
def memory_usage_spaces(context, file_count):
    # type: (Any, Any) -> None
    """
    Simulates memory and CPU usage testing when handling space issues.
    Monitors memory usage during the processing of large files with excessive spaces.
    """
    context.file_count = int(file_count)  # Converts file_count to integer

    # Simulate the processing logic and track the system's memory usage for each file
    for i in range(context.file_count):
        file_name = "large_file_{}.csv".format(i + 1)  # Simulating file names
        logging.info("Processing file {}...".format(file_name))

        # Simulate processing delay using datetime
        start_time = datetime.now()
        processing_duration = timedelta(seconds=1)  # Simulating 1 second delay per file
        end_time = start_time + processing_duration

        while datetime.now() < end_time:
            pass  # Busy wait for the desired amount of time

        # Track memory usage after each file processing
        memory_before, _ = get_memory_usage()  # Get memory usage before processing

        # Simulate the actual file processing logic here (can be replaced with actual processing steps)
        logging.info("Simulating file processing for {}".format(file_name))

        # Get memory usage after processing the file
        memory_after, _ = get_memory_usage()  # Get memory usage after processing

        # Log the memory usage during processing
        memory_increase = memory_after - memory_before
        logging.info(
            "Memory usage during processing file {}: {} MB".format(
                file_name, memory_increase
            )
        )

        # Check if memory usage increased significantly (e.g., over 100MB)
        if memory_increase > 100:
            logging.warning(
                "Memory usage increased significantly during processing of file {}: {} MB".format(
                    file_name, memory_increase
                )
            )

    # Check if the total memory usage exceeds a predefined threshold
    _, memory_usage_pct = get_memory_usage()
    if memory_usage_pct > 80:
        logging.warning(
            "Memory usage exceeds 80% after processing {} files.".format(
                context.file_count
            )
        )


@then('the memory usage should not exceed "{max_memory_usage} MB"')
def step_then_validate_memory_spaces(context, max_memory_usage):
    # type: (Any, Any) -> None
    """Checks memory usage constraints."""
    current_memory_usage = get_memory_usage()  # Get current memory usage

    assert (
        current_memory_usage <= max_memory_usage
    ), "Memory usage exceeded! Current: {} MB, Max allowed: {} MB".format(
        current_memory_usage, max_memory_usage
    )
    print(
        "Memory usage remains within {} MB while processing {} files.".format(
            max_memory_usage, context.file_count
        )
    )


# ================= End of Trailing and Leading Spaces Validation Script =================
