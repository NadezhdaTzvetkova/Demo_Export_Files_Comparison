import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.worksheet.protection import SheetProtection

# ✅ Ensure save directory exists
SAVE_DIR = "test_data/generated_files"
os.makedirs(SAVE_DIR, exist_ok=True)

# ✅ Function to create an Excel file with sheet protection
def create_protected_excel(file_name, sheet_name, protection_type, password=None):
    """Generate an Excel file with specified sheet protection settings."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Sample headers and data
    headers = ["Transaction ID", "Amount", "Transaction Date", "Account Number"]
    data = [
        [1001, 200.50, "2024-01-01", "ACC123"],
        [1002, 450.00, "2024-01-02", "ACC456"],
        [1003, 300.75, "2024-01-03", "ACC789"],
    ]

    # Write headers
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write data
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Apply protection based on type
    if protection_type == "Read-Only":
        ws.protection.sheet = True  # Entire sheet read-only
    elif protection_type == "Password-Protected" and password:
        ws.protection.set_password(password)
    elif protection_type == "Partially Locked":
        ws.protection.sheet = True
        ws.protection.enable()
        ws.protection.enable_select_locked_cells = True

    # Save the Excel file
    file_path = os.path.join(SAVE_DIR, file_name)
    wb.save(file_path)
    print(f"✅ File {file_name} generated successfully with {protection_type} protection.")

# ✅ Generate protected sheet test files
protected_sheets_files = {
    "transactions_protected.xlsx": ("Transactions Data", "Read-Only"),
    "transactions_password.xlsx": ("Summary Sheet", "Password-Protected"),
    "transactions_locked_rows.xlsx": ("Account Balances", "Partially Locked"),
}

for file_name, (sheet_name, protection_type) in protected_sheets_files.items():
    create_protected_excel(file_name, sheet_name, protection_type, password="secure123" if "password" in file_name else None)

# ✅ Error handling test cases for protected sheets
error_handling_files = {
    "transactions_protected.xlsx": ("Transactions Data", "Read-Only"),
    "summary_protected.xlsx": ("Summary Report", "Password-Protected"),
    "locked_accounts.xlsx": ("Account Balances", "Partially Locked"),
}

for file_name, (sheet_name, protection_type) in error_handling_files.items():
    create_protected_excel(file_name, sheet_name, protection_type, password="secure123" if "password" in file_name else None)

# ✅ Batch processing test cases
batch_processing_files = ["batch_high.xlsx", "batch_medium.xlsx", "batch_low.xlsx"]

for file_name in batch_processing_files:
    create_protected_excel(file_name, "Transactions", "Read-Only")

# ✅ Performance testing (simulate larger datasets)
def create_large_protected_excel(file_name, sheet_name, row_count, protection_type, password=None):
    """Generate a large Excel file to test performance under protected sheet conditions."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Sample headers
    headers = ["Transaction ID", "Amount", "Transaction Date", "Account Number"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Generate large dataset
    for row_idx in range(2, row_count + 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=np.random.uniform(10, 5000))
        ws.cell(row=row_idx, column=3, value=f"2024-01-{(row_idx-1) % 30 + 1:02d}")
        ws.cell(row=row_idx, column=4, value=f"ACC{np.random.randint(100, 999)}")

    # Apply protection based on type
    if protection_type == "Read-Only":
        ws.protection.sheet = True
    elif protection_type == "Password-Protected" and password:
        ws.protection.set_password(password)
    elif protection_type == "Partially Locked":
        ws.protection.sheet = True
        ws.protection.enable()
        ws.protection.enable_select_locked_cells = True

    # Save the file
    file_path = os.path.join(SAVE_DIR, file_name)
    wb.save(file_path)
    print(f"✅ Large file {file_name} generated successfully with {protection_type} protection.")

# ✅ Create performance test files
create_large_protected_excel("performance_2015_2020.xlsx", "Transactions", 100, "Read-Only")
create_large_protected_excel("performance_2021_2023.xlsx", "Summary Sheet", 500, "Password-Protected", password="secure123")

# ✅ Security validation test cases
security_validation_files = {
    "transactions_legacy.xlsx": ("Legacy Data", "Read-Only"),
    "transactions_modified.xlsx": ("Custom Schema", "Password-Protected"),
    "transactions_test.xlsx": ("Test Environment", "Partially Locked"),
}

for file_name, (sheet_name, protection_type) in security_validation_files.items():
    create_protected_excel(file_name, sheet_name, protection_type, password="secure123" if "password" in file_name else None)

print("\n🎉 All protected sheet validation test files generated successfully!")
