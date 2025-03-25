import os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment

# ✅ Ensure save directory exists
SAVE_DIR = "test_data/generated_files"
os.makedirs(SAVE_DIR, exist_ok=True)


# ✅ Function to create an Excel file with merged cells
def create_excel_with_merged_cells(file_name, merge_positions):
    """Generate an Excel file with specified merged cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Sample headers and data
    headers = ["Transaction ID", "Amount", "Transaction Date", "Account Number"]
    data = [
        [1001, 200.50, "2024-01-01", "ACC123"],
        [1002, 450.00, "2024-01-02", "ACC456"],
        [1003, 300.75, "2024-01-03", "ACC789"],
    ]

    # Write headers
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write data
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Apply merged cells based on `merge_positions`
    for merge_range in merge_positions:
        ws.merge_cells(merge_range)
        start_cell = merge_range.split(":")[0]
        ws[start_cell].alignment = Alignment(horizontal="center", vertical="center")

    # Save the Excel file
    file_path = os.path.join(SAVE_DIR, file_name)
    wb.save(file_path)
    print(f"✅ File {file_name} generated successfully with merged cells.")


# ✅ Generate merged cell test files
merged_cell_files = {
    "transactions_merged_cells.xlsx": ["B2:C2"],  # Merge Amount and Transaction Date
    "transactions_partially_merged.xlsx": ["D2:E2"],  # Merge Account Number (partial)
}

for file_name, merge_positions in merged_cell_files.items():
    create_excel_with_merged_cells(file_name, merge_positions)

# ✅ Error handling test cases for merged cells
error_handling_files = {
    "transactions_2020.xlsx": ["A2:A3"],  # Merge Transaction ID (causes issue)
    "transactions_2021.xlsx": ["C2:C3"],  # Merge Date Column
    "transactions_test.xlsx": ["D2:D3"],  # Merge Account Number
}

for file_name, merge_positions in error_handling_files.items():
    create_excel_with_merged_cells(file_name, merge_positions)

# ✅ Batch processing test cases
batch_processing_files = ["batch_high.xlsx", "batch_medium.xlsx", "batch_low.xlsx"]

for file_name in batch_processing_files:
    create_excel_with_merged_cells(file_name, ["B2:C2"])  # Example merge


# ✅ Performance testing (simulate larger datasets)
def create_large_excel(file_name, row_count, merge_positions):
    """Generate a large Excel file to test performance under merged cell conditions."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Sample headers
    headers = ["Transaction ID", "Amount", "Transaction Date", "Account Number"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Generate large dataset
    for row_idx in range(2, row_count + 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=np.random.uniform(10, 5000))
        ws.cell(row=row_idx, column=3, value=f"2024-01-{(row_idx-1) % 30 + 1:02d}")
        ws.cell(row=row_idx, column=4, value=f"ACC{np.random.randint(100, 999)}")

    # Apply merged cells
    for merge_range in merge_positions:
        ws.merge_cells(merge_range)
        start_cell = merge_range.split(":")[0]
        ws[start_cell].alignment = Alignment(horizontal="center", vertical="center")

    # Save the file
    file_path = os.path.join(SAVE_DIR, file_name)
    wb.save(file_path)
    print(f"✅ Large file {file_name} generated successfully.")


# ✅ Create performance test files
create_large_excel("performance_2015_2020.xlsx", 100, ["B2:C2"])
create_large_excel("performance_2021_2023.xlsx", 500, ["D2:E2"])

# ✅ Schema validation test cases
schema_validation_files = {
    "transactions_legacy.xlsx": ["B2:B3"],  # Legacy format with merged cells
    "transactions_modified.xlsx": ["C2:C3"],  # Custom schema with merged date
    "transactions_test.xlsx": ["D2:D3"],  # Test environment schema issue
}

for file_name, merge_positions in schema_validation_files.items():
    create_excel_with_merged_cells(file_name, merge_positions)

print("\n🎉 All merged cell validation test files generated successfully!")
